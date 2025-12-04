#!/usr/bin/env python3
"""
AUTO-CONSOLIDATE MASTER CHARTS

Merges individual Excel master chart files into one consolidated reference workbook.

Purpose:
- User creates multiple master charts: HIV_Master_Chart.xlsx, Antibiotics_Master_Chart.xlsx, etc.
- This script consolidates them into Pharmacology_Master_Reference.xlsx
- Each topic becomes a separate sheet in the reference workbook
- Creates alphabetical index sheet for quick drug/condition lookup

Usage:
    python Auto_Consolidate_Master_Charts.py <master_chart_file> <reference_workbook_path>

Example:
    python Auto_Consolidate_Master_Charts.py "HIV_Master_Chart.xlsx" "Pharmacology_Master_Reference.xlsx"

Features:
- Preserves formatting (fonts, colors, borders) from source
- Creates/updates Index sheet with alphabetical drug list
- Tracks which drugs are in which sheets
- Non-destructive: keeps existing sheets, adds/updates new ones
"""

import sys
import os
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def copy_cell_style(source_cell, target_cell):
    """
    Copy all style properties from source cell to target cell.
    Based on best practices from Stack Overflow.
    """
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def copy_worksheet_to_workbook(source_wb, source_sheet_name, target_wb, target_sheet_name):
    """
    Copy worksheet from source workbook to target workbook.
    Preserves cell values, formatting, column widths, and row heights.

    Args:
        source_wb: Source workbook (already loaded)
        source_sheet_name: Name of sheet to copy from source
        target_wb: Target workbook
        target_sheet_name: Name for sheet in target workbook

    Returns:
        target_sheet: The created/updated sheet in target workbook
    """
    source_sheet = source_wb[source_sheet_name]

    # Remove existing sheet if present (we're updating)
    if target_sheet_name in target_wb.sheetnames:
        del target_wb[target_sheet_name]

    # Create new sheet
    target_sheet = target_wb.create_sheet(target_sheet_name)

    # Copy column widths
    for col_letter in source_sheet.column_dimensions:
        if source_sheet.column_dimensions[col_letter].width:
            target_sheet.column_dimensions[col_letter].width = \
                source_sheet.column_dimensions[col_letter].width

    # Copy row heights
    for row_num in source_sheet.row_dimensions:
        if source_sheet.row_dimensions[row_num].height:
            target_sheet.row_dimensions[row_num].height = \
                source_sheet.row_dimensions[row_num].height

    # Copy cells (values and formatting)
    # Use max_row and max_column for efficiency (best practice)
    max_row = source_sheet.max_row
    max_col = source_sheet.max_column

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            source_cell = source_sheet.cell(row_idx, col_idx)
            target_cell = target_sheet.cell(row_idx, col_idx)

            # Copy value
            target_cell.value = source_cell.value

            # Copy style
            copy_cell_style(source_cell, target_cell)

    # Copy merged cells
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))

    # Freeze panes if present
    if source_sheet.freeze_panes:
        target_sheet.freeze_panes = source_sheet.freeze_panes

    return target_sheet


def extract_drug_names_from_sheet(sheet):
    """
    Extract all drug names from master chart sheet.
    Assumes first row is header, column B contains drug names.

    Args:
        sheet: Worksheet object

    Returns:
        list: Drug names (strings)
    """
    drug_names = []

    # Skip header row (row 1), start from row 2
    for row_idx in range(2, sheet.max_row + 1):
        # Column B typically contains "Drug Name (Brand)" in master charts
        drug_cell = sheet.cell(row_idx, 2)  # Column B
        if drug_cell.value:
            drug_name = str(drug_cell.value).strip()
            if drug_name:  # Skip empty cells
                drug_names.append(drug_name)

    return drug_names


def create_or_update_index_sheet(wb, drug_sheet_mapping):
    """
    Create or update the Index sheet with alphabetical drug listing.

    Args:
        wb: Workbook object
        drug_sheet_mapping: dict {drug_name: sheet_name}
    """
    # Remove existing Index sheet if present
    if "Index" in wb.sheetnames:
        del wb["Index"]

    # Create Index sheet as first sheet
    index_sheet = wb.create_sheet("Index", 0)

    # Set column widths
    index_sheet.column_dimensions['A'].width = 40  # Drug name
    index_sheet.column_dimensions['B'].width = 30  # Sheet location

    # Header row
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    index_sheet.cell(1, 1).value = "Drug / Condition"
    index_sheet.cell(1, 1).font = header_font
    index_sheet.cell(1, 1).fill = header_fill
    index_sheet.cell(1, 1).alignment = header_alignment

    index_sheet.cell(1, 2).value = "Located In"
    index_sheet.cell(1, 2).font = header_font
    index_sheet.cell(1, 2).fill = header_fill
    index_sheet.cell(1, 2).alignment = header_alignment

    index_sheet.row_dimensions[1].height = 30
    index_sheet.freeze_panes = 'A2'

    # Sort drugs alphabetically
    sorted_drugs = sorted(drug_sheet_mapping.items())

    # Add drugs to index
    current_row = 2
    current_letter = None
    row_colors = ['E8F5E9', 'FFFFFF']  # Alternating light green and white
    color_idx = 0

    for drug_name, sheet_name in sorted_drugs:
        # Add letter section headers
        first_letter = drug_name[0].upper()
        if first_letter != current_letter:
            # Add letter header
            index_sheet.merge_cells(start_row=current_row, start_column=1,
                                   end_row=current_row, end_column=2)
            letter_cell = index_sheet.cell(current_row, 1)
            letter_cell.value = f"═══ {first_letter} ═══"
            letter_cell.font = Font(name='Calibri', size=12, bold=True, color='1976D2')
            letter_cell.alignment = Alignment(horizontal='center', vertical='center')
            letter_cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            index_sheet.row_dimensions[current_row].height = 25
            current_row += 1
            current_letter = first_letter
            color_idx = 0  # Reset alternating colors

        # Add drug entry
        bg_color = row_colors[color_idx % 2]

        drug_cell = index_sheet.cell(current_row, 1)
        drug_cell.value = drug_name
        drug_cell.font = Font(name='Calibri', size=11)
        drug_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        drug_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        sheet_cell = index_sheet.cell(current_row, 2)
        sheet_cell.value = sheet_name
        sheet_cell.font = Font(name='Calibri', size=11, color='1565C0')
        sheet_cell.alignment = Alignment(horizontal='left', vertical='top')
        sheet_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        current_row += 1
        color_idx += 1

    return index_sheet


def consolidate_master_chart(master_chart_path, reference_workbook_path):
    """
    Consolidate a master chart file into the reference workbook.

    Args:
        master_chart_path: Path to individual master chart file (e.g., HIV_Master_Chart.xlsx)
        reference_workbook_path: Path to consolidated reference workbook

    Returns:
        dict: Summary of consolidation
    """
    print(f"📁 Reading master chart: {master_chart_path}")

    # Load source workbook
    if not os.path.exists(master_chart_path):
        return {"error": f"Source file not found: {master_chart_path}"}

    source_wb = load_workbook(master_chart_path)

    # Determine topic name from filename
    # Example: "HIV_Master_Chart.xlsx" → "HIV Drugs"
    basename = os.path.basename(master_chart_path)
    topic_name = basename.replace("_Master_Chart.xlsx", "").replace("_", " ")

    # Load or create reference workbook
    if os.path.exists(reference_workbook_path):
        print(f"📂 Loading existing reference: {reference_workbook_path}")
        ref_wb = load_workbook(reference_workbook_path)
    else:
        print(f"📝 Creating new reference: {reference_workbook_path}")
        ref_wb = Workbook()
        # Remove default sheet
        if "Sheet" in ref_wb.sheetnames:
            del ref_wb["Sheet"]

    # Copy master chart sheet to reference workbook
    # Assume source has sheet named "Master Chart" or first sheet
    if "Master Chart" in source_wb.sheetnames:
        source_sheet_name = "Master Chart"
    else:
        source_sheet_name = source_wb.sheetnames[0]

    print(f"📋 Copying sheet '{source_sheet_name}' to reference as '{topic_name}'")
    copy_worksheet_to_workbook(source_wb, source_sheet_name, ref_wb, topic_name)

    # Extract drug names from the copied sheet
    drugs = extract_drug_names_from_sheet(ref_wb[topic_name])
    print(f"✓ Extracted {len(drugs)} drugs from {topic_name}")

    # Build complete drug-sheet mapping from all sheets (except Index)
    drug_sheet_mapping = {}
    for sheet_name in ref_wb.sheetnames:
        if sheet_name == "Index":
            continue
        sheet_drugs = extract_drug_names_from_sheet(ref_wb[sheet_name])
        for drug in sheet_drugs:
            drug_sheet_mapping[drug] = sheet_name

    # Create/update Index sheet
    print(f"📇 Creating/updating Index sheet with {len(drug_sheet_mapping)} total drugs")
    create_or_update_index_sheet(ref_wb, drug_sheet_mapping)

    # Save reference workbook
    print(f"💾 Saving reference: {reference_workbook_path}")
    ref_wb.save(reference_workbook_path)

    return {
        "topic": topic_name,
        "drugs_added": len(drugs),
        "total_drugs_in_reference": len(drug_sheet_mapping),
        "reference_path": reference_workbook_path
    }


def main():
    """Main entry point for script."""
    if len(sys.argv) != 3:
        print("Usage: python Auto_Consolidate_Master_Charts.py <master_chart_file> <reference_workbook>")
        print("")
        print("Example:")
        print('  python Auto_Consolidate_Master_Charts.py "HIV_Master_Chart.xlsx" "Pharmacology_Master_Reference.xlsx"')
        sys.exit(1)

    master_chart_path = sys.argv[1]
    reference_workbook_path = sys.argv[2]

    print("═══════════════════════════════════════")
    print("  AUTO-CONSOLIDATE MASTER CHARTS")
    print("═══════════════════════════════════════")
    print("")

    result = consolidate_master_chart(master_chart_path, reference_workbook_path)

    if "error" in result:
        print(f"❌ ERROR: {result['error']}")
        sys.exit(1)

    print("")
    print("═══════════════════════════════════════")
    print("  CONSOLIDATION COMPLETE")
    print("═══════════════════════════════════════")
    print(f"Topic: {result['topic']}")
    print(f"Drugs added: {result['drugs_added']}")
    print(f"Total drugs in reference: {result['total_drugs_in_reference']}")
    print(f"Reference file: {result['reference_path']}")
    print("")
    print("✅ Master chart consolidated successfully!")


if __name__ == '__main__':
    main()
