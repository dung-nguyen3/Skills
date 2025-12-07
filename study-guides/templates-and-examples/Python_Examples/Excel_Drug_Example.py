#!/usr/bin/env python3
"""
COMPLETE EXCEL DRUG CHART EXAMPLE
Reference implementation for creating comprehensive drug charts

This file shows a COMPLETE working example with all tabs.
Templates should reference this file rather than including all code inline.

Structure:
- Tab 1: Drug Details (comparison tables by drug class)
- Tab 2: Key Comparisons (mechanisms, toxicities, uses)
- Tab 3: Master Chart (all drugs in one table)
- Tab 4 (Optional): FYI Drugs (future reference drugs with page numbers)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# COLOR SCHEME - SOFT PASTELS WITH BLACK TEXT
# =============================================================================

# Main color sets for drug class tables (rotate through these)
# Color Set 0: Ice Blue
ICE_BLUE_HEADER = 'B4C6E7'
ICE_BLUE_MAIN = 'D9E2F3'
ICE_BLUE_ROW_LABEL = 'C5D3ED'

# Color Set 1: Seafoam
SEAFOAM_HEADER = 'A8CCA8'
SEAFOAM_MAIN = 'C8E6C9'
SEAFOAM_ROW_LABEL = 'B8D9B9'

# Color Set 2: Light Orchid
LIGHT_ORCHID_HEADER = 'B8A4D0'
LIGHT_ORCHID_MAIN = 'D1C4E9'
LIGHT_ORCHID_ROW_LABEL = 'C4B4DC'

# Color Set 3: Champagne
CHAMPAGNE_HEADER = 'E0D0B0'
CHAMPAGNE_MAIN = 'F7E7CE'
CHAMPAGNE_ROW_LABEL = 'EBDBBF'

# Color Set 4: Sky Blue
SKY_BLUE_HEADER = '9DC3E6'
SKY_BLUE_MAIN = 'BDD7EE'
SKY_BLUE_ROW_LABEL = 'AECDEA'

# Color Set 5: Pale Azure
PALE_AZURE_HEADER = 'D0E8FF'
PALE_AZURE_MAIN = 'F0F8FF'
PALE_AZURE_ROW_LABEL = 'E0F0FF'

# Color Set 6: Blush Pink
BLUSH_PINK_HEADER = 'E8C4CC'
BLUSH_PINK_MAIN = 'FCE4EC'
BLUSH_PINK_ROW_LABEL = 'F2D4DC'

# Color Set 7: Soft Lilac
SOFT_LILAC_HEADER = 'D0C8DC'
SOFT_LILAC_MAIN = 'EDE7F6'
SOFT_LILAC_ROW_LABEL = 'DED7E9'

# Color Set 8: Soft Tangerine
SOFT_TANGERINE_HEADER = 'E0C8B0'
SOFT_TANGERINE_MAIN = 'FFE8D6'
SOFT_TANGERINE_ROW_LABEL = 'EFD8C3'

# Color Set 9: Powder Blue
POWDER_BLUE_HEADER = 'A0C4E8'
POWDER_BLUE_MAIN = 'BBDEFB'
POWDER_BLUE_ROW_LABEL = 'ADD1F1'

# Special background colors
MAIN_TITLE_COLOR = '4472C4'  # Dark blue for sheet titles

# Color sets array for easy rotation
COLOR_SETS = [
    {'header': ICE_BLUE_HEADER, 'main': ICE_BLUE_MAIN, 'row_label': ICE_BLUE_ROW_LABEL},
    {'header': SEAFOAM_HEADER, 'main': SEAFOAM_MAIN, 'row_label': SEAFOAM_ROW_LABEL},
    {'header': LIGHT_ORCHID_HEADER, 'main': LIGHT_ORCHID_MAIN, 'row_label': LIGHT_ORCHID_ROW_LABEL},
    {'header': CHAMPAGNE_HEADER, 'main': CHAMPAGNE_MAIN, 'row_label': CHAMPAGNE_ROW_LABEL},
    {'header': SKY_BLUE_HEADER, 'main': SKY_BLUE_MAIN, 'row_label': SKY_BLUE_ROW_LABEL},
    {'header': PALE_AZURE_HEADER, 'main': PALE_AZURE_MAIN, 'row_label': PALE_AZURE_ROW_LABEL},
    {'header': BLUSH_PINK_HEADER, 'main': BLUSH_PINK_MAIN, 'row_label': BLUSH_PINK_ROW_LABEL},
    {'header': SOFT_LILAC_HEADER, 'main': SOFT_LILAC_MAIN, 'row_label': SOFT_LILAC_ROW_LABEL},
    {'header': SOFT_TANGERINE_HEADER, 'main': SOFT_TANGERINE_MAIN, 'row_label': SOFT_TANGERINE_ROW_LABEL},
    {'header': POWDER_BLUE_HEADER, 'main': POWDER_BLUE_MAIN, 'row_label': POWDER_BLUE_ROW_LABEL},
]

# Border style - white borders between cells
thin_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='FFFFFF'),
    bottom=Side(style='thin', color='FFFFFF')
)

def get_color_set(index):
    """Get color set by index (rotates through available sets)"""
    return COLOR_SETS[index % len(COLOR_SETS)]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def apply_cell_style(cell, text='', bold=False, font_size=10, bg_color=None,
                     border=True, alignment='left', wrap=True):
    """Apply comprehensive cell styling"""
    cell.value = text
    cell.font = Font(name='Calibri', size=font_size, bold=bold, color='000000')
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical='top',
        wrap_text=wrap
    )

    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    if border:
        cell.border = thin_border

def create_header_row(ws, headers, row=1):
    """Create formatted header row"""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx)
        apply_cell_style(
            cell, text=header, bold=True, font_size=12,
            bg_color=MAIN_TITLE_COLOR, alignment='center'
        )
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')

    ws.row_dimensions[row].height = 25
    ws.freeze_panes = f'A{row+1}'

def set_column_widths(ws, widths):
    """Set column widths from dictionary"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def add_comparison_table(ws, start_row, title, headers, data_rows, colors):
    """
    Add a key comparison table for Key Comparisons tab

    Args:
        ws: worksheet object
        start_row: starting row number
        title: table title (e.g., "MECHANISMS ACROSS CLASSES")
        headers: list of column headers
        data_rows: list of lists, each inner list is a row of data
        colors: dict with 'header' and 'main' color hex codes

    Returns:
        next available row number
    """
    from openpyxl.utils import get_column_letter

    current_row = start_row
    num_cols = len(headers)
    end_col_letter = get_column_letter(num_cols)

    # Title row (merged)
    ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = title
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='000000')
    title_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    current_row += 1

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(name='Calibri', bold=True, size=11, color='000000')
        cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 23
    current_row += 1

    # Data rows - ALL rows use same color (colors['main'])
    for row_data in data_rows:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            # First column is bold
            if col_idx == 1:
                cell.font = Font(name='Calibri', bold=True, size=10, color='000000')
            else:
                cell.font = Font(name='Calibri', size=10, color='000000')
            # ALL cells get pastel background
            cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 43
        current_row += 1

    return current_row

# =============================================================================
# TAB 1: DRUG DETAILS
# =============================================================================

def create_drug_details_tab(wb):
    """
    Tab 1: Drug Details
    - Drug class comparison tables
    - Drugs as columns, properties as rows (10 categories)
    - Merged cells for shared class properties
    - Category order: MOA, Route, Uses, Combination, Adverse Effects, Contraindications, Interactions, PK, Special, [Other]
    - DOC notation: 🟢 DOC for [condition] inline with uses
    - Toxicity marking: ⚠️ within Adverse Effects row
    """
    ws = wb.active
    ws.title = "Drug Details"

    # Column widths
    set_column_widths(ws, {
        'A': 30,  # Property name
        'B': 25,  # Drug 1
        'C': 25,  # Drug 2
        'D': 25,  # Drug 3
        'E': 25,  # Drug 4
        'F': 25,  # Drug 5
    })

    current_row = 1
    class_index = 0

    # =========================================================================
    # EXAMPLE: NRTI Drug Class
    # =========================================================================

    # Class header
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws.cell(current_row, 1)
    colors = get_color_set(class_index)
    apply_cell_style(cell, text="NUCLEOSIDE REVERSE TRANSCRIPTASE INHIBITORS (NRTIs)",
                    bold=True, font_size=18, bg_color=colors['header'], alignment='center')
    cell.font = Font(name='Calibri', size=18, bold=True, color='000000')
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Drug names header
    color = colors['main']
    headers = ['Property', 'Tenofovir (Viread)', 'Emtricitabine (Emtriva)',
               'Lamivudine (Epivir)', 'Abacavir (Ziagen)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx)
        if col_idx == 1:
            apply_cell_style(cell, text=header, bold=True, font_size=14,
                            bg_color=colors['row_label'], alignment='center')
        else:
            apply_cell_style(cell, text=header, bold=True, font_size=16,
                            bg_color=color, alignment='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # Class-wide properties (merged across all drugs)
    class_properties = [
        ('Drug Class', 'NRTI - Nucleoside Reverse Transcriptase Inhibitor'),
        ('Mechanism', 'Competes with natural nucleosides → incorporated into viral DNA → chain termination'),
        ('Route', 'Oral'),
    ]

    for prop_name, prop_value in class_properties:
        # Property name in column A
        cell_a = ws.cell(current_row, 1)
        apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

        # Apply styling to ALL cells BEFORE merging (B-E = columns 2-5)
        for col_idx in range(2, 6):  # Columns 2, 3, 4, 5 = B, C, D, E
            cell = ws.cell(current_row, col_idx)
            apply_cell_style(cell, text=prop_value if col_idx==2 else '',
                           font_size=12, bg_color=color)

        # Now merge columns B-E
        ws.merge_cells(start_row=current_row, start_column=2,
                      end_row=current_row, end_column=5)

        ws.row_dimensions[current_row].height = 48 if prop_name != 'Route' else 32
        current_row += 1

    # Drug-specific properties (individual values per drug)
    # Category order: Uses, Combination Therapy, Adverse Effects, Contraindications, Drug Interactions, Pharmacokinetics, Special Considerations
    # Note: Uses includes 🟢 DOC notation, Adverse Effects includes ⚠️ for toxicity
    drug_properties = [
        ('Uses',
         '🟢 DOC for HIV (first-line)',
         '🟢 DOC for HIV (first-line)',
         'HIV, HBV',
         'HIV (HLA-B*5701 negative)'),
        ('Combination Therapy',
         'Often with emtricitabine (Truvada)',
         'Often with tenofovir (Truvada)',
         'Triple therapy combinations',
         'Triple therapy combinations'),
        ('Adverse Effects',
         '⚠️ Nephrotoxicity, ⚠️ ↓ bone density',
         'Minimal (well tolerated)',
         'Minimal (well tolerated)',
         '⚠️ Hypersensitivity reaction (life-threatening)'),
        ('Contraindications',
         'CrCl <50 mL/min',
         'None',
         'None',
         'HLA-B*5701 positive'),
        ('Drug Interactions',
         'Limited interactions',
         'Limited interactions',
         'Limited interactions',
         'No significant interactions'),
        ('Pharmacokinetics',
         'Renal elimination',
         'Renal elimination',
         'Renal elimination',
         'Hepatic metabolism'),
        ('Special Considerations',
         'Monitor renal function',
         'Safe in pregnancy',
         'Safe in pregnancy, HBV coinfection',
         '❗ Screen HLA-B*5701 BEFORE use'),
    ]

    for prop_row in drug_properties:
        prop_name = prop_row[0]

        # Property name
        cell_a = ws.cell(current_row, 1)
        apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

        # Individual drug values
        for col_idx, value in enumerate(prop_row[1:], start=2):
            cell = ws.cell(current_row, col_idx)
            apply_cell_style(cell, text=value, font_size=12, bg_color=color)

        ws.row_dimensions[current_row].height = 48
        current_row += 1

    # Space before next class
    current_row += 2
    class_index += 1

    # =========================================================================
    # Add more drug classes here following same pattern...
    # =========================================================================

    return ws

# =============================================================================
# TAB 2: KEY COMPARISONS
# =============================================================================

def create_key_comparisons_tab(wb):
    """
    Tab 2: Key Comparisons
    - Side-by-side comparisons across drug classes
    - Each comparison table uses ONE consistent color
    - Only compare 1 category at a time per table
    - Group similar information together (e.g., group drugs with same toxicity near each other)

    Specific comparison tables to include:
    1. Combined Drug Comparison Chart (if applicable)
    2. Adverse Effects Comparison
    3. First-line vs Second-line Drugs
    4. Toxicity Comparison
    5. Mechanisms Comparison
    6. Drug Interactions Comparison
    7. DOC Comparison by indication
    """
    ws = wb.create_sheet("Key Comparisons")

    # Column widths
    set_column_widths(ws, {
        'A': 30,
        'B': 35,
        'C': 35,
        'D': 35,
        'E': 35,
    })

    current_row = 1

    # Comparison 0: Combined Drug Comparison Chart (if applicable) (Powder Blue)
    # Shows combination drugs with their individual components
    colors = get_color_set(9)
    current_row = add_comparison_table(
        ws, current_row,
        'COMBINED DRUG COMPARISON CHART',
        ['Drug 1', 'Drug 2', 'Combined Product', 'Uses'],
        [
            ['Tenofovir', 'Emtricitabine', 'Truvada', '🟢 DOC for HIV (first-line), PrEP'],
            ['Tenofovir alafenamide', 'Emtricitabine', 'Descovy', 'HIV, PrEP (better renal profile)'],
            ['Abacavir', 'Lamivudine', 'Epzicom', 'HIV (HLA-B*5701 negative)'],
        ],
        colors
    )
    current_row += 2

    # Comparison 1: Mechanisms Across Classes (Ice Blue)
    # Note: Only compare 1 category (Mechanism) at a time
    colors = get_color_set(0)
    current_row = add_comparison_table(
        ws, current_row,
        'MECHANISMS ACROSS CLASSES',
        ['Drug Class', 'Mechanism', 'Target', 'Result'],
        [
            ['NRTI', 'Nucleoside analog → chain termination', 'Reverse transcriptase',
             'Viral DNA synthesis stops'],
            ['NNRTI', 'Allosteric binding', 'Reverse transcriptase',
             'Enzyme cannot function'],
            ['Integrase Inhibitor', 'Blocks integration', 'Integrase enzyme',
             'Viral DNA cannot insert into host'],
            ['Protease Inhibitor', 'Competitive inhibition', 'HIV protease',
             'Immature non-infectious virions'],
        ],
        colors
    )
    current_row += 2

    # Comparison 2: Major Toxicities (Seafoam)
    # Note: Group similar toxicities together (e.g., hepatotoxicity entries near each other)
    colors = get_color_set(1)
    current_row = add_comparison_table(
        ws, current_row,
        'MAJOR TOXICITIES COMPARISON',
        ['Drug Class', 'Key Toxicity', 'Monitoring', 'Management'],
        [
            # Grouped: Hepatotoxicity and severe reactions
            ['NRTI', '⚠️ Lactic acidosis, ⚠️ hepatotoxicity, lipodystrophy',
             'Lactate levels, LFTs, lipid panel', 'Discontinue if severe'],
            ['NNRTI', '⚠️ Rash (Stevens-Johnson), ⚠️ hepatotoxicity',
             'Skin examination, LFTs', 'Stop immediately if severe rash'],
            # Grouped: Metabolic effects
            ['Protease Inhibitor', 'GI upset, hyperglycemia, hyperlipidemia',
             'Blood glucose, lipid panel', 'Manage metabolic effects'],
            # Well-tolerated group
            ['Integrase Inhibitor', 'Generally well tolerated, weight gain',
             'Weight monitoring', 'Lifestyle modifications'],
        ],
        colors
    )
    current_row += 2

    # Comparison 3: Clinical Uses (Light Orchid)
    colors = get_color_set(2)
    current_row = add_comparison_table(
        ws, current_row,
        'CLINICAL USES COMPARISON',
        ['Drug Class', 'Primary Use', 'Special Indications', 'Notes'],
        [
            ['NRTI', '🟢 First-line HIV therapy', 'PrEP (Truvada, Descovy)',
             'Backbone of most regimens'],
            ['NNRTI', 'HIV treatment (combination)', 'Resource-limited settings',
             'Cannot be used as monotherapy'],
            ['Protease Inhibitor', 'HIV treatment (salvage)', 'Treatment-experienced patients',
             'Often boosted with ritonavir'],
            ['Integrase Inhibitor', '🟢 First-line HIV therapy', 'Preferred in many guidelines',
             'High barrier to resistance'],
        ],
        colors
    )
    current_row += 2

    # Comparison 4: Drug Interactions (Champagne)
    colors = get_color_set(3)
    current_row = add_comparison_table(
        ws, current_row,
        'MAJOR DRUG INTERACTIONS',
        ['Drug Class', 'Key Interactions', 'Mechanism', 'Clinical Impact'],
        [
            ['NRTI', 'Limited interactions', 'Renally eliminated',
             'Generally safe with other drugs'],
            ['NNRTI', 'CYP450 inducers/inhibitors', 'Hepatic metabolism',
             '⚠️ Many significant interactions'],
            ['Protease Inhibitor', 'CYP3A4 substrates', 'CYP3A4 inhibition',
             '⚠️ Extensive interaction profile'],
            ['Integrase Inhibitor', 'Cation-containing antacids', 'Chelation',
             'Separate administration by 2-4 hours'],
        ],
        colors
    )
    current_row += 2

    # Comparison 5: Clinical Decision-Making (Sky Blue)
    colors = get_color_set(4)
    current_row = add_comparison_table(
        ws, current_row,
        'CLINICAL DECISION-MAKING GUIDE',
        ['Scenario', 'Preferred Choice', 'Rationale'],
        [
            ['Treatment-naïve patient', '🟢 Integrase inhibitor + 2 NRTIs',
             'Best efficacy, tolerability, high barrier to resistance'],
            ['Pregnancy', 'Integrase inhibitor-based regimen',
             'Safest profile, extensive pregnancy data'],
            ['Renal impairment', 'Avoid tenofovir TDF, use TAF',
             'Reduced nephrotoxicity with TAF formulation'],
            ['Treatment-experienced with resistance', 'Boosted PI + newer agents',
             'Higher barrier to resistance, salvage therapy'],
        ],
        colors
    )

    return ws

# =============================================================================
# TAB 3: MASTER CHART
# =============================================================================

def create_master_chart_tab(wb):
    """
    Tab 3: Master Chart
    - ALL drugs in one comprehensive table
    - Rows = individual drugs
    - Frozen header for scrolling
    - Column order: Drug Class, Drug Name, Mechanism, Route, Uses, Combination, Adverse Effects, Contraindications, Interactions, PK, Special
    - Uses includes 🟢 DOC notation, Adverse Effects includes ⚠️ for toxicity
    """
    ws = wb.create_sheet("Master Chart")

    # Column widths
    set_column_widths(ws, {
        'A': 20,  # Drug Class
        'B': 25,  # Drug Name (Brand)
        'C': 35,  # Mechanism of Action
        'D': 15,  # Route
        'E': 35,  # Uses
        'F': 30,  # Combination Therapy
        'G': 35,  # Adverse Effects
        'H': 30,  # Contraindications
        'I': 30,  # Drug Interactions
        'J': 25,  # Pharmacokinetics
        'K': 30,  # Special Considerations
    })

    # Headers (11 columns)
    headers = ['Drug Class', 'Drug Name (Brand)', 'Mechanism of Action', 'Route',
               'Uses', 'Combination Therapy', 'Adverse Effects', 'Contraindications',
               'Drug Interactions', 'Pharmacokinetics', 'Special Considerations']
    create_header_row(ws, headers, 1)

    # Master chart data (ALL drugs from ALL classes)
    # Column order: Drug Class, Drug Name, Mechanism, Route, Uses, Combination, Adverse Effects, Contraindications, Interactions, PK, Special
    master_data = [
        ('NRTI', 'Tenofovir (Viread)', 'Nucleoside analog → chain termination', 'Oral',
         '🟢 DOC for HIV (first-line)', 'Often with emtricitabine (Truvada)',
         '⚠️ Nephrotoxicity, ⚠️ ↓ bone density', 'CrCl <50 mL/min',
         'Limited interactions', 'Renal elimination', 'Monitor renal function'),
        ('NRTI', 'Emtricitabine (Emtriva)', 'Nucleoside analog → chain termination', 'Oral',
         '🟢 DOC for HIV (first-line)', 'Often with tenofovir (Truvada)',
         'Minimal (well tolerated)', 'None',
         'Limited interactions', 'Renal elimination', 'Safe in pregnancy'),
        ('NRTI', 'Lamivudine (Epivir)', 'Nucleoside analog → chain termination', 'Oral',
         'HIV, HBV', 'Triple therapy combinations',
         'Minimal (well tolerated)', 'None',
         'Limited interactions', 'Renal elimination', 'HBV coinfection'),
        ('NRTI', 'Abacavir (Ziagen)', 'Nucleoside analog → chain termination', 'Oral',
         'HIV (HLA-B*5701 negative)', 'Triple therapy combinations',
         '⚠️ Hypersensitivity reaction (life-threatening)', 'HLA-B*5701 positive',
         'No significant interactions', 'Hepatic metabolism', '❗ Screen HLA-B*5701 BEFORE use'),
        # Add all other drugs here...
    ]

    current_row = 2
    class_index = -1  # Start at -1 so first class becomes 0
    prev_class = None

    for row_data in master_data:
        drug_class = row_data[0]

        # Change color when class changes
        if drug_class != prev_class:
            class_index += 1
            prev_class = drug_class

        # Get color set and use 'main' shade for data cells
        colors = get_color_set(class_index)
        color = colors['main']

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            # First column (drug class) is bold
            if col_idx == 1:
                cell.font = Font(name='Calibri', bold=True, size=10, color='000000')
            else:
                cell.font = Font(name='Calibri', size=10, color='000000')
            # All cells get drug class color
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
            cell.value = value

        ws.row_dimensions[current_row].height = 60
        current_row += 1

    return ws

# =============================================================================
# TAB 4 (OPTIONAL): FYI DRUGS
# =============================================================================

def create_fyi_tab(wb, fyi_drugs):
    """
    Tab 4 (Optional): FYI Drugs
    - Simple 2-column table: Drug Name | Page Number
    - Only include if source has FYI/Future Reference section
    - Minimal styling

    Args:
        wb: Workbook object
        fyi_drugs: List of tuples (drug_name, page_number)
                   Example: [("Cevimeline (Evoxac)", "25"), ("Rivastigmine (Exelon)", "25")]
    """
    ws = wb.create_sheet("FYI Drugs")

    # Column widths
    set_column_widths(ws, {
        'A': 35,  # Drug Name
        'B': 15,  # Page Number
    })

    # Headers
    headers = ['Drug Name', 'Page Number']
    create_header_row(ws, headers, 1)

    # Data rows
    current_row = 2
    for drug_name, page_number in fyi_drugs:
        # Drug name
        cell_a = ws.cell(current_row, 1)
        cell_a.value = drug_name
        cell_a.font = Font(name='Calibri', size=10, color='000000')
        cell_a.alignment = Alignment(wrap_text=True, vertical='top')
        cell_a.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell_a.border = thin_border

        # Page number
        cell_b = ws.cell(current_row, 2)
        cell_b.value = page_number
        cell_b.font = Font(name='Calibri', size=10, color='000000')
        cell_b.alignment = Alignment(horizontal='center', vertical='top')
        cell_b.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell_b.border = thin_border

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    return ws

# =============================================================================
# MAIN FUNCTION
# =============================================================================

def create_drug_chart(output_path, fyi_drugs=None):
    """
    Create complete drug chart with 3 main tabs (+ optional FYI tab)

    Args:
        output_path: Path to save the Excel file
        fyi_drugs: Optional list of tuples (drug_name, page_number) for FYI drugs
                   Example: [("Cevimeline (Evoxac)", "25"), ("Rivastigmine (Exelon)", "25")]

    Tabs:
        - Tab 1: Drug Details (comparison tables by drug class)
        - Tab 2: Key Comparisons (specific comparison tables across classes)
        - Tab 3: Master Chart (all drugs in one comprehensive table)
        - Tab 4 (Optional): FYI Drugs (if fyi_drugs list provided)
    """
    # Create workbook
    wb = Workbook()

    # Create all tabs
    create_drug_details_tab(wb)
    create_key_comparisons_tab(wb)
    create_master_chart_tab(wb)

    # Create FYI tab if FYI drugs provided
    if fyi_drugs:
        create_fyi_tab(wb, fyi_drugs)

    # Save
    wb.save(output_path)
    print(f"✅ Drug chart created: {output_path}")

if __name__ == '__main__':
    # Example usage without FYI drugs
    create_drug_chart('HIV_Drugs_Chart.xlsx')

    # Example usage with FYI drugs
    # fyi_drugs_example = [
    #     ("Cevimeline (Evoxac)", "25"),
    #     ("Rivastigmine (Exelon)", "25"),
    #     ("Galantamine (Razadyne)", "25"),
    # ]
    # create_drug_chart('HIV_Drugs_Chart_with_FYI.xlsx', fyi_drugs=fyi_drugs_example)
