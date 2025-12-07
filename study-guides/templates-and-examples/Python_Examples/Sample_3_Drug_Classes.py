#!/usr/bin/env python3
"""
SAMPLE: 3 Drug Classes - Complete Implementation
Demonstrates the new 3-tab drug chart structure with:
- 3 different HIV drug classes (NRTIs, NNRTIs, Protease Inhibitors)
- 10-category standardized order
- Emoji notation (🟢 DOC, ⚠️ toxicity)
- No analogies or memory tricks
- Tab 2 specific comparison tables
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# COLOR SCHEME - SOFT PASTELS WITH BLACK TEXT
# =============================================================================

# Color Set 0: Ice Blue
ICE_BLUE_HEADER = 'B4C6E7'
ICE_BLUE_MAIN = 'D9E2F3'
ICE_BLUE_ROW_LABEL = 'C5D3ED'

# Color Set 1: Seafoam
SEAFOAM_HEADER = 'A8CCA8'
SEAFOAM_MAIN = 'C8E6C9'
SEAFOAM_ROW_LABEL = 'B8D9B9'

# Color Set 2: Light Orchid
LIGHT_ORCHID_HEADER = 'B8A4D0'
LIGHT_ORCHID_MAIN = 'D1C4E9'
LIGHT_ORCHID_ROW_LABEL = 'C4B4DC'

# Special colors
MAIN_TITLE_COLOR = '4472C4'

# Color sets array
COLOR_SETS = [
    {'header': ICE_BLUE_HEADER, 'main': ICE_BLUE_MAIN, 'row_label': ICE_BLUE_ROW_LABEL},
    {'header': SEAFOAM_HEADER, 'main': SEAFOAM_MAIN, 'row_label': SEAFOAM_ROW_LABEL},
    {'header': LIGHT_ORCHID_HEADER, 'main': LIGHT_ORCHID_MAIN, 'row_label': LIGHT_ORCHID_ROW_LABEL},
]

# Border style
thin_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='FFFFFF'),
    bottom=Side(style='thin', color='FFFFFF')
)

def get_color_set(index):
    return COLOR_SETS[index % len(COLOR_SETS)]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def apply_cell_style(cell, text='', bold=False, font_size=10, bg_color=None,
                     border=True, alignment='left', wrap=True):
    cell.value = text
    cell.font = Font(name='Calibri', size=font_size, bold=bold, color='000000')
    cell.alignment = Alignment(horizontal=alignment, vertical='top', wrap_text=wrap)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    if border:
        cell.border = thin_border

def create_header_row(ws, headers, row=1):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx)
        apply_cell_style(cell, text=header, bold=True, font_size=12,
                        bg_color=MAIN_TITLE_COLOR, alignment='center')
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    ws.row_dimensions[row].height = 25
    ws.freeze_panes = f'A{row+1}'

def set_column_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def add_comparison_table(ws, start_row, title, headers, data_rows, colors):
    current_row = start_row
    num_cols = len(headers)
    end_col_letter = get_column_letter(num_cols)

    # Title row
    ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = title
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='000000')
    title_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    current_row += 1

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(name='Calibri', bold=True, size=11, color='000000')
        cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 23
    current_row += 1

    # Data rows
    for row_data in data_rows:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            if col_idx == 1:
                cell.font = Font(name='Calibri', bold=True, size=10, color='000000')
            else:
                cell.font = Font(name='Calibri', size=10, color='000000')
            cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 43
        current_row += 1

    return current_row

# =============================================================================
# TAB 1: DRUG DETAILS
# =============================================================================

def create_drug_details_tab(wb):
    ws = wb.active
    ws.title = "Drug Details"

    set_column_widths(ws, {
        'A': 30, 'B': 25, 'C': 25, 'D': 25, 'E': 25, 'F': 25,
    })

    current_row = 1

    # =========================================================================
    # DRUG CLASS 1: NRTIs (Ice Blue)
    # =========================================================================
    colors = get_color_set(0)

    # Class header
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws.cell(current_row, 1)
    apply_cell_style(cell, text="NUCLEOSIDE REVERSE TRANSCRIPTASE INHIBITORS (NRTIs)",
                    bold=True, font_size=18, bg_color=colors['header'], alignment='center')
    cell.font = Font(name='Calibri', size=18, bold=True, color='000000')
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # Drug names header
    color = colors['main']
    headers = ['Property', 'Tenofovir (Viread)', 'Emtricitabine (Emtriva)',
               'Lamivudine (Epivir)', 'Abacavir (Ziagen)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx)
        if col_idx == 1:
            apply_cell_style(cell, text=header, bold=True, font_size=14,
                            bg_color=colors['row_label'], alignment='center')
        else:
            apply_cell_style(cell, text=header, bold=True, font_size=16,
                            bg_color=color, alignment='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # Class-wide properties
    class_properties = [
        ('Mechanism of Action', 'Competes with natural nucleosides → incorporated into viral DNA → chain termination'),
        ('Route', 'Oral'),
    ]

    for prop_name, prop_value in class_properties:
        cell_a = ws.cell(current_row, 1)
        apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

        for col_idx in range(2, 6):
            cell = ws.cell(current_row, col_idx)
            apply_cell_style(cell, text=prop_value if col_idx==2 else '', font_size=12, bg_color=color)

        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        ws.row_dimensions[current_row].height = 48 if prop_name != 'Route' else 32
        current_row += 1

    # Drug-specific properties
    drug_properties = [
        ('Uses',
         '🟢 DOC for HIV (first-line), PrEP',
         '🟢 DOC for HIV (first-line), PrEP',
         'HIV, HBV',
         'HIV (HLA-B*5701 negative)'),
        ('Combination Therapy',
         'Often with emtricitabine (Truvada)',
         'Often with tenofovir (Truvada)',
         'Triple therapy combinations',
         'Triple therapy combinations'),
        ('Adverse Effects',
         '⚠️ Nephrotoxicity, ⚠️ ↓ bone density, GI upset',
         'Minimal (well tolerated)',
         'Minimal (well tolerated)',
         '⚠️ Hypersensitivity reaction (life-threatening), fever, rash'),
        ('Contraindications',
         'CrCl <50 mL/min',
         'None',
         'None',
         'HLA-B*5701 positive'),
        ('Drug Interactions',
         'Limited interactions (renal elimination)',
         'Limited interactions (renal elimination)',
         'Limited interactions (renal elimination)',
         'Alcohol (↑ levels)'),
        ('Pharmacokinetics',
         'Renal elimination, 60% bioavailability',
         'Renal elimination, 93% bioavailability',
         'Renal elimination, 86% bioavailability',
         'Hepatic metabolism (alcohol dehydrogenase)'),
        ('Special Considerations',
         'Monitor renal function, bone density',
         'Safe in pregnancy',
         'Safe in pregnancy, effective for HBV coinfection',
         '❗ Screen HLA-B*5701 BEFORE initiating therapy'),
    ]

    for prop_row in drug_properties:
        prop_name = prop_row[0]
        cell_a = ws.cell(current_row, 1)
        apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

        for col_idx, value in enumerate(prop_row[1:], start=2):
            cell = ws.cell(current_row, col_idx)
            apply_cell_style(cell, text=value, font_size=12, bg_color=color)

        ws.row_dimensions[current_row].height = 48
        current_row += 1

    current_row += 2

    # =========================================================================
    # DRUG CLASS 2: NNRTIs (Seafoam)
    # =========================================================================
    colors = get_color_set(1)

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws.cell(current_row, 1)
    apply_cell_style(cell, text="NON-NUCLEOSIDE REVERSE TRANSCRIPTASE INHIBITORS (NNRTIs)",
                    bold=True, font_size=18, bg_color=colors['header'], alignment='center')
    cell.font = Font(name='Calibri', size=18, bold=True, color='000000')
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    color = colors['main']
    headers = ['Property', 'Efavirenz (Sustiva)', 'Nevirapine (Viramune)', 'Etravirine (Intelence)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx)
        if col_idx == 1:
            apply_cell_style(cell, text=header, bold=True, font_size=14,
                            bg_color=colors['row_label'], alignment='center')
        else:
            apply_cell_style(cell, text=header, bold=True, font_size=16,
                            bg_color=color, alignment='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    class_properties = [
        ('Mechanism of Action', 'Binds to allosteric site on reverse transcriptase → conformational change → enzyme dysfunction'),
        ('Route', 'Oral'),
    ]

    for prop_name, prop_value in class_properties:
        cell_a = ws.cell(current_row, 1)
        apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

        for col_idx in range(2, 5):
            cell = ws.cell(current_row, col_idx)
            apply_cell_style(cell, text=prop_value if col_idx==2 else '', font_size=12, bg_color=color)

        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        ws.row_dimensions[current_row].height = 48 if prop_name != 'Route' else 32
        current_row += 1

    drug_properties = [
        ('Uses',
         'HIV (combination therapy)',
         'HIV (combination therapy)',
         'HIV (treatment-experienced patients)'),
        ('Combination Therapy',
         'Always combined with NRTIs',
         'Always combined with NRTIs',
         'Used with boosted PIs'),
        ('Adverse Effects',
         '⚠️ CNS effects (vivid dreams, dizziness, confusion), rash, ⚠️ teratogenic',
         '⚠️ Rash (including Stevens-Johnson syndrome), ⚠️ hepatotoxicity',
         'Rash (mild), nausea'),
        ('Contraindications',
         'Pregnancy (1st trimester), severe hepatic impairment',
         'Severe hepatic impairment',
         'None (used in treatment-experienced)'),
        ('Drug Interactions',
         '⚠️ CYP3A4 inducer (↓ levels of many drugs)',
         '⚠️ CYP3A4 inducer (↓ levels of many drugs)',
         'CYP3A4, CYP2C9, CYP2C19 substrate'),
        ('Pharmacokinetics',
         'Hepatic metabolism (CYP3A4), long half-life',
         'Hepatic metabolism (CYP3A4)',
         'Hepatic metabolism, take with food'),
        ('Special Considerations',
         'Take at bedtime (↓ CNS effects), avoid in pregnancy',
         'Monitor LFTs closely, progressive dose escalation',
         'Resistance mutations common with older NNRTIs'),
    ]

    for prop_row in drug_properties:
        prop_name = prop_row[0]
        cell_a = ws.cell(current_row, 1)
        apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

        for col_idx, value in enumerate(prop_row[1:], start=2):
            cell = ws.cell(current_row, col_idx)
            apply_cell_style(cell, text=value, font_size=12, bg_color=color)

        ws.row_dimensions[current_row].height = 48
        current_row += 1

    current_row += 2

    # =========================================================================
    # DRUG CLASS 3: Protease Inhibitors (Light Orchid)
    # =========================================================================
    colors = get_color_set(2)

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws.cell(current_row, 1)
    apply_cell_style(cell, text="PROTEASE INHIBITORS (PIs)",
                    bold=True, font_size=18, bg_color=colors['header'], alignment='center')
    cell.font = Font(name='Calibri', size=18, bold=True, color='000000')
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    color = colors['main']
    headers = ['Property', 'Darunavir (Prezista)', 'Atazanavir (Reyataz)', 'Lopinavir (Kaletra)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx)
        if col_idx == 1:
            apply_cell_style(cell, text=header, bold=True, font_size=14,
                            bg_color=colors['row_label'], alignment='center')
        else:
            apply_cell_style(cell, text=header, bold=True, font_size=16,
                            bg_color=color, alignment='center')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    class_properties = [
        ('Mechanism of Action', 'Competitive inhibition of HIV protease → immature, non-infectious virions'),
        ('Route', 'Oral'),
    ]

    for prop_name, prop_value in class_properties:
        cell_a = ws.cell(current_row, 1)
        apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

        for col_idx in range(2, 5):
            cell = ws.cell(current_row, col_idx)
            apply_cell_style(cell, text=prop_value if col_idx==2 else '', font_size=12, bg_color=color)

        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        ws.row_dimensions[current_row].height = 48 if prop_name != 'Route' else 32
        current_row += 1

    drug_properties = [
        ('Uses',
         '🟢 DOC for treatment-experienced patients',
         'HIV (treatment-naïve)',
         'HIV (salvage therapy)'),
        ('Combination Therapy',
         'Always boosted with ritonavir or cobicistat',
         'Boosted with ritonavir or cobicistat',
         'Co-formulated with ritonavir'),
        ('Adverse Effects',
         'GI upset, ⚠️ rash (contains sulfa moiety), hyperlipidemia',
         '⚠️ Hyperbilirubinemia (indirect), ⚠️ PR interval prolongation, nephrolithiasis',
         'GI upset, ⚠️ pancreatitis, hyperlipidemia, hyperglycemia'),
        ('Contraindications',
         'Sulfa allergy',
         'Severe hepatic impairment, 2nd/3rd degree AV block',
         'Hepatic impairment'),
        ('Drug Interactions',
         '⚠️ CYP3A4 substrate and inhibitor (extensive interactions)',
         '⚠️ CYP3A4 inhibitor, ↑ unconjugated bilirubin',
         '⚠️ CYP3A4 inhibitor (extensive interactions)'),
        ('Pharmacokinetics',
         'Hepatic metabolism (CYP3A4), take with food',
         'Hepatic metabolism, take with food',
         'Hepatic metabolism, bioavailability ↑ with food'),
        ('Special Considerations',
         'High genetic barrier to resistance',
         'Monitor ECG, may cause jaundice (benign)',
         'Monitor lipids, glucose, amylase/lipase'),
    ]

    for prop_row in drug_properties:
        prop_name = prop_row[0]
        cell_a = ws.cell(current_row, 1)
        apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

        for col_idx, value in enumerate(prop_row[1:], start=2):
            cell = ws.cell(current_row, col_idx)
            apply_cell_style(cell, text=value, font_size=12, bg_color=color)

        ws.row_dimensions[current_row].height = 48
        current_row += 1

    return ws

# =============================================================================
# TAB 2: KEY COMPARISONS
# =============================================================================

def create_key_comparisons_tab(wb):
    ws = wb.create_sheet("Key Comparisons")

    set_column_widths(ws, {
        'A': 30, 'B': 35, 'C': 35, 'D': 35, 'E': 35,
    })

    current_row = 1

    # Comparison 1: Mechanisms Across Classes
    colors = get_color_set(0)
    current_row = add_comparison_table(
        ws, current_row,
        'MECHANISMS OF ACTION COMPARISON',
        ['Drug Class', 'Mechanism', 'Target', 'Result'],
        [
            ['NRTI', 'Nucleoside analog → chain termination', 'Reverse transcriptase',
             'Viral DNA synthesis stops'],
            ['NNRTI', 'Allosteric binding → conformational change', 'Reverse transcriptase',
             'Enzyme cannot function'],
            ['Protease Inhibitor', 'Competitive inhibition', 'HIV protease',
             'Immature, non-infectious virions'],
        ],
        colors
    )
    current_row += 2

    # Comparison 2: Major Toxicities (grouped by type)
    colors = get_color_set(1)
    current_row = add_comparison_table(
        ws, current_row,
        'MAJOR TOXICITIES COMPARISON',
        ['Drug Class', 'Key Toxicities', 'Monitoring', 'Management'],
        [
            # Grouped: Hepatotoxicity
            ['NRTI', '⚠️ Lactic acidosis, ⚠️ hepatotoxicity (rare), ⚠️ nephrotoxicity (tenofovir)',
             'Lactate, LFTs, SCr', 'Discontinue if severe'],
            ['NNRTI', '⚠️ Hepatotoxicity, ⚠️ rash (Stevens-Johnson possible)',
             'LFTs, skin examination', 'Stop immediately if severe'],
            # Grouped: Metabolic
            ['Protease Inhibitor', 'Hyperlipidemia, hyperglycemia, lipodystrophy',
             'Lipid panel, glucose', 'Statin therapy, glucose control'],
        ],
        colors
    )
    current_row += 2

    # Comparison 3: Drug of Choice by Indication
    colors = get_color_set(2)
    current_row = add_comparison_table(
        ws, current_row,
        'DRUG OF CHOICE (DOC) BY INDICATION',
        ['Clinical Scenario', 'First Choice', 'Alternative', 'Avoid'],
        [
            ['Treatment-naïve HIV', '🟢 Tenofovir + Emtricitabine (NRTI backbone)',
             'Abacavir + Lamivudine', 'Monotherapy (never use)'],
            ['Treatment-experienced with resistance', '🟢 Darunavir (boosted PI)',
             'Dolutegravir (integrase inhibitor)', 'Unboosted PIs'],
            ['HIV + HBV coinfection', '🟢 Tenofovir + Emtricitabine or Lamivudine',
             'Entecavir (for HBV alone)', 'NNRTIs (no HBV activity)'],
            ['Pregnancy', 'Integrase inhibitor-based regimen',
             'Boosted PI-based regimen', 'Efavirenz (teratogenic)'],
        ],
        colors
    )
    current_row += 2

    # Comparison 4: First-line vs Second-line
    colors = get_color_set(0)
    current_row = add_comparison_table(
        ws, current_row,
        'FIRST-LINE VS SECOND-LINE THERAPY',
        ['Treatment Line', 'Preferred Regimen', 'Rationale', 'When to Use'],
        [
            ['First-line (naïve)', '2 NRTIs + Integrase Inhibitor',
             'Best efficacy, tolerability, high barrier to resistance',
             'Treatment-naïve patients'],
            ['Second-line (experienced)', 'Boosted PI + 2 NRTIs',
             'Higher resistance barrier, salvage therapy',
             'Treatment failure, resistance'],
            ['Alternative first-line', '2 NRTIs + NNRTI',
             'Cost-effective, once-daily dosing',
             'Resource-limited settings'],
        ],
        colors
    )
    current_row += 2

    # Comparison 5: Drug Interactions
    colors = get_color_set(1)
    current_row = add_comparison_table(
        ws, current_row,
        'DRUG INTERACTIONS BY CLASS',
        ['Drug Class', 'Interaction Mechanism', 'Clinical Significance', 'Management'],
        [
            ['NRTI', 'Limited (renal elimination)',
             'Few significant interactions',
             'Dose adjust for renal impairment'],
            ['NNRTI', '⚠️ CYP3A4 induction',
             'Decreases levels of many medications',
             'Check drug interactions, adjust doses'],
            ['Protease Inhibitor', '⚠️ CYP3A4 inhibition',
             'Increases levels of many medications',
             'Extensive interaction profile, frequent monitoring'],
        ],
        colors
    )

    return ws

# =============================================================================
# TAB 3: MASTER CHART
# =============================================================================

def create_master_chart_tab(wb):
    ws = wb.create_sheet("Master Chart")

    set_column_widths(ws, {
        'A': 20, 'B': 25, 'C': 35, 'D': 15, 'E': 35, 'F': 30,
        'G': 35, 'H': 30, 'I': 30, 'J': 25, 'K': 30,
    })

    headers = ['Drug Class', 'Drug Name (Brand)', 'Mechanism of Action', 'Route',
               'Uses', 'Combination Therapy', 'Adverse Effects', 'Contraindications',
               'Drug Interactions', 'Pharmacokinetics', 'Special Considerations']
    create_header_row(ws, headers, 1)

    # All drugs from all classes
    master_data = [
        # NRTIs (Ice Blue - index 0)
        ('NRTI', 'Tenofovir (Viread)', 'Nucleoside analog → chain termination', 'Oral',
         '🟢 DOC for HIV (first-line), PrEP', 'With emtricitabine (Truvada)',
         '⚠️ Nephrotoxicity, ⚠️ ↓ bone density, GI upset', 'CrCl <50 mL/min',
         'Limited interactions', 'Renal elimination', 'Monitor renal function, bone density'),
        ('NRTI', 'Emtricitabine (Emtriva)', 'Nucleoside analog → chain termination', 'Oral',
         '🟢 DOC for HIV (first-line), PrEP', 'With tenofovir (Truvada)',
         'Minimal (well tolerated)', 'None',
         'Limited interactions', 'Renal elimination', 'Safe in pregnancy'),
        ('NRTI', 'Lamivudine (Epivir)', 'Nucleoside analog → chain termination', 'Oral',
         'HIV, HBV', 'Triple therapy combinations',
         'Minimal (well tolerated)', 'None',
         'Limited interactions', 'Renal elimination', 'Effective for HBV coinfection'),
        ('NRTI', 'Abacavir (Ziagen)', 'Nucleoside analog → chain termination', 'Oral',
         'HIV (HLA-B*5701 negative)', 'Triple therapy combinations',
         '⚠️ Hypersensitivity reaction (life-threatening)', 'HLA-B*5701 positive',
         'Alcohol (↑ levels)', 'Hepatic metabolism', '❗ Screen HLA-B*5701 BEFORE use'),

        # NNRTIs (Seafoam - index 1)
        ('NNRTI', 'Efavirenz (Sustiva)', 'Allosteric RT binding → dysfunction', 'Oral',
         'HIV (combination therapy)', 'With NRTIs',
         '⚠️ CNS effects, rash, ⚠️ teratogenic', 'Pregnancy (1st trimester)',
         '⚠️ CYP3A4 inducer', 'Hepatic metabolism', 'Take at bedtime, avoid in pregnancy'),
        ('NNRTI', 'Nevirapine (Viramune)', 'Allosteric RT binding → dysfunction', 'Oral',
         'HIV (combination therapy)', 'With NRTIs',
         '⚠️ Rash, ⚠️ hepatotoxicity', 'Severe hepatic impairment',
         '⚠️ CYP3A4 inducer', 'Hepatic metabolism', 'Monitor LFTs, dose escalation'),
        ('NNRTI', 'Etravirine (Intelence)', 'Allosteric RT binding → dysfunction', 'Oral',
         'HIV (treatment-experienced)', 'With boosted PIs',
         'Rash (mild), nausea', 'None',
         'CYP substrate', 'Hepatic metabolism', 'Active against resistant strains'),

        # Protease Inhibitors (Light Orchid - index 2)
        ('PI', 'Darunavir (Prezista)', 'Protease inhibition → immature virions', 'Oral',
         '🟢 DOC for treatment-experienced', 'Boosted with ritonavir/cobicistat',
         'GI upset, ⚠️ rash, hyperlipidemia', 'Sulfa allergy',
         '⚠️ CYP3A4 substrate/inhibitor', 'Hepatic metabolism', 'High resistance barrier'),
        ('PI', 'Atazanavir (Reyataz)', 'Protease inhibition → immature virions', 'Oral',
         'HIV (treatment-naïve)', 'Boosted with ritonavir/cobicistat',
         '⚠️ Hyperbilirubinemia, ⚠️ PR prolongation', '2nd/3rd degree AV block',
         '⚠️ CYP3A4 inhibitor', 'Hepatic metabolism', 'Monitor ECG, benign jaundice'),
        ('PI', 'Lopinavir (Kaletra)', 'Protease inhibition → immature virions', 'Oral',
         'HIV (salvage therapy)', 'Co-formulated with ritonavir',
         '⚠️ Pancreatitis, hyperlipidemia', 'Hepatic impairment',
         '⚠️ CYP3A4 inhibitor', 'Hepatic metabolism', 'Monitor lipids, glucose'),
    ]

    current_row = 2
    class_index = -1
    prev_class = None

    for row_data in master_data:
        drug_class = row_data[0]

        if drug_class != prev_class:
            class_index += 1
            prev_class = drug_class

        colors = get_color_set(class_index)
        color = colors['main']

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            if col_idx == 1:
                cell.font = Font(name='Calibri', bold=True, size=10, color='000000')
            else:
                cell.font = Font(name='Calibri', size=10, color='000000')
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
            cell.value = value

        ws.row_dimensions[current_row].height = 60
        current_row += 1

    return ws

# =============================================================================
# MAIN FUNCTION
# =============================================================================

def create_sample_chart():
    wb = Workbook()

    create_drug_details_tab(wb)
    create_key_comparisons_tab(wb)
    create_master_chart_tab(wb)

    output_path = 'HIV_3_Drug_Classes_Sample.xlsx'
    wb.save(output_path)
    print(f"✅ Sample drug chart created: {output_path}")
    print(f"\nStructure:")
    print(f"  Tab 1: Drug Details (3 drug classes)")
    print(f"    - NRTIs (Ice Blue)")
    print(f"    - NNRTIs (Seafoam)")
    print(f"    - Protease Inhibitors (Light Orchid)")
    print(f"  Tab 2: Key Comparisons (5 comparison tables)")
    print(f"  Tab 3: Master Chart (10 drugs total)")
    print(f"\nFeatures:")
    print(f"  ✓ 10-category standardized order")
    print(f"  ✓ Emoji notation (🟢 DOC, ⚠️ toxicity)")
    print(f"  ✓ Grouped comparisons in Tab 2")
    print(f"  ✓ Color-coded by drug class")
    print(f"  ✓ No analogies or memory tricks")

if __name__ == '__main__':
    create_sample_chart()
