#!/usr/bin/env python3
"""
SAMPLE: 3 Drug Classes to Show Color Rotation
Demonstrates the fixed styling with multiple drug classes
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# COLOR SCHEME - 3-Shade System
# =============================================================================

MAIN_TITLE_COLOR = '4472C4'  # Dark blue for headers

# Color Set 0: Ice Blue
ICE_BLUE_HEADER = 'B4C6E7'
ICE_BLUE_MAIN = 'D9E2F3'
ICE_BLUE_ROW_LABEL = 'C5D3ED'

# Color Set 1: Seafoam
SEAFOAM_HEADER = 'A5D6A7'
SEAFOAM_MAIN = 'C8E6C9'
SEAFOAM_ROW_LABEL = 'B7DDB9'

# Color Set 2: Light Orchid
LIGHT_ORCHID_HEADER = 'B39DDB'
LIGHT_ORCHID_MAIN = 'D1C4E9'
LIGHT_ORCHID_ROW_LABEL = 'C2B2E0'

MNEMONIC_BG = 'E6F3FF'
CLINICAL_PEARL_BG = 'E8F5E9'
ANALOGY_BOX_BG = 'FFF9E6'

COLOR_SETS = [
    {'header': ICE_BLUE_HEADER, 'main': ICE_BLUE_MAIN, 'row_label': ICE_BLUE_ROW_LABEL},
    {'header': SEAFOAM_HEADER, 'main': SEAFOAM_MAIN, 'row_label': SEAFOAM_ROW_LABEL},
    {'header': LIGHT_ORCHID_HEADER, 'main': LIGHT_ORCHID_MAIN, 'row_label': LIGHT_ORCHID_ROW_LABEL},
]

thin_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='FFFFFF'),
    bottom=Side(style='thin', color='FFFFFF')
)

def get_color_set(index):
    return COLOR_SETS[index % len(COLOR_SETS)]

def apply_cell_style(cell, text='', bold=False, font_size=10, bg_color=None, alignment='left'):
    cell.value = text
    cell.font = Font(name='Calibri', size=font_size, bold=bold, color='000000')
    cell.alignment = Alignment(horizontal=alignment, vertical='top', wrap_text=True)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    cell.border = thin_border

def set_column_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def add_mnemonic_row(ws, row, text, span_cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row, 1)
    apply_cell_style(cell, text=f"💡 MEMORY TRICKS: {text}", font_size=11, bg_color=MNEMONIC_BG)
    ws.row_dimensions[row].height = 60

def add_comparison_table(ws, start_row, title, headers, data_rows, colors):
    current_row = start_row
    num_cols = len(headers)
    end_col = get_column_letter(num_cols)

    # Title
    ws.merge_cells(f'A{current_row}:{end_col}{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = title
    cell.font = Font(name='Calibri', bold=True, size=14, color='000000')
    cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 25
    current_row += 1

    # Headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(current_row, col_idx)
        cell.value = header
        cell.font = Font(name='Calibri', bold=True, size=11, color='000000')
        cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 23
    current_row += 1

    # Data rows - all same color
    for row_data in data_rows:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            cell.value = value
            cell.font = Font(name='Calibri', bold=(col_idx==1), size=10, color='000000')
            cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 43
        current_row += 1

    return current_row

# =============================================================================
# CREATE WORKBOOK
# =============================================================================

wb = Workbook()

# =============================================================================
# TAB 1: DRUG DETAILS (3 Classes)
# =============================================================================

ws = wb.active
ws.title = "Drug Details"

set_column_widths(ws, {
    'A': 30, 'B': 25, 'C': 25, 'D': 25, 'E': 40
})

current_row = 1

# CLASS 1: NRTIs (Ice Blue)
colors = get_color_set(0)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
cell = ws.cell(current_row, 1)
apply_cell_style(cell, text="NUCLEOSIDE REVERSE TRANSCRIPTASE INHIBITORS (NRTIs)",
                bold=True, font_size=18, bg_color=colors['header'], alignment='center')
cell.font = Font(name='Calibri', size=18, bold=True, color='000000')
ws.row_dimensions[current_row].height = 28
current_row += 1

# Drug names
headers = ['Property', 'Tenofovir', 'Emtricitabine', 'Lamivudine', 'Analogy']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(current_row, col_idx)
    if col_idx == 1:
        apply_cell_style(cell, text=header, bold=True, font_size=14, bg_color=colors['row_label'], alignment='center')
    else:
        apply_cell_style(cell, text=header, bold=True, font_size=16, bg_color=colors['main'], alignment='center')
ws.row_dimensions[current_row].height = 30
current_row += 1

# Properties
properties = [
    ('Mechanism', 'Nucleoside analog → chain termination', 'Fake bricks stop construction'),
    ('Uses', '🟢 HIV first-line', ''),
    ('Adverse Effects', '⚠️ Nephrotoxicity (Tenofovir)', ''),
]

for prop_name, prop_value, analogy in properties:
    cell_a = ws.cell(current_row, 1)
    apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

    # Apply styling to ALL cells BEFORE merging (B, C, D)
    for col_idx in range(2, 5):  # Columns 2, 3, 4 = B, C, D
        cell = ws.cell(current_row, col_idx)
        apply_cell_style(cell, text=prop_value if col_idx==2 else '',
                        font_size=12, bg_color=colors['main'])

    # Now merge cells
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

    cell_analogy = ws.cell(current_row, 5)
    apply_cell_style(cell_analogy, text=analogy, font_size=10,
                    bg_color=ANALOGY_BOX_BG if analogy else colors['main'])

    ws.row_dimensions[current_row].height = 48
    current_row += 1

add_mnemonic_row(ws, current_row, '"NRTI = Nucleoside Rival Terminating Infection"', span_cols=5)
current_row += 2

# CLASS 2: NNRTIs (Seafoam)
colors = get_color_set(1)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
cell = ws.cell(current_row, 1)
apply_cell_style(cell, text="NON-NUCLEOSIDE REVERSE TRANSCRIPTASE INHIBITORS (NNRTIs)",
                bold=True, font_size=18, bg_color=colors['header'], alignment='center')
cell.font = Font(name='Calibri', size=18, bold=True, color='000000')
ws.row_dimensions[current_row].height = 28
current_row += 1

# Drug names
headers = ['Property', 'Efavirenz', 'Rilpivirine', 'Nevirapine', 'Analogy']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(current_row, col_idx)
    if col_idx == 1:
        apply_cell_style(cell, text=header, bold=True, font_size=14, bg_color=colors['row_label'], alignment='center')
    else:
        apply_cell_style(cell, text=header, bold=True, font_size=16, bg_color=colors['main'], alignment='center')
ws.row_dimensions[current_row].height = 30
current_row += 1

# Properties
properties = [
    ('Mechanism', 'Allosteric inhibition of reverse transcriptase', 'Jamming the gears'),
    ('Uses', 'HIV treatment (combination)', ''),
    ('Adverse Effects', '⚠️ CNS effects, rash', ''),
]

for prop_name, prop_value, analogy in properties:
    cell_a = ws.cell(current_row, 1)
    apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

    # Apply styling to ALL cells BEFORE merging (B, C, D)
    for col_idx in range(2, 5):  # Columns 2, 3, 4 = B, C, D
        cell = ws.cell(current_row, col_idx)
        apply_cell_style(cell, text=prop_value if col_idx==2 else '',
                        font_size=12, bg_color=colors['main'])

    # Now merge cells
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

    cell_analogy = ws.cell(current_row, 5)
    apply_cell_style(cell_analogy, text=analogy, font_size=10,
                    bg_color=ANALOGY_BOX_BG if analogy else colors['main'])

    ws.row_dimensions[current_row].height = 48
    current_row += 1

add_mnemonic_row(ws, current_row, '"NNRTI = Non-Nucleoside Not Required Triphosphorylation Inside"', span_cols=5)
current_row += 2

# CLASS 3: Protease Inhibitors (Light Orchid)
colors = get_color_set(2)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
cell = ws.cell(current_row, 1)
apply_cell_style(cell, text="PROTEASE INHIBITORS (PIs)",
                bold=True, font_size=18, bg_color=colors['header'], alignment='center')
cell.font = Font(name='Calibri', size=18, bold=True, color='000000')
ws.row_dimensions[current_row].height = 28
current_row += 1

# Drug names
headers = ['Property', 'Ritonavir', 'Atazanavir', 'Darunavir', 'Analogy']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(current_row, col_idx)
    if col_idx == 1:
        apply_cell_style(cell, text=header, bold=True, font_size=14, bg_color=colors['row_label'], alignment='center')
    else:
        apply_cell_style(cell, text=header, bold=True, font_size=16, bg_color=colors['main'], alignment='center')
ws.row_dimensions[current_row].height = 30
current_row += 1

# Properties
properties = [
    ('Mechanism', 'Inhibits HIV protease → immature virions', 'Factory shutdown'),
    ('Uses', 'HIV salvage therapy', ''),
    ('Adverse Effects', '⚠️ GI upset, hyperglycemia, hyperlipidemia', ''),
]

for prop_name, prop_value, analogy in properties:
    cell_a = ws.cell(current_row, 1)
    apply_cell_style(cell_a, text=prop_name, bold=True, font_size=14, bg_color=colors['row_label'])

    # Apply styling to ALL cells BEFORE merging (B, C, D)
    for col_idx in range(2, 5):  # Columns 2, 3, 4 = B, C, D
        cell = ws.cell(current_row, col_idx)
        apply_cell_style(cell, text=prop_value if col_idx==2 else '',
                        font_size=12, bg_color=colors['main'])

    # Now merge cells
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)

    cell_analogy = ws.cell(current_row, 5)
    apply_cell_style(cell_analogy, text=analogy, font_size=10,
                    bg_color=ANALOGY_BOX_BG if analogy else colors['main'])

    ws.row_dimensions[current_row].height = 48
    current_row += 1

add_mnemonic_row(ws, current_row, '"All PIs end in -navir"', span_cols=5)

# =============================================================================
# TAB 2: KEY COMPARISONS
# =============================================================================

ws2 = wb.create_sheet("Key Comparisons")
set_column_widths(ws2, {'A': 30, 'B': 35, 'C': 35, 'D': 35})

current_row = 1

# Comparison 1: Mechanisms (Ice Blue)
current_row = add_comparison_table(
    ws2, current_row,
    'MECHANISMS ACROSS CLASSES',
    ['Drug Class', 'Mechanism', 'Target', 'Result'],
    [
        ['NRTI', 'Nucleoside analog → chain termination', 'Reverse transcriptase', 'DNA synthesis stops'],
        ['NNRTI', 'Allosteric binding', 'Reverse transcriptase', 'Enzyme dysfunction'],
        ['Protease Inhibitor', 'Competitive inhibition', 'HIV protease', 'Immature virions'],
    ],
    get_color_set(0)
)
current_row += 2

# Comparison 2: Toxicities (Seafoam)
current_row = add_comparison_table(
    ws2, current_row,
    'MAJOR TOXICITIES',
    ['Drug Class', 'Key Toxicity', 'Monitoring'],
    [
        ['NRTI', '⚠️ Lactic acidosis, nephrotoxicity', 'Lactate, renal function'],
        ['NNRTI', 'Rash, hepatotoxicity', 'Skin exam, LFTs'],
        ['Protease Inhibitor', 'Hyperglycemia, hyperlipidemia', 'Glucose, lipids'],
    ],
    get_color_set(1)
)
current_row += 2

# Comparison 3: Clinical Uses (Light Orchid)
current_row = add_comparison_table(
    ws2, current_row,
    'CLINICAL USES',
    ['Drug Class', 'Primary Use', 'Notes'],
    [
        ['NRTI', '🟢 First-line HIV therapy', 'Backbone of regimens'],
        ['NNRTI', 'HIV treatment (combination)', 'Cannot use as monotherapy'],
        ['Protease Inhibitor', 'Salvage therapy', 'Treatment-experienced patients'],
    ],
    get_color_set(2)
)

# =============================================================================
# TAB 3: MASTER CHART
# =============================================================================

ws3 = wb.create_sheet("Master Chart")
set_column_widths(ws3, {
    'A': 20, 'B': 25, 'C': 35, 'D': 35, 'E': 35
})

# Header
headers = ['Drug Class', 'Drug Name', 'Mechanism', 'Uses', 'Adverse Effects']
for col_idx, header in enumerate(headers, start=1):
    cell = ws3.cell(1, col_idx)
    cell.value = header
    cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=MAIN_TITLE_COLOR, end_color=MAIN_TITLE_COLOR, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws3.row_dimensions[1].height = 25
ws3.freeze_panes = 'A2'

# Data - showing color rotation by class
data = [
    ('NRTI', 'Tenofovir', 'Nucleoside analog → chain termination', '🟢 HIV first-line', '⚠️ Nephrotoxicity'),
    ('NRTI', 'Emtricitabine', 'Nucleoside analog → chain termination', '🟢 HIV first-line', 'Minimal'),
    ('NRTI', 'Lamivudine', 'Nucleoside analog → chain termination', 'HIV, HBV', 'Minimal'),
    ('NNRTI', 'Efavirenz', 'Allosteric inhibition', 'HIV treatment', '⚠️ CNS effects, rash'),
    ('NNRTI', 'Rilpivirine', 'Allosteric inhibition', 'HIV treatment', 'Rash, insomnia'),
    ('NNRTI', 'Nevirapine', 'Allosteric inhibition', 'HIV treatment', 'Hepatotoxicity, rash'),
    ('Protease Inhibitor', 'Ritonavir', 'Inhibits HIV protease', 'Boosting agent', 'GI upset'),
    ('Protease Inhibitor', 'Atazanavir', 'Inhibits HIV protease', 'HIV treatment', 'Hyperbilirubinemia'),
    ('Protease Inhibitor', 'Darunavir', 'Inhibits HIV protease', 'HIV salvage', 'Rash'),
]

current_row = 2
class_index = -1
prev_class = None

for row_data in data:
    drug_class = row_data[0]

    if drug_class != prev_class:
        class_index += 1
        prev_class = drug_class

    colors = get_color_set(class_index)
    color = colors['main']

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws3.cell(current_row, col_idx)
        cell.value = value
        cell.font = Font(name='Calibri', bold=(col_idx==1), size=10, color='000000')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

    ws3.row_dimensions[current_row].height = 60
    current_row += 1

# =============================================================================
# TAB 4: HIGH-YIELD & PEARLS
# =============================================================================

ws4 = wb.create_sheet("High-Yield & Pearls")
set_column_widths(ws4, {'A': 100})

current_row = 1

# Clinical Pearls
colors = get_color_set(0)
ws4.merge_cells(f'A{current_row}:A{current_row}')
cell = ws4.cell(current_row, 1)
cell.value = "CLINICAL PEARLS"
cell.font = Font(name='Calibri', bold=True, size=14, color='000000')
cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.border = thin_border
ws4.row_dimensions[current_row].height = 23
current_row += 1

pearls = """• NRTIs are backbone of most HIV regimens
• Always screen HLA-B*5701 before starting Abacavir
• Protease inhibitors often boosted with ritonavir
• NNRTIs have low barrier to resistance"""

cell = ws4.cell(current_row, 1)
cell.value = pearls
cell.font = Font(name='Calibri', size=11, color='000000')
cell.fill = PatternFill(start_color=CLINICAL_PEARL_BG, end_color=CLINICAL_PEARL_BG, fill_type='solid')
cell.alignment = Alignment(wrap_text=True, vertical='top')
cell.border = thin_border
ws4.row_dimensions[current_row].height = 60
current_row += 2

# Mnemonics
colors = get_color_set(1)
cell = ws4.cell(current_row, 1)
cell.value = "MEMORY TRICKS"
cell.font = Font(name='Calibri', bold=True, size=14, color='000000')
cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.border = thin_border
ws4.row_dimensions[current_row].height = 23
current_row += 1

mnemonics = """💡 All PIs end in -navir
💡 Tenofovir = Tender kidneys (nephrotoxicity)
💡 NRTI needs activation by Nucleotide kinases"""

cell = ws4.cell(current_row, 1)
cell.value = mnemonics
cell.font = Font(name='Calibri', size=11, color='000000')
cell.fill = PatternFill(start_color=MNEMONIC_BG, end_color=MNEMONIC_BG, fill_type='solid')
cell.alignment = Alignment(wrap_text=True, vertical='top')
cell.border = thin_border
ws4.row_dimensions[current_row].height = 60

# Save
wb.save('HIV_3_Classes_Sample.xlsx')
print("✅ Sample created: HIV_3_Classes_Sample.xlsx")
print("\n📊 COLOR DEMONSTRATION:")
print("  Tab 1 (Drug Details):")
print("    - NRTIs: Ice Blue (#D9E2F3)")
print("    - NNRTIs: Seafoam (#C8E6C9)")
print("    - Protease Inhibitors: Light Orchid (#D1C4E9)")
print("\n  Tab 2 (Key Comparisons):")
print("    - Table 1 (Mechanisms): Ice Blue - ALL rows same color")
print("    - Table 2 (Toxicities): Seafoam - ALL rows same color")
print("    - Table 3 (Uses): Light Orchid - ALL rows same color")
print("\n  Tab 3 (Master Chart):")
print("    - NRTIs (3 drugs): ALL columns Ice Blue")
print("    - NNRTIs (3 drugs): ALL columns Seafoam")
print("    - PIs (3 drugs): ALL columns Light Orchid")
