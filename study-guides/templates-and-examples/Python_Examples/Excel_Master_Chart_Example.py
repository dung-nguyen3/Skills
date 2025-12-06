#!/usr/bin/env python3
"""
EXCEL MASTER CHART EXAMPLE
Single-sheet comprehensive reference chart

This is a simpler format than the 4-tab drug chart.
Use for: Quick-reference charts, condition lists, lab values

Structure:
- Single sheet: "Master Chart"
- Frozen header row
- Color-coded rows by category/class
- All items in one searchable table
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =============================================================================
# COLOR SCHEME - 3-Shade System (See Excel_Color_Reference.txt)
# =============================================================================

MAIN_TITLE_COLOR = '4472C4'  # Dark blue for headers

# Color Set 0: Ice Blue
ICE_BLUE_HEADER = 'B4C6E7'
ICE_BLUE_MAIN = 'D9E2F3'
ICE_BLUE_ROW_LABEL = 'C5D3ED'

# Color Set 1: Seafoam
SEAFOAM_HEADER = 'A5D6A7'
SEAFOAM_MAIN = 'C8E6C9'
SEAFOAM_ROW_LABEL = 'B7DDB9'

# Color Set 2: Light Orchid
LIGHT_ORCHID_HEADER = 'B39DDB'
LIGHT_ORCHID_MAIN = 'D1C4E9'
LIGHT_ORCHID_ROW_LABEL = 'C2B2E0'

# Color Set 3: Champagne
CHAMPAGNE_HEADER = 'F4D9B3'
CHAMPAGNE_MAIN = 'F7E7CE'
CHAMPAGNE_ROW_LABEL = 'F6E0C0'

# Color Set 4: Sky Blue
SKY_BLUE_HEADER = '9ECAE1'
SKY_BLUE_MAIN = 'BDD7EE'
SKY_BLUE_ROW_LABEL = 'ACD0E7'

# Color Set 5: Pale Azure
PALE_AZURE_HEADER = 'D9ECFF'
PALE_AZURE_MAIN = 'F0F8FF'
PALE_AZURE_ROW_LABEL = 'E5F2FF'

# Color Set 6: Blush Pink
BLUSH_PINK_HEADER = 'F8BBD0'
BLUSH_PINK_MAIN = 'FCE4EC'
BLUSH_PINK_ROW_LABEL = 'FAD2DE'

# Color Set 7: Soft Lilac
SOFT_LILAC_HEADER = 'D1C4E9'
SOFT_LILAC_MAIN = 'EDE7F6'
SOFT_LILAC_ROW_LABEL = 'DFD6ED'

# Color Set 8: Soft Tangerine
SOFT_TANGERINE_HEADER = 'FFD4B3'
SOFT_TANGERINE_MAIN = 'FFE8D6'
SOFT_TANGERINE_ROW_LABEL = 'FFDEC4'

# Color Set 9: Powder Blue
POWDER_BLUE_HEADER = '90CAF9'
POWDER_BLUE_MAIN = 'BBDEFB'
POWDER_BLUE_ROW_LABEL = 'A5D6FA'

COLOR_SETS = [
    {'header': ICE_BLUE_HEADER, 'main': ICE_BLUE_MAIN, 'row_label': ICE_BLUE_ROW_LABEL},
    {'header': SEAFOAM_HEADER, 'main': SEAFOAM_MAIN, 'row_label': SEAFOAM_ROW_LABEL},
    {'header': LIGHT_ORCHID_HEADER, 'main': LIGHT_ORCHID_MAIN, 'row_label': LIGHT_ORCHID_ROW_LABEL},
    {'header': CHAMPAGNE_HEADER, 'main': CHAMPAGNE_MAIN, 'row_label': CHAMPAGNE_ROW_LABEL},
    {'header': SKY_BLUE_HEADER, 'main': SKY_BLUE_MAIN, 'row_label': SKY_BLUE_ROW_LABEL},
    {'header': PALE_AZURE_HEADER, 'main': PALE_AZURE_MAIN, 'row_label': PALE_AZURE_ROW_LABEL},
    {'header': BLUSH_PINK_HEADER, 'main': BLUSH_PINK_MAIN, 'row_label': BLUSH_PINK_ROW_LABEL},
    {'header': SOFT_LILAC_HEADER, 'main': SOFT_LILAC_MAIN, 'row_label': SOFT_LILAC_ROW_LABEL},
    {'header': SOFT_TANGERINE_HEADER, 'main': SOFT_TANGERINE_MAIN, 'row_label': SOFT_TANGERINE_ROW_LABEL},
    {'header': POWDER_BLUE_HEADER, 'main': POWDER_BLUE_MAIN, 'row_label': POWDER_BLUE_ROW_LABEL},
]

def get_color_set(index):
    """Get color set by index (rotates through available sets)"""
    return COLOR_SETS[index % len(COLOR_SETS)]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def set_cell_style(cell, text, bold=False, bg_color=None):
    """Apply cell styling"""
    cell.value = text
    cell.font = Font(name='Calibri', size=10, bold=bold, color='000000')
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    # White borders
    cell.border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

# =============================================================================
# CREATE MASTER CHART
# =============================================================================

def create_master_chart(output_path, headers, column_widths, data_rows):
    """
    Create single-sheet master chart

    Args:
        output_path: File path to save
        headers: List of column headers (user-defined)
        column_widths: Dict of column letters → widths
        data_rows: List of tuples (category, values...)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Chart"

    # Set column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ==========================================================================
    # HEADER ROW
    # ==========================================================================
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header_text
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=MAIN_TITLE_COLOR, end_color=MAIN_TITLE_COLOR, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = 'A2'  # Freeze header

    # ==========================================================================
    # DATA ROWS
    # ==========================================================================
    current_row = 2
    category_index = -1  # Start at -1 so first category becomes 0
    prev_category = None

    for row_data in data_rows:
        category = row_data[0]

        # Change color when category changes
        if category != prev_category:
            category_index += 1
            prev_category = category

        # Get color set and use 'main' shade for data cells
        colors = get_color_set(category_index)
        color = colors['main']

        # Apply data to cells
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)  # First column (category) is bold
            set_cell_style(cell, text=value, bold=bold, bg_color=color)

        ws.row_dimensions[current_row].height = 60
        current_row += 1

    # Save
    wb.save(output_path)
    print(f"✅ Master chart created: {output_path}")

# =============================================================================
# EXAMPLE USAGE
# =============================================================================

if __name__ == '__main__':
    # Example 1: Drug Master Chart
    headers = [
        'Drug Class',
        'Drug Name (Brand)',
        'Route',
        'Mechanism',
        'Uses',
        'Adverse Effects',
        'Contraindications',
        'Special Considerations'
    ]

    column_widths = {
        'A': 20,  # Drug Class
        'B': 25,  # Drug Name
        'C': 15,  # Route
        'D': 35,  # Mechanism
        'E': 35,  # Uses
        'F': 35,  # Adverse Effects
        'G': 30,  # Contraindications
        'H': 30,  # Special Considerations
    }

    data_rows = [
        ('NRTI', 'Tenofovir (Viread)', 'Oral', 'Nucleoside analog → chain termination',
         '🟢 HIV - First line', '⚠️ Nephrotoxicity, ↓ bone density', 'CrCl <50',
         'Monitor renal function'),
        ('NRTI', 'Emtricitabine (Emtriva)', 'Oral', 'Nucleoside analog → chain termination',
         '🟢 HIV - First line', 'Minimal', 'None', 'Safe in pregnancy'),
        ('NRTI', 'Lamivudine (Epivir)', 'Oral', 'Nucleoside analog → chain termination',
         'HIV, HBV', 'Minimal', 'None', 'HBV coinfection'),
        ('NNRTI', 'Efavirenz (Sustiva)', 'Oral', 'Allosteric RT inhibition',
         'HIV', '⚠️ CNS effects, rash', 'Pregnancy (1st trimester)', 'Take at bedtime'),
        ('NNRTI', 'Rilpivirine (Edurant)', 'Oral', 'Allosteric RT inhibition',
         'HIV (with food)', 'Rash, insomnia', 'None', 'Take with meal'),
        ('Integrase Inhibitor', 'Dolutegravir (Tivicay)', 'Oral', 'Blocks integration',
         '🟢 HIV - Preferred', 'Minimal, insomnia', 'None', 'High barrier to resistance'),
    ]

    create_master_chart('HIV_Master_Chart.xlsx', headers, column_widths, data_rows)

    print("\n" + "="*70)

    # Example 2: Condition Master Chart
    headers2 = [
        'Condition',
        'Epidemiology',
        'Risk Factors',
        'Clinical Presentation',
        'Diagnosis',
        'Treatment'
    ]

    column_widths2 = {
        'A': 25,
        'B': 30,
        'C': 30,
        'D': 40,
        'E': 35,
        'F': 35,
    }

    data_rows2 = [
        ('Bacterial Pharyngitis', 'Common in children 5-15 years', 'Close contact, school',
         'Sore throat, fever, exudates, swollen lymph nodes',
         'Rapid strep test, throat culture', '🟢 Penicillin or amoxicillin'),
        ('Viral Pharyngitis', 'Most common cause of pharyngitis', 'Crowded places, winter',
         'Sore throat, rhinorrhea, cough', 'Clinical diagnosis', 'Supportive care'),
        ('Mononucleosis', 'Adolescents and young adults', 'EBV transmission',
         'Sore throat, fever, fatigue, splenomegaly', 'Monospot test, atypical lymphocytes',
         'Supportive, avoid contact sports'),
    ]

    create_master_chart('Pharyngitis_Master_Chart.xlsx', headers2, column_widths2, data_rows2)

    print("\n" + "="*70)

    # Example 3: Lab Values Master Chart
    headers3 = [
        'Test Name',
        'Normal Range',
        'Increased In',
        'Decreased In',
        'Clinical Significance'
    ]

    column_widths3 = {
        'A': 25,
        'B': 20,
        'C': 40,
        'D': 40,
        'E': 40,
    }

    data_rows3 = [
        ('Hemoglobin (Hgb)', '13-17 g/dL (M), 12-15 g/dL (F)',
         'Polycythemia, dehydration, COPD, high altitude',
         'Anemia (blood loss, hemolysis, marrow failure)',
         'Oxygen-carrying capacity'),
        ('Platelet Count', '150,000-400,000/μL',
         'Thrombocytosis (infection, malignancy, post-splenectomy)',
         'Thrombocytopenia (ITP, DIC, HIT, aplasia)',
         'Bleeding risk if <50,000'),
        ('White Blood Cell Count', '4,000-11,000/μL',
         'Infection, inflammation, leukemia, stress',
         'Bone marrow suppression, autoimmune, overwhelming infection',
         'Immune function'),
    ]

    create_master_chart('Lab_Values_Master_Chart.xlsx', headers3, column_widths3, data_rows3)
