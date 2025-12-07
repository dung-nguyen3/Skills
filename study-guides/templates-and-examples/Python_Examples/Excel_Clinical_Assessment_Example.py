#!/usr/bin/env python3
"""
COMPLETE EXCEL CLINICAL ASSESSMENT CHART EXAMPLE
Reference implementation for creating comprehensive 4-tab clinical assessment charts

This file shows a COMPLETE working example with all 4 tabs.
Use for clinical conditions requiring H&P-focused study (diseases, syndromes).

Structure:
- Tab 1: Key Comparisons (clinical differentials, side-by-side)
- Tab 2: Master Chart (7 clinical columns, onset-based grouping)
- Tab 3: H&P Guide (comprehensive: OLDCAARTS, PMH/FH/SH, ROS, Physical Exam, Patient Education, Red Flags)
- Tab 4: Summary (mnemonics, "If X Think Y", RED FLAGS, high-yield pearls)

Example: Common Pain Syndromes (placeholder content)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# COLOR SCHEME (See Excel_Color_Reference.txt for details)
# =============================================================================

MAIN_TITLE_BG = '4472C4'  # Dark blue - for sheet titles and master chart header
MNEMONIC_BG = 'E6F3FF'    # Light blue for mnemonics
CLINICAL_PEARL_BG = 'E8F5E9'  # Light green for clinical pearls
RED_FLAG_BG = 'FFEBEE'    # Light red for red flags

# H&P Guide tab colors
HP_SECTION_BG = '4472C4'      # Dark blue for main section headers
HP_HISTORY_BG = 'C8E6C9'      # Light green for history sections (OLDCAARTS, PMH/FH/SH)
HP_ROS_BG = 'D1C4E9'          # Light purple for ROS section
HP_PE_BG = 'F7E7CE'           # Peach for Physical Exam section
HP_EDUCATION_BG = 'BDD7EE'    # Light blue for Patient Education
HP_RED_FLAG_BG = 'FCE4EC'     # Pink for Red Flags section

# ONSET-BASED COLORS for Master Chart
# Conditions are grouped by onset pattern, not generic category
ONSET_COLORS = {
    'Acute': 'FCE4EC',           # Blush Pink - emergencies
    'Acute/Subacute': 'FFE8D6',  # Soft Tangerine - days to weeks
    'Subacute': 'F7E7CE',        # Champagne - weeks
    'Chronic': 'C8E6C9',         # Seafoam - months to years
    'Chronic Inflammatory': 'BDD7EE',  # Sky Blue - chronic with inflammation
}

# COLOR_SETS for Key Comparisons tab (3-shade system matching template)
ICE_BLUE_HEADER = 'B4C6E7'
ICE_BLUE_MAIN = 'D9E2F3'
ICE_BLUE_ROW_LABEL = 'C5D3ED'

SEAFOAM_HEADER = 'A5D6A7'
SEAFOAM_MAIN = 'C8E6C9'
SEAFOAM_ROW_LABEL = 'B7DDB9'

LIGHT_ORCHID_HEADER = 'B39DDB'
LIGHT_ORCHID_MAIN = 'D1C4E9'
LIGHT_ORCHID_ROW_LABEL = 'C2B2E0'

CHAMPAGNE_HEADER = 'F4D9B3'
CHAMPAGNE_MAIN = 'F7E7CE'
CHAMPAGNE_ROW_LABEL = 'F6E0C0'

SKY_BLUE_HEADER = '9ECAE1'
SKY_BLUE_MAIN = 'BDD7EE'
SKY_BLUE_ROW_LABEL = 'ACD0E7'

COLOR_SETS = [
    {'header': ICE_BLUE_HEADER, 'main': ICE_BLUE_MAIN, 'row_label': ICE_BLUE_ROW_LABEL},
    {'header': SEAFOAM_HEADER, 'main': SEAFOAM_MAIN, 'row_label': SEAFOAM_ROW_LABEL},
    {'header': LIGHT_ORCHID_HEADER, 'main': LIGHT_ORCHID_MAIN, 'row_label': LIGHT_ORCHID_ROW_LABEL},
    {'header': CHAMPAGNE_HEADER, 'main': CHAMPAGNE_MAIN, 'row_label': CHAMPAGNE_ROW_LABEL},
    {'header': SKY_BLUE_HEADER, 'main': SKY_BLUE_MAIN, 'row_label': SKY_BLUE_ROW_LABEL},
]

def get_color_set(index):
    """Get color set by index (rotates through available sets)"""
    return COLOR_SETS[index % len(COLOR_SETS)]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def apply_cell_style(cell, text='', bold=False, font_size=11, bg_color=None,
                     border=True, alignment='left', wrap=True, font_color='000000'):
    """Apply comprehensive cell styling"""
    cell.value = text
    cell.font = Font(name='Calibri', size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical='top',
        wrap_text=wrap
    )

    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    if border:
        cell.border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

def create_comparison_header(ws, title, row, span_cols=4, table_index=0):
    """Create merged title row for a comparison table"""
    colors = get_color_set(table_index)
    # Apply styling to ALL cells BEFORE merging
    for col_idx in range(1, span_cols + 1):
        cell = ws.cell(row, col_idx)
        apply_cell_style(cell, text=title if col_idx==1 else '', bold=True, font_size=14,
                        bg_color=colors['header'], alignment='center', font_color='000000')
        cell.font = Font(name='Calibri', size=14, bold=True, color='000000')

    # Now merge cells
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 30

def create_column_headers(ws, headers, row, start_col=1, table_index=0):
    """Create column headers for comparison table"""
    colors = get_color_set(table_index)
    header_color = colors['main']
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row, col_idx)
        apply_cell_style(cell, text=header, bold=True, font_size=11,
                        bg_color=header_color, alignment='center', font_color='000000')
    ws.row_dimensions[row].height = 25

def create_master_header_row(ws, headers, row=1):
    """Create formatted header row with freeze for Master Chart"""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx)
        apply_cell_style(
            cell, text=header, bold=True, font_size=12,
            bg_color=MAIN_TITLE_BG, alignment='center', font_color='FFFFFF'
        )
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')

    ws.row_dimensions[row].height = 28
    ws.freeze_panes = 'B2'  # Freeze header row AND condition column

def set_column_widths(ws, widths):
    """Set column widths from dictionary"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def add_mnemonic_row(ws, row, mnemonic_text, span_cols=4):
    """Add mnemonic row directly below a comparison table"""
    cell_a = ws.cell(row, 1)
    apply_cell_style(cell_a, text='MEMORY AID', bold=True, font_size=11,
                    bg_color=MNEMONIC_BG, alignment='left', font_color='0000FF')

    # Apply styling to ALL cells BEFORE merging (columns 2 through span_cols)
    for col_idx in range(2, span_cols + 1):
        cell = ws.cell(row, col_idx)
        apply_cell_style(cell, text=mnemonic_text if col_idx==2 else '', bg_color=MNEMONIC_BG)

    # Now merge cells
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 60

def add_section_header(ws, row, title, span_cols=1):
    """Add section header in Summary tab"""
    if span_cols > 1:
        # Apply styling to ALL cells BEFORE merging
        for col_idx in range(1, span_cols + 1):
            cell = ws.cell(row, col_idx)
            apply_cell_style(cell, text=title if col_idx==1 else '', bold=True, font_size=14,
                            bg_color=MAIN_TITLE_BG, alignment='center', font_color='FFFFFF')
            cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')

        # Now merge cells
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    else:
        cell = ws.cell(row, 1)
        apply_cell_style(cell, text=title, bold=True, font_size=14,
                        bg_color=MAIN_TITLE_BG, alignment='center', font_color='FFFFFF')
        cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws.row_dimensions[row].height = 30

def get_onset_color(onset_category):
    """Return color hex code based on onset category"""
    return ONSET_COLORS.get(onset_category, 'C8E6C9')  # Default to Seafoam

# =============================================================================
# TAB 1: KEY COMPARISONS
# =============================================================================

def create_key_comparisons_tab(wb):
    """
    Tab 1: Key Comparisons
    - Multiple side-by-side comparison tables for clinical differentials
    - One comparison type per table
    - Mnemonics placed directly below relevant tables
    """
    ws = wb.active
    ws.title = "Key Comparisons"

    # Column widths
    set_column_widths(ws, {
        'A': 25,  # Category/Row label
        'B': 35,  # Condition 1
        'C': 35,  # Condition 2
        'D': 35,  # Condition 3
    })

    current_row = 1

    # =========================================================================
    # COMPARISON TABLE 1: Acute vs Chronic Pain (Ice Blue)
    # =========================================================================
    table_index = 0

    create_comparison_header(ws, "ACUTE vs CHRONIC PAIN PRESENTATION", current_row,
                            span_cols=3, table_index=table_index)
    current_row += 1

    headers = ['Feature', 'Acute Pain', 'Chronic Pain']
    create_column_headers(ws, headers, current_row, table_index=table_index)
    current_row += 1

    comparison_data = [
        ('Duration', '<6 weeks', '>3 months'),
        ('Onset', 'Sudden, identifiable trigger', 'Gradual, may be insidious'),
        ('Character', 'Sharp, localized', 'Dull, diffuse, may radiate'),
        ('Associated Symptoms', 'Autonomic signs (sweating, tachycardia)', 'Sleep disturbance, mood changes'),
        ('Treatment Focus', 'Address underlying cause', 'Multimodal, functional improvement'),
    ]

    colors = get_color_set(table_index)
    table_main = colors['main']
    for row_data in comparison_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_main)
        current_row += 1

    current_row += 3

    # =========================================================================
    # COMPARISON TABLE 2: Inflammatory vs Mechanical (Seafoam)
    # =========================================================================
    table_index = 1

    create_comparison_header(ws, "INFLAMMATORY vs MECHANICAL PAIN", current_row,
                            span_cols=3, table_index=table_index)
    current_row += 1

    headers = ['Feature', 'Inflammatory', 'Mechanical']
    create_column_headers(ws, headers, current_row, table_index=table_index)
    current_row += 1

    comparison_data = [
        ('Age of Onset', '<40 years', 'Any age'),
        ('Type of Onset', 'Insidious (gradual)', 'Acute (sudden) or gradual'),
        ('Morning Stiffness', '⭐ >60 minutes', '<30 minutes'),
        ('Effect of Exercise', 'Improves with activity', 'Worsens with activity'),
        ('Effect of Rest', 'No improvement', 'Improves with rest'),
        ('Night Pain', 'Present, improves on rising', 'Less common'),
    ]

    colors = get_color_set(table_index)
    table_main = colors['main']
    for row_data in comparison_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_main)
        current_row += 1

    # Mnemonic
    current_row += 1
    add_mnemonic_row(ws, current_row,
        'Inflammatory = "REST makes it WORSE, ACTIVITY makes it BETTER"', span_cols=3)
    current_row += 3

    # =========================================================================
    # COMPARISON TABLE 3: Physical Exam Tests (Light Orchid)
    # =========================================================================
    table_index = 2

    create_comparison_header(ws, "PHYSICAL EXAM TESTS INTERPRETATION", current_row,
                            span_cols=4, table_index=table_index)
    current_row += 1

    headers = ['Test', 'How to Perform', 'Positive Finding', 'Indicates']
    create_column_headers(ws, headers, current_row, table_index=table_index)
    current_row += 1

    test_data = [
        ('Example Test A', 'Patient position, maneuver description', 'Pain reproduction, specific sign', 'Condition A'),
        ('Example Test B', 'Patient position, maneuver description', 'Positive if symptom reproduced', 'Condition B'),
        ('Example Test C', 'Patient position, maneuver description', 'Abnormal reflex or finding', 'Condition C'),
    ]

    colors = get_color_set(table_index)
    table_main = colors['main']
    for row_data in test_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_main)
        current_row += 1

    return ws

# =============================================================================
# TAB 2: MASTER CHART (Clinical Assessment Focus)
# =============================================================================

def create_master_chart_tab(wb):
    """
    Tab 2: Master Chart
    - 7 clinical columns for H&P-focused study
    - Conditions grouped by onset category
    - Color-coded by onset (not generic category)
    """
    ws = wb.create_sheet("Master Chart")

    # Column widths - 7 clinical columns
    set_column_widths(ws, {
        'A': 28,  # Condition
        'B': 18,  # Category/Onset
        'C': 45,  # Clinical Presentation
        'D': 45,  # Physical Exam Findings
        'E': 38,  # Diagnosis/Imaging
        'F': 42,  # Treatment
        'G': 45,  # Patient Education
    })

    # Headers - 7 clinical columns
    headers = ['Condition', 'Category', 'Clinical Presentation', 'Physical Exam Findings',
               'Diagnosis/Imaging', 'Treatment', 'Patient Education']
    create_master_header_row(ws, headers, 1)

    # Example conditions data - organized by ONSET
    conditions = [
        # === ACUTE (Blush Pink) ===
        {
            'condition': 'Acute Condition Example',
            'category': 'Acute',
            'presentation': '⚠️ EMERGENCY\nSudden onset symptoms\nSevere pain\nRapid progression',
            'physical_exam': '⚠️ Emergency exam findings\nVital sign abnormalities\nSpecific signs to look for',
            'diagnosis': '⚠️ URGENT imaging\nStat labs\nEmergency workup',
            'treatment': '⚠️ EMERGENCY\nImmediate intervention\nUrgent referral',
            'education': '⚠️ SEEK EMERGENCY CARE\nRed flags to watch\nTime-sensitive condition'
        },

        # === ACUTE/SUBACUTE (Soft Tangerine) ===
        {
            'condition': 'Acute/Subacute Condition A',
            'category': 'Acute/Subacute',
            'presentation': 'Onset over days to weeks\nProgressive symptoms\nMay have triggering event',
            'physical_exam': 'Specific exam findings\nPositive tests\nNeurologic exam if indicated',
            'diagnosis': 'Clinical diagnosis or\nImaging if not improving\nLabs as indicated',
            'treatment': '🟢 First-line: [Treatment]\nActivity modification\nPhysical therapy',
            'education': 'Expected course\nActivity recommendations\nWhen to follow up'
        },
        {
            'condition': 'Acute/Subacute Condition B',
            'category': 'Acute/Subacute',
            'presentation': 'Symptom pattern description\n⭐ Hallmark finding\nAssociated symptoms',
            'physical_exam': 'Inspection findings\nPalpation findings\n⭐ Key test result',
            'diagnosis': 'First-line imaging\nLabs to consider\nCriteria for diagnosis',
            'treatment': '🟢 First-line treatment\nAlternatives\nReferral criteria',
            'education': 'Prognosis\nLifestyle modifications\nFollow-up timing'
        },

        # === CHRONIC (Seafoam) ===
        {
            'condition': 'Chronic Condition A',
            'category': 'Chronic',
            'presentation': 'Months to years duration\nGradual progression\nFluctuating symptoms',
            'physical_exam': 'Chronic changes visible\nFunctional limitations\nSpecific findings',
            'diagnosis': 'Imaging findings\nNormal labs expected\nClinical criteria',
            'treatment': '🟢 Conservative management\nPhysical therapy\nMedications PRN',
            'education': 'Chronic condition management\nLifestyle important\nSelf-care strategies'
        },
        {
            'condition': 'Chronic Condition B',
            'category': 'Chronic',
            'presentation': 'Long-standing symptoms\nAssociated findings\nImpact on function',
            'physical_exam': 'Degenerative changes\nROM limitations\nNo inflammatory signs',
            'diagnosis': 'Weight-bearing imaging\nTypical findings: [list]\nNormal inflammatory markers',
            'treatment': '🟢 NSAIDs, activity modification\nPT for strengthening\nSurgery if failed conservative',
            'education': 'Progressive but manageable\nWeight management\nLow-impact exercise'
        },

        # === CHRONIC INFLAMMATORY (Sky Blue) ===
        {
            'condition': 'Chronic Inflammatory Condition',
            'category': 'Chronic Inflammatory',
            'presentation': 'Chronic with inflammatory features\n⭐ Morning stiffness >1 hour\nSystemic symptoms possible',
            'physical_exam': 'Inflammatory signs (warmth, swelling)\nSymmetric involvement\nExtra-articular manifestations',
            'diagnosis': '⭐ Specific labs (RF, anti-CCP, etc.)\nElevated inflammatory markers\nImaging for erosions',
            'treatment': '🟢 DMARDs (disease-modifying)\nReferral to specialist\nMultidisciplinary care',
            'education': 'Autoimmune/inflammatory nature\n❗ Early treatment prevents damage\nLifelong management'
        },
    ]

    current_row = 2
    for condition in conditions:
        onset = condition['category']
        bg_color = get_onset_color(onset)

        apply_cell_style(ws.cell(current_row, 1), condition['condition'], bold=True, font_size=10, bg_color=bg_color)
        apply_cell_style(ws.cell(current_row, 2), condition['category'], font_size=10, bg_color=bg_color)
        apply_cell_style(ws.cell(current_row, 3), condition['presentation'], font_size=10, bg_color=bg_color)
        apply_cell_style(ws.cell(current_row, 4), condition['physical_exam'], font_size=10, bg_color=bg_color)
        apply_cell_style(ws.cell(current_row, 5), condition['diagnosis'], font_size=10, bg_color=bg_color)
        apply_cell_style(ws.cell(current_row, 6), condition['treatment'], font_size=10, bg_color=bg_color)
        apply_cell_style(ws.cell(current_row, 7), condition['education'], font_size=10, bg_color=bg_color)

        ws.row_dimensions[current_row].height = 80
        current_row += 1

    return ws

# =============================================================================
# TAB 3: H&P GUIDE (Comprehensive Clinical Encounter Guide)
# =============================================================================

def create_hp_guide_tab(wb):
    """
    Tab 3: H&P Guide
    - Comprehensive clinical encounter guide
    - OLDCAARTS history taking
    - PMH/FH/SH sections
    - Review of Systems (ROS)
    - Physical Exam with verbalizations (Inspection, Palpation, ROM, Neuro, Special Tests)
    - Patient Education (Disease Process, Management, Lifestyle, Medication Risks)
    - Red Flags / When to Return
    """
    ws = wb.create_sheet("H&P Guide")

    set_column_widths(ws, {
        'A': 35,   # Component/Label
        'B': 85,   # Verbalization/Instruction
    })

    current_row = 1

    # Title - Apply styling to ALL cells BEFORE merging
    for col_idx in range(1, 3):  # Columns 1, 2
        cell = ws.cell(current_row, col_idx)
        apply_cell_style(cell, text="FOCUSED H&P PROTOCOL - Physical Exam Verbalizations" if col_idx==1 else '',
                        bold=True, font_size=16, bg_color=MAIN_TITLE_BG,
                        alignment='center', font_color='FFFFFF')

    # Now merge cells
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.row_dimensions[current_row].height = 35
    current_row += 2

    # Helper function for protocol sections
    def add_protocol_section(title, row):
        # Apply styling to ALL cells BEFORE merging
        for col_idx in range(1, 3):  # Columns 1, 2
            cell = ws.cell(row, col_idx)
            apply_cell_style(cell, text=title if col_idx==1 else '', bold=True, font_size=13,
                            bg_color=HP_SECTION_BG, alignment='center', font_color='FFFFFF')

        # Now merge cells
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 28
        return row + 1

    def add_protocol_row(label, content, row, is_verbalization=False, is_red_flag=False):
        cell_a = ws.cell(row, 1)
        cell_b = ws.cell(row, 2)

        if is_red_flag:
            apply_cell_style(cell_a, text=label, bold=True, font_size=11,
                           bg_color=HP_RED_FLAG_BG, font_color='FFFFFF')
            apply_cell_style(cell_b, text=content, font_size=11,
                           bg_color=HP_RED_FLAG_BG, font_color='FFFFFF')
        elif is_verbalization:
            apply_cell_style(cell_a, text=label, bold=True, font_size=11, bg_color=HP_VERBALIZATION_BG)
            apply_cell_style(cell_b, text=f'"{content}"', font_size=11, bg_color=HP_VERBALIZATION_BG)
        else:
            apply_cell_style(cell_a, text=label, bold=True, font_size=11, bg_color=HP_SUBSECTION_BG)
            apply_cell_style(cell_b, text=content, font_size=11, bg_color=HP_SUBSECTION_BG)

        ws.row_dimensions[row].height = max(30, len(content.split('\n')) * 18)
        return row + 1

    # =========================================================================
    # SPINE - INSPECTION
    # =========================================================================
    current_row = add_protocol_section("SPINE - INSPECTION", current_row)

    current_row = add_protocol_row("Patient Position",
        "Patient standing, back to examiner with gown open", current_row)

    current_row = add_protocol_row("VERBALIZATION",
        "I am inspecting for spinal alignment, lateral curvature, and shoulder and iliac crest symmetry.",
        current_row, is_verbalization=True)

    current_row = add_protocol_row("Items to Inspect",
        "• Skin: lesions, ecchymosis, rashes\n• Spine: alignment, lateral curvature, hyper/hypolordosis, kyphosis\n• Shoulders: asymmetry\n• Iliac crests: asymmetry\n• Forward bend test: abnormal spinal curvature",
        current_row)
    current_row += 1

    # =========================================================================
    # SPINE - PALPATION
    # =========================================================================
    current_row = add_protocol_section("SPINE - PALPATION", current_row)

    current_row = add_protocol_row("VERBALIZATION",
        "I am palpating the cervical spine, thoracic spine, lumbar spine, and paraspinal musculature for tenderness and deformity.",
        current_row, is_verbalization=True)

    current_row = add_protocol_row("Technique",
        "Palpate vertebral processes and spinal musculature starting at base of skull to top of sacrum", current_row)

    current_row = add_protocol_row("What to Palpate",
        "• Cervical spine (C1-C7)\n• Thoracic spine (T1-T12)\n• Lumbar spine (L1-L5)\n• Paraspinal musculature",
        current_row)
    current_row += 1

    # =========================================================================
    # SPINE - RANGE OF MOTION
    # =========================================================================
    current_row = add_protocol_section("SPINE - RANGE OF MOTION", current_row)

    current_row = add_protocol_row("VERBALIZATION",
        "I am assessing range of motion of the cervical and thoracolumbar spine.",
        current_row, is_verbalization=True)

    current_row = add_protocol_row("CERVICAL SPINE",
        "• Flexion: \"Bend head forward with chin to chest\"\n• Extension: \"Bend head backward\"\n• Lateral bending: \"Bend head to one side, trying to touch ear to shoulder\"\n• Rotation: \"Rotate head to look over shoulder while keeping rest of spine straight\"",
        current_row)

    current_row = add_protocol_row("THORACOLUMBAR SPINE",
        "(stabilize pelvis)\n• Flexion: \"Bend forward at the hips\"\n• Extension: \"Bend backward at the hips\"\n• Lateral bending: \"Bend to the side like picking up a suitcase\"\n• Rotation: \"Rotate to look behind you using your entire back\"",
        current_row)
    current_row += 1

    # =========================================================================
    # LOWER EXTREMITY - INSPECTION
    # =========================================================================
    current_row = add_protocol_section("LOWER EXTREMITY - INSPECTION", current_row)

    current_row = add_protocol_row("VERBALIZATION",
        "I am assessing the hips, knees, ankles and toes for asymmetry, malalignment, deformities, leg length discrepancy, varicosities, swelling, discoloration, and hair distribution.",
        current_row, is_verbalization=True)
    current_row += 1

    # =========================================================================
    # LOWER EXTREMITY - PALPATION
    # =========================================================================
    current_row = add_protocol_section("LOWER EXTREMITY - PALPATION", current_row)

    le_palpation = [
        ("Pelvis/Hips", "I am palpating the anterior superior iliac spine and greater trochanter."),
        ("Upper Leg", "I am palpating the quadriceps and hamstrings."),
        ("Knee", "I am palpating the patella and the medial and lateral joint lines."),
        ("Knee (tendons)", "I am palpating the patellar tendon and the quadriceps tendon."),
        ("Popliteal fossa", "I am palpating the popliteal fossa for the popliteal pulse, swelling, aneurysm, cyst, and lymphadenopathy."),
        ("Lower Leg", "I am palpating the tibia and fibula."),
        ("Gastrocnemius", "I am palpating the gastrocnemius for palpable cords and asymmetry."),
        ("Soft tissues", "I am palpating the lower extremity for edema."),
        ("Ankle", "I am palpating the Achilles tendon and calcaneus. / I am palpating the medial and lateral malleoli."),
        ("Foot", "I am palpating the metatarsals and phalanges. / I am palpating the MTP joints and IP joints."),
    ]

    for landmark, verbalization in le_palpation:
        current_row = add_protocol_row(landmark, verbalization, current_row, is_verbalization=True)
    current_row += 1

    # =========================================================================
    # NEUROLOGIC EXAM
    # =========================================================================
    current_row = add_protocol_section("NEUROLOGIC EXAM", current_row)

    current_row = add_protocol_row("GAIT - Verbalization",
        "I am observing the patient's gait for asymmetry, weakness and foot drop.",
        current_row, is_verbalization=True)

    current_row = add_protocol_row("Gait Testing",
        "• Casual walk: Observe gait and posture\n• Walk on heels: Assesses L4-L5 (dorsiflexion)\n• Walk on toes: Assesses S1 (plantar flexion)\n• Heel-to-toe (tandem): Balance assessment",
        current_row)

    current_row = add_protocol_row("STRENGTH - Verbalization",
        "I am assessing strength bilaterally.",
        current_row, is_verbalization=True)

    current_row = add_protocol_row("Strength Testing",
        "• Hip flexion (L1-L2): \"Lift your knee up as I try to press it down\"\n• Hip abduction: \"Push your legs outward against my resistance\"\n• Hip adduction: \"Push your legs inward against my resistance\"\n• Knee flexion (L5-S1): \"Pull your leg in against my resistance\"\n• Knee extension (L3-L4): \"Kick your leg out against my resistance\"\n• Ankle dorsiflexion (L4-L5): \"Pull your foot up towards the ceiling against my hand\"\n• Ankle plantar flexion (S1-S2): \"Push down against my hand like your foot is on the accelerator\"",
        current_row)

    current_row = add_protocol_row("REFLEXES",
        "(compare bilaterally, assess for hyper/hyporeflexia)\n• Patellar (L3-L4): Strike patellar tendon, observe knee extension\n• Achilles (S1-S2): Strike Achilles tendon, observe plantar flexion\n• Babinski (UMN): Stroke heel to 5th toe then across ball of foot (makes \"7\")\n  - Normal: flexion of toes\n  - Abnormal (positive): fanning of toes",
        current_row)
    current_row += 1

    # =========================================================================
    # SPECIALIZED TESTS
    # =========================================================================
    current_row = add_protocol_section("SPECIALIZED TESTS", current_row)

    current_row = add_protocol_row("STRAIGHT LEG RAISE (SLR)",
        "With pt supine, lift leg off table until pain elicited. Note angle. May dorsiflex foot.\nSEATED OPTION: Extend knee, evaluate for backward movement, pain, paresthesia into LE.\nPOSITIVE: Radiating pain down leg → lumbosacral/sciatic radiculopathy",
        current_row)

    current_row = add_protocol_row("CROSSED SLR",
        "Perform SLR on asymptomatic leg.\nPOSITIVE: Pain in symptomatic leg when raising contralateral leg = highly specific for disc herniation",
        current_row)

    current_row = add_protocol_row("SPURLING MANEUVER",
        "With pt seated/standing, position pt to rotate and extend head. Apply mild compression at top of head.\nPOSITIVE: Pain/paresthesia into upper extremity → cervical radiculopathy",
        current_row)

    current_row = add_protocol_row("FABER/PATRICK TEST",
        "Pt supine, position leg in figure-four (flexion, abduction, external rotation of hip).\nPlace one hand at contralateral ASIS and other hand on ipsilateral knee.\nPOSITIVE: Reproduced pain → hip or SI joint pathology",
        current_row)
    current_row += 1

    # =========================================================================
    # SENSORY DERMATOMES
    # =========================================================================
    current_row = add_protocol_section("SENSORY DERMATOMES", current_row)

    current_row = add_protocol_row("VERBALIZATION",
        "Close your eyes. Tell me when you feel me touch you and if it feels sharp or dull.",
        current_row, is_verbalization=True)

    current_row = add_protocol_row("Dermatome Distributions",
        "• L4: Knee, medial shin (anterior leg)\n• L5: Anterolateral shin, dorsal foot, great toe\n• S1: Posterior leg, lateral/plantar foot\n• Saddle area (S2-S4): EMERGENCY - perineal sensory changes suggest Cauda Equina Syndrome",
        current_row)
    current_row += 1

    # =========================================================================
    # RED FLAGS - EMERGENCY SIGNS
    # =========================================================================
    current_row = add_protocol_section("RED FLAGS - EMERGENCY SIGNS", current_row)

    current_row = add_protocol_row("⚠️ Cauda Equina Syndrome",
        "Saddle paresthesia, bowel/bladder dysfunction, bilateral LE weakness, decreased anal sphincter tone",
        current_row, is_red_flag=True)

    current_row = add_protocol_row("⚠️ Myelopathy (UMN signs)",
        "Hyperreflexia, Babinski (up-going toes), clonus, hypertonia/spasticity, Lhermitte sign",
        current_row, is_red_flag=True)

    current_row = add_protocol_row("⚠️ Infection/Cancer",
        "Fever, night sweats, unexplained weight loss, immunosuppression, IV drug use, recent infection",
        current_row, is_red_flag=True)

    return ws

# =============================================================================
# TAB 4: SUMMARY
# =============================================================================

def create_summary_tab(wb):
    """
    Tab 4: Summary
    - Mnemonics with full breakdowns
    - "If X Think Y" clinical associations
    - RED FLAGS section (required for clinical charts)
    - High-yield pearls
    """
    ws = wb.create_sheet("Summary")

    set_column_widths(ws, {'A': 100})

    current_row = 1

    # =========================================================================
    # Section 1: MNEMONICS
    # =========================================================================

    add_section_header(ws, current_row, "MNEMONICS")
    current_row += 2

    mnemonics = [
        ('Example Mnemonic', 'E = Example meaning\nX = eXplanation\nA = Another point\nM = Memory aid'),
        ('Clinical Memory Trick', 'Phrase or acronym that helps remember key finding'),
    ]

    for mnemonic_name, mnemonic_content in mnemonics:
        cell = ws.cell(current_row, 1)
        full_text = f'MNEMONIC: "{mnemonic_name}"\n\n{mnemonic_content}'
        apply_cell_style(cell, text=full_text, bg_color=MNEMONIC_BG)
        ws.row_dimensions[current_row].height = 80
        current_row += 1

    current_row += 2

    # =========================================================================
    # Section 2: IF X THINK Y
    # =========================================================================

    add_section_header(ws, current_row, '"IF X THINK Y" CLINICAL ASSOCIATIONS')
    current_row += 2

    associations = [
        "If you see [finding A] → Think [Condition A]",
        "If you see [finding B] + [finding C] → Think [Condition B]",
        "If you see [hallmark finding] → Think [Specific diagnosis]",
        "If you see [emergency sign] → Think [Emergency condition] - URGENT",
    ]

    for assoc in associations:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=assoc, bg_color='F3E5F5')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2

    # =========================================================================
    # Section 3: RED FLAGS (REQUIRED for clinical charts)
    # =========================================================================

    add_section_header(ws, current_row, "RED FLAGS - URGENT EVALUATION")
    current_row += 2

    red_flags = [
        "⚠️ [Red flag finding 1] → Requires immediate evaluation",
        "⚠️ [Red flag finding 2] → Urgent referral needed",
        "⚠️ [Red flag finding 3] → Rule out [serious condition]",
        "⚠️ Progressive neurologic deficit → Emergency imaging",
        "⚠️ Constitutional symptoms (fever, weight loss) → Workup for malignancy/infection",
    ]

    for flag in red_flags:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=flag, bg_color=RED_FLAG_BG)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2

    # =========================================================================
    # Section 4: HIGH-YIELD PEARLS
    # =========================================================================

    add_section_header(ws, current_row, "HIGH-YIELD PEARLS")
    current_row += 2

    pearls = [
        "• [Key differentiating feature between similar conditions]",
        "• [First-line treatment that's commonly tested]",
        "• [Hallmark finding that's pathognomonic]",
        "• [Common exam question or clinical decision point]",
        "• [Important timing or dosing information]",
    ]

    for pearl in pearls:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=pearl, bg_color=CLINICAL_PEARL_BG)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    return ws

# =============================================================================
# MAIN FUNCTION
# =============================================================================

def create_clinical_assessment_chart(output_path):
    """
    Create complete 4-tab clinical assessment chart

    Args:
        output_path: Path to save the Excel file

    Tabs:
        1. Key Comparisons - clinical differentials side-by-side
        2. Master Chart - 7 clinical columns, onset-based grouping
        3. H&P Guide - comprehensive (OLDCAARTS, PMH/FH/SH, ROS, PE, Patient Education, Red Flags)
        4. Summary - mnemonics, "If X Think Y", red flags, pearls
    """
    wb = Workbook()

    # Create all 4 tabs
    create_key_comparisons_tab(wb)
    create_master_chart_tab(wb)
    create_hp_guide_tab(wb)
    create_summary_tab(wb)

    # Save
    wb.save(output_path)
    print(f"Clinical assessment chart created: {output_path}")

if __name__ == '__main__':
    # Example usage
    create_clinical_assessment_chart('Example_Clinical_Assessment_Chart.xlsx')
