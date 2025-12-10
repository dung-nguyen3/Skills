#!/usr/bin/env python3
"""
COMPLETE EXCEL COMPARISON CHART EXAMPLE
Reference implementation for creating comprehensive 3-tab comparison charts

This file shows a COMPLETE working example with all 3 tabs.
Use for ANY medical topic comparisons (conditions, mechanisms, concepts).

Structure:
- Tab 1: Key Comparisons (multiple side-by-side tables, one category per table)
- Tab 2: Master Chart (all items in one comprehensive table)
- Tab 3: Summary (mnemonics, "If X Think Y", high-yield pearls)

Category Selection:
Categories are selected from comprehensive menus (A-M) in the template based on
source content analysis. This example uses Menu B (Medical Conditions) categories:
- Mechanism comparison (Table 1)
- Clinical presentation comparison (Table 2)
- Treatment comparison (Table 3)

Categories are flexible - use menus in Excel_Comparison_Chart_REVISED.txt to
select relevant categories for any medical topic (diagnostic tools, drugs,
conditions, mechanisms, anatomy, procedures, etc.).

Example: Hypersensitivity Reactions Type I-IV
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# COLOR SCHEME - 3-Shade System (See Excel_Color_Reference.txt)
# =============================================================================

MAIN_TITLE_BG = '4472C4'  # Dark blue - ONLY for sheet titles
CLINICAL_PEARL_BG = 'E8F5E9'  # Light green for clinical pearls

# Color Set 0: Ice Blue
ICE_BLUE_HEADER = 'B4C6E7'
ICE_BLUE_MAIN = 'D9E2F3'
ICE_BLUE_ROW_LABEL = 'C5D3ED'

# Color Set 1: Seafoam
SEAFOAM_HEADER = 'A8CCA8'
SEAFOAM_MAIN = 'C8E6C9'
SEAFOAM_ROW_LABEL = 'B8D9B9'

# Color Set 2: Light Orchid
LIGHT_ORCHID_HEADER = 'B8A4D0'
LIGHT_ORCHID_MAIN = 'D1C4E9'
LIGHT_ORCHID_ROW_LABEL = 'C4B4DC'

# Color Set 3: Champagne
CHAMPAGNE_HEADER = 'E0D0B0'
CHAMPAGNE_MAIN = 'F7E7CE'
CHAMPAGNE_ROW_LABEL = 'EBDBBF'

# Color Set 4: Sky Blue
SKY_BLUE_HEADER = '9DC3E6'
SKY_BLUE_MAIN = 'BDD7EE'
SKY_BLUE_ROW_LABEL = 'AECDEA'

# Color Set 5: Pale Azure
PALE_AZURE_HEADER = 'D0E8FF'
PALE_AZURE_MAIN = 'F0F8FF'
PALE_AZURE_ROW_LABEL = 'E0F0FF'

# Color Set 6: Blush Pink
BLUSH_PINK_HEADER = 'E8C4CC'
BLUSH_PINK_MAIN = 'FCE4EC'
BLUSH_PINK_ROW_LABEL = 'F2D4DC'

# Color Set 7: Soft Lilac
SOFT_LILAC_HEADER = 'D0C8DC'
SOFT_LILAC_MAIN = 'EDE7F6'
SOFT_LILAC_ROW_LABEL = 'DED7E9'

# Color Set 8: Soft Tangerine
SOFT_TANGERINE_HEADER = 'E0C8B0'
SOFT_TANGERINE_MAIN = 'FFE8D6'
SOFT_TANGERINE_ROW_LABEL = 'EFD8C3'

# Color Set 9: Powder Blue
POWDER_BLUE_HEADER = 'A0C4E8'
POWDER_BLUE_MAIN = 'BBDEFB'
POWDER_BLUE_ROW_LABEL = 'ADD1F1'

COLOR_SETS = [
    {'header': ICE_BLUE_HEADER, 'main': ICE_BLUE_MAIN, 'row_label': ICE_BLUE_ROW_LABEL},
    {'header': SEAFOAM_HEADER, 'main': SEAFOAM_MAIN, 'row_label': SEAFOAM_ROW_LABEL},
    {'header': LIGHT_ORCHID_HEADER, 'main': LIGHT_ORCHID_MAIN, 'row_label': LIGHT_ORCHID_ROW_LABEL},
    {'header': CHAMPAGNE_HEADER, 'main': CHAMPAGNE_MAIN, 'row_label': CHAMPAGNE_ROW_LABEL},
    {'header': SKY_BLUE_HEADER, 'main': SKY_BLUE_MAIN, 'row_label': SKY_BLUE_ROW_LABEL},
    {'header': PALE_AZURE_HEADER, 'main': PALE_AZURE_MAIN, 'row_label': PALE_AZURE_ROW_LABEL},
    {'header': BLUSH_PINK_HEADER, 'main': BLUSH_PINK_MAIN, 'row_label': BLUSH_PINK_ROW_LABEL},
    {'header': SOFT_LILAC_HEADER, 'main': SOFT_LILAC_MAIN, 'row_label': SOFT_LILAC_ROW_LABEL},
    {'header': SOFT_TANGERINE_HEADER, 'main': SOFT_TANGERINE_MAIN, 'row_label': SOFT_TANGERINE_ROW_LABEL},
    {'header': POWDER_BLUE_HEADER, 'main': POWDER_BLUE_MAIN, 'row_label': POWDER_BLUE_ROW_LABEL},
]

def get_color_set(index):
    """Get color set by index (rotates through available sets)"""
    return COLOR_SETS[index % len(COLOR_SETS)]

# COLOR RULE: Same table/category = same color set
# - Table title row: colors['header']
# - Column headers: colors['main'] (lighter than header)
# - Data rows: colors['main'] (same as column headers)
# Rotate to next color SET only when starting a NEW table/category

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def apply_cell_style(cell, text='', bold=False, font_size=10, bg_color=None,
                     border=True, alignment='left', wrap=True, font_color='000000'):
    """Apply comprehensive cell styling"""
    cell.value = text
    cell.font = Font(name='Calibri', size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical='top',
        wrap_text=wrap
    )

    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    if border:
        cell.border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

def create_comparison_header(ws, title, row, span_cols=5, table_index=0):
    """Create merged title row for a comparison table

    Args:
        table_index: Index for color selection (uses COLOR_SETS[table_index]['header'])
    """
    colors = get_color_set(table_index)
    # Apply styling to ALL cells BEFORE merging
    for col_idx in range(1, span_cols + 1):
        cell = ws.cell(row, col_idx)
        apply_cell_style(cell, text=title if col_idx==1 else '', bold=True, font_size=14,
                        bg_color=colors['header'], alignment='center', font_color='000000')
        cell.font = Font(name='Calibri', size=14, bold=True, color='000000')

    # Now merge cells
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 25

def create_column_headers(ws, headers, row, start_col=1, table_index=0):
    """Create column headers for comparison table

    Args:
        table_index: Index for color selection (uses COLOR_SETS[table_index]['main'])
    """
    colors = get_color_set(table_index)
    header_color = colors['main']
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row, col_idx)
        apply_cell_style(cell, text=header, bold=True, font_size=11,
                        bg_color=header_color, alignment='center', font_color='000000')
    ws.row_dimensions[row].height = 25

def create_master_header_row(ws, headers, row=1):
    """Create formatted header row with freeze for Master Chart

    Uses dark blue (#4472C4) with white text - NOT the two-shade system.
    Master Chart header is always dark blue, data rows use MAIN_COLORS.
    """
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx)
        apply_cell_style(
            cell, text=header, bold=True, font_size=12,
            bg_color=MAIN_TITLE_BG, alignment='center', font_color='FFFFFF'
        )
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')

    ws.row_dimensions[row].height = 25
    ws.freeze_panes = f'A{row+1}'

def set_column_widths(ws, widths):
    """Set column widths from dictionary"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def add_section_header(ws, row, title):
    """Add section header in Summary tab

    Uses MAIN_TITLE_BG (dark blue) for sheet-level section headers
    """
    cell = ws.cell(row, 1)
    apply_cell_style(cell, text=title, bold=True, font_size=14,
                    bg_color=MAIN_TITLE_BG, alignment='center', font_color='FFFFFF')
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws.row_dimensions[row].height = 30

# =============================================================================
# TAB 1: KEY COMPARISONS
# =============================================================================

def create_key_comparisons_tab(wb):
    """
    Tab 1: Key Comparisons
    - Multiple side-by-side comparison tables
    - One category per table (e.g., separate tables for Mechanism, Clinical, Treatment)
    - Mnemonics placed directly below relevant tables
    """
    ws = wb.active
    ws.title = "Key Comparisons"

    # Column widths
    set_column_widths(ws, {
        'A': 25,  # Category/Row label
        'B': 35,  # Item 1 (e.g., Type I)
        'C': 35,  # Item 2 (e.g., Type II)
        'D': 35,  # Item 3 (e.g., Type III)
        'E': 35,  # Item 4 (e.g., Type IV)
    })

    current_row = 1

    # =========================================================================
    # COMPARISON TABLE 1: Mechanism (uses color set 0 = Ice Blue)
    # =========================================================================
    table_index = 0

    create_comparison_header(ws, "HYPERSENSITIVITY REACTIONS: MECHANISM", current_row,
                            span_cols=5, table_index=table_index)
    current_row += 1

    # Column headers (items being compared)
    headers = ['Category', 'Type I (Immediate)', 'Type II (Cytotoxic)',
               'Type III (Immune Complex)', 'Type IV (Delayed)']
    create_column_headers(ws, headers, current_row, table_index=table_index)
    current_row += 1

    # Comparison data for Mechanism table
    mechanism_data = [
        ('Antibody Involved', 'IgE', 'IgG, IgM', 'IgG, IgM', 'None (T-cell mediated)'),
        ('Time to Reaction', 'Minutes', 'Hours', 'Hours to days', '24-72 hours'),
        ('Mechanism', 'Mast cell degranulation', 'Complement activation, ADCC',
         'Immune complex deposition', 'T-cell activation'),
        ('Mediators', 'Histamine, leukotrienes', 'Complement, NK cells',
         'Complement, neutrophils', 'Cytokines (IFN-γ, TNF)'),
    ]

    # Table 1 uses ONE color SET: HEADER for title, MAIN for headers/data rows
    colors = get_color_set(table_index)
    table_1_main = colors['main']
    for row_data in mechanism_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)  # First column bold
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_1_main)
        current_row += 1

    current_row += 3  # Space before next table

    # =========================================================================
    # COMPARISON TABLE 2: Clinical Presentation (uses color set 1 = Seafoam)
    # =========================================================================
    table_index = 1

    create_comparison_header(ws, "HYPERSENSITIVITY REACTIONS: CLINICAL PRESENTATION", current_row,
                            span_cols=5, table_index=table_index)
    current_row += 1

    create_column_headers(ws, headers, current_row, table_index=table_index)
    current_row += 1

    clinical_data = [
        ('Classic Examples', 'Anaphylaxis, allergic rhinitis, asthma',
         'Hemolytic anemia, Goodpasture', 'SLE, serum sickness, PSGN',
         'Contact dermatitis, TB test'),
        ('Target Organs', 'Skin, airways, GI, cardiovascular', 'Blood cells, basement membrane',
         'Joints, kidneys, skin', 'Skin, organs with granulomas'),
        ('Key Finding', 'Wheal and flare', 'Coombs test positive',
         'Vasculitis, glomerulonephritis', 'Granulomas, indurated skin'),
        ('Severity', 'Can be life-threatening', 'Variable', 'Chronic inflammation', 'Usually localized'),
    ]

    # Table 2 uses ONE color SET: HEADER for title, MAIN for headers/data rows
    colors = get_color_set(table_index)
    table_2_main = colors['main']
    for row_data in clinical_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_2_main)
        current_row += 1

    current_row += 3  # Space before next table

    # =========================================================================
    # COMPARISON TABLE 3: Treatment (uses color set 2 = Light Orchid)
    # =========================================================================
    table_index = 2

    create_comparison_header(ws, "HYPERSENSITIVITY REACTIONS: TREATMENT", current_row,
                            span_cols=5, table_index=table_index)
    current_row += 1

    create_column_headers(ws, headers, current_row, table_index=table_index)
    current_row += 1

    treatment_data = [
        ('First-Line Treatment', 'Epinephrine (anaphylaxis), antihistamines',
         'Stop offending agent, supportive', 'Corticosteroids, immunosuppressants',
         'Remove antigen, topical steroids'),
        ('Adjunct Therapy', 'Corticosteroids, bronchodilators', 'Plasmapheresis (severe)',
         'Plasmapheresis', 'Systemic steroids if severe'),
        ('Prevention', 'Allergen avoidance, desensitization', 'Avoid trigger medications',
         'Treat underlying condition', 'Avoid known antigens'),
    ]

    # Table 3 uses ONE color SET: HEADER for title, MAIN for headers/data rows
    colors = get_color_set(table_index)
    table_3_main = colors['main']
    for row_data in treatment_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_3_main)
        current_row += 1

    return ws

# =============================================================================
# TAB 2: MASTER CHART
# =============================================================================

def create_master_chart_tab(wb):
    """
    Tab 2: Master Chart
    - All items in one comprehensive table
    - One row per condition/type
    - Frozen header for scrolling
    """
    ws = wb.create_sheet("Master Chart")

    # Column widths
    set_column_widths(ws, {
        'A': 20,  # Type
        'B': 25,  # Antibody
        'C': 15,  # Onset
        'D': 35,  # Mechanism
        'E': 35,  # Examples
        'F': 35,  # Key Finding
        'G': 35,  # Treatment
    })

    # Headers
    headers = ['Type', 'Antibody', 'Onset', 'Mechanism', 'Examples', 'Key Finding', 'Treatment']
    create_master_header_row(ws, headers, 1)

    # Master chart data (all hypersensitivity types)
    master_data = [
        ('Type I', 'IgE', 'Minutes', 'Mast cell degranulation → histamine release',
         'Anaphylaxis, allergic rhinitis, asthma, urticaria',
         'Wheal and flare, elevated tryptase', 'Epinephrine, antihistamines, steroids'),
        ('Type II', 'IgG, IgM', 'Hours', 'Antibody binds cell surface → complement/ADCC',
         'Hemolytic anemia, Goodpasture, Graves, MG',
         'Direct Coombs test positive', 'Stop trigger, plasmapheresis'),
        ('Type III', 'IgG, IgM', 'Hours-days', 'Immune complex deposition → inflammation',
         'SLE, serum sickness, PSGN, RA',
         'Vasculitis, glomerulonephritis', 'Steroids, immunosuppressants'),
        ('Type IV', 'None', '24-72 hrs', 'T-cell mediated → cytokine release',
         'Contact dermatitis, TB skin test, transplant rejection',
         'Granulomas, induration', 'Remove antigen, steroids'),
    ]

    current_row = 2
    # Master Chart: Category-based coloring (same category = same color)
    # In this example, each row is a different category (Type I-IV)
    # Data rows use font size 10 (not 11) per template spec
    category_index = -1  # Start at -1 so first category becomes 0
    prev_category = None

    for row_data in master_data:
        category = row_data[0]  # First column is the category

        # Change color when category changes
        if category != prev_category:
            category_index += 1
            prev_category = category

        # Get color set and use 'main' shade for data cells
        colors = get_color_set(category_index)
        color = colors['main']

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)  # First column bold
            apply_cell_style(cell, text=value, bold=bold, font_size=10, bg_color=color)
        ws.row_dimensions[current_row].height = 50  # Taller rows for content
        current_row += 1

    return ws

# =============================================================================
# TAB 3: SUMMARY
# =============================================================================

def create_summary_tab(wb):
    """
    Tab 3: Summary
    - Mnemonics with full breakdowns
    - "If X Think Y" associations
    - Critical values
    - Key definitions
    - High-yield pearls
    """
    ws = wb.create_sheet("Summary")

    # Column width - wide for content
    set_column_widths(ws, {
        'A': 100,
    })

    current_row = 1

    # =========================================================================
    # Section 1: IF X THINK Y
    # =========================================================================

    add_section_header(ws, current_row, '"IF X THINK Y" ASSOCIATIONS')
    current_row += 1

    associations = [
        "If immediate reaction after drug/food/sting → Think Type I hypersensitivity",
        "If hemolytic anemia with positive Coombs → Think Type II hypersensitivity",
        "If arthritis + nephritis + rash → Think Type III (immune complex disease)",
        "If contact dermatitis or PPD+ → Think Type IV (delayed/T-cell)",
        "If needs epinephrine → Think Type I anaphylaxis",
        "If granulomas on biopsy → Think Type IV hypersensitivity",
    ]

    for assoc in associations:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=assoc, bg_color='F3E5F5')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2

    # =========================================================================
    # Section 2: KEY DEFINITIONS
    # =========================================================================

    add_section_header(ws, current_row, "KEY DEFINITIONS")
    current_row += 1

    definitions = [
        "Anaphylaxis: Severe, life-threatening systemic Type I reaction",
        "ADCC: Antibody-dependent cellular cytotoxicity (Type II mechanism)",
        "Immune complex: Antigen-antibody aggregate that deposits in tissues (Type III)",
        "Granuloma: Collection of macrophages/giant cells (Type IV hallmark)",
        "Arthus reaction: Localized Type III reaction at injection site",
        "Serum sickness: Systemic Type III reaction (fever, rash, arthralgia, lymphadenopathy)",
    ]

    for defn in definitions:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=defn, bg_color='E8F5E9')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2

    # =========================================================================
    # Section 3: HIGH-YIELD PEARLS
    # =========================================================================

    add_section_header(ws, current_row, "HIGH-YIELD PEARLS")
    current_row += 1

    pearls = [
        "• Type I is the ONLY type that involves IgE",
        "• Type IV is the ONLY type that is antibody-INDEPENDENT (T-cell mediated)",
        "• Type II and III both involve IgG/IgM but differ by target (cell surface vs soluble)",
        "• Epinephrine is FIRST-LINE for anaphylaxis - don't delay!",
        "• PPD test reads at 48-72 hours because Type IV is DELAYED",
        "• Goodpasture = Type II (anti-basement membrane antibodies)",
        "• SLE = Type III (immune complex deposition in multiple organs)",
        "• Contact dermatitis = Type IV (think poison ivy)",
    ]

    for pearl in pearls:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=pearl, bg_color='E0F2F1')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    return ws

# =============================================================================
# MAIN FUNCTION
# =============================================================================

def create_comparison_chart(output_path):
    """
    Create complete 3-tab comparison chart

    Args:
        output_path: Path to save the Excel file
    """
    # Create workbook
    wb = Workbook()

    # Create all tabs
    create_key_comparisons_tab(wb)
    create_master_chart_tab(wb)
    create_summary_tab(wb)

    # Save
    wb.save(output_path)
    print(f"Comparison chart created: {output_path}")

if __name__ == '__main__':
    # Example usage
    create_comparison_chart('Hypersensitivity_Comparison_Chart.xlsx')
