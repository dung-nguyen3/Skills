#!/usr/bin/env python3
"""
Generate Excel file with pre-formatted drug class chart templates.
Each chart has proper colors, fonts, and sizes ready to copy/paste.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_drug_class_templates():
    wb = Workbook()

    # Define all 10 color sets (matching Excel Drugs Chart 11-1 template.py)
    COLOR_SETS = [
        {
            'name': 'Ice Blue',
            'header': 'B4C6E7',
            'row_label': 'C5D3ED',
            'main': 'D9E2F3'
        },
        {
            'name': 'Seafoam',
            'header': 'A8CCA8',
            'row_label': 'B8D9B9',
            'main': 'C8E6C9'
        },
        {
            'name': 'Light Orchid',
            'header': 'B8A4D0',
            'row_label': 'C4B4DC',
            'main': 'D1C4E9'
        },
        {
            'name': 'Champagne',
            'header': 'E0D0B0',
            'row_label': 'EBDBBF',
            'main': 'F7E7CE'
        },
        {
            'name': 'Sky Blue',
            'header': '9DC3E6',
            'row_label': 'AECDEA',
            'main': 'BDD7EE'
        },
        {
            'name': 'Pale Azure',
            'header': 'D0E8FF',
            'row_label': 'E0F0FF',
            'main': 'F0F8FF'
        },
        {
            'name': 'Blush Pink',
            'header': 'E8C4CC',
            'row_label': 'F0D4DA',
            'main': 'FCE4EC'
        },
        {
            'name': 'Soft Lilac',
            'header': 'C8C0E0',
            'row_label': 'D8D4EB',
            'main': 'EDE7F6'
        },
        {
            'name': 'Soft Tangerine',
            'header': 'E8C8A8',
            'row_label': 'F0D8BF',
            'main': 'FFE8D6'
        },
        {
            'name': 'Powder Blue',
            'header': '9BBEE0',
            'row_label': 'ABCEEA',
            'main': 'BBDEFB'
        }
    ]

    # White border (invisible)
    white_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )

    # Create Instructions sheet first
    ws_instructions = wb.active
    ws_instructions.title = "HOW TO USE"

    ws_instructions['A1'] = 'DRUG CLASS CHART TEMPLATES - INSTRUCTIONS'
    ws_instructions['A1'].font = Font(size=18, bold=True)
    ws_instructions.merge_cells('A1:F1')

    instructions = [
        '',
        'HOW TO USE THESE TEMPLATES:',
        '',
        '1. Click on the tab for the color scheme you want (Ice Blue, Seafoam, etc.)',
        '2. Select the entire drug class chart (click and drag from A1 to last cell)',
        '3. Copy (Ctrl+C or Cmd+C)',
        '4. Go to your study guide Excel file',
        '5. Click where you want the chart',
        '6. Paste (Ctrl+V or Cmd+V) - all formatting is preserved!',
        '',
        'FORMATTING INCLUDED:',
        '• Header row: Bold, Size 18, centered',
        '• Drug names: Bold, Size 16, centered',
        '• Row labels (first column): Bold, Size 14, centered',
        '• Content cells: Size 12, left-aligned, word wrap enabled',
        '• All cells have proper background colors',
        '• White borders (invisible) between cells',
        '',
        'CUSTOMIZING:',
        '• Add more rows: Right-click row number → Insert',
        '• Add more columns: Right-click column letter → Insert',
        '• Change text: Just type in the cells',
        '• Copy formatting: Use Format Painter (paintbrush icon)',
        '',
        'EACH TAB CONTAINS:',
        '• 3-drug chart template (3 columns)',
        '• 2-drug chart template (2 columns)',
        '• Single drug chart template (1 column)',
        '',
        'COLOR SCHEMES (in order):',
        '1. Ice Blue - General topics',
        '2. Seafoam - Normal findings',
        '3. Light Orchid - Special topics',
        '4. Champagne - Warnings/cautions',
        '5. Sky Blue - Diagnostic',
        '6. Pale Azure - Alternative blue',
        '7. Blush Pink - Important alerts',
        '8. Soft Lilac - Alternative purple',
        '9. Soft Tangerine - Highlights',
        '10. Powder Blue - Alternative blue'
    ]

    for i, text in enumerate(instructions, start=2):
        ws_instructions.cell(row=i, column=1, value=text)
        if text.startswith('HOW TO USE') or text.startswith('FORMATTING') or text.startswith('CUSTOMIZING') or text.startswith('EACH TAB') or text.startswith('COLOR SCHEMES'):
            ws_instructions.cell(row=i, column=1).font = Font(size=14, bold=True)
        else:
            ws_instructions.cell(row=i, column=1).font = Font(size=12)

    ws_instructions.column_dimensions['A'].width = 80

    # Create a sheet for each color scheme
    for color_set in COLOR_SETS:
        ws = wb.create_sheet(title=color_set['name'])

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 35

        current_row = 1

        # ===== 3-DRUG CHART TEMPLATE =====
        ws.cell(row=current_row, column=1, value='3-DRUG CHART TEMPLATE')
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color='FF0000')
        current_row += 1

        # Header row (Drug Class Name)
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col, value='DRUG CLASS NAME' if col == 1 else '')
            cell.fill = PatternFill(start_color=color_set['header'], end_color=color_set['header'], fill_type='solid')
            cell.font = Font(size=18, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = white_border
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Drug names row
        drug_names = ['', 'Drug Name 1', 'Drug Name 2', 'Drug Name 3']
        for col, name in enumerate(drug_names, start=1):
            cell = ws.cell(row=current_row, column=col, value=name)
            cell.fill = PatternFill(start_color=color_set['row_label'], end_color=color_set['row_label'], fill_type='solid')
            cell.font = Font(size=16, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = white_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Data rows
        row_labels = ['MOA', 'Indications', 'Adverse Effects', 'Contraindications', 'Special Notes']
        for label in row_labels:
            # Row label (first column)
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.fill = PatternFill(start_color=color_set['row_label'], end_color=color_set['row_label'], fill_type='solid')
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = white_border

            # Content cells
            for col in range(2, 5):
                cell = ws.cell(row=current_row, column=col, value='')
                cell.fill = PatternFill(start_color=color_set['main'], end_color=color_set['main'], fill_type='solid')
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = white_border

            ws.row_dimensions[current_row].height = 60
            current_row += 1

        current_row += 2

        # ===== 2-DRUG CHART TEMPLATE =====
        ws.cell(row=current_row, column=1, value='2-DRUG CHART TEMPLATE')
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color='FF0000')
        current_row += 1

        # Header row
        for col in range(1, 4):
            cell = ws.cell(row=current_row, column=col, value='DRUG CLASS NAME' if col == 1 else '')
            cell.fill = PatternFill(start_color=color_set['header'], end_color=color_set['header'], fill_type='solid')
            cell.font = Font(size=18, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = white_border
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Drug names row
        drug_names = ['', 'Drug Name 1', 'Drug Name 2']
        for col, name in enumerate(drug_names, start=1):
            cell = ws.cell(row=current_row, column=col, value=name)
            cell.fill = PatternFill(start_color=color_set['row_label'], end_color=color_set['row_label'], fill_type='solid')
            cell.font = Font(size=16, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = white_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Data rows
        for label in row_labels:
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.fill = PatternFill(start_color=color_set['row_label'], end_color=color_set['row_label'], fill_type='solid')
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = white_border

            for col in range(2, 4):
                cell = ws.cell(row=current_row, column=col, value='')
                cell.fill = PatternFill(start_color=color_set['main'], end_color=color_set['main'], fill_type='solid')
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = white_border

            ws.row_dimensions[current_row].height = 60
            current_row += 1

        current_row += 2

        # ===== SINGLE DRUG CHART TEMPLATE =====
        ws.cell(row=current_row, column=1, value='SINGLE DRUG CHART TEMPLATE')
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True, color='FF0000')
        current_row += 1

        # Header row
        for col in range(1, 3):
            cell = ws.cell(row=current_row, column=col, value='DRUG CLASS NAME' if col == 1 else '')
            cell.fill = PatternFill(start_color=color_set['header'], end_color=color_set['header'], fill_type='solid')
            cell.font = Font(size=18, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = white_border
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Drug name row
        cell = ws.cell(row=current_row, column=1, value='')
        cell.fill = PatternFill(start_color=color_set['row_label'], end_color=color_set['row_label'], fill_type='solid')
        cell.border = white_border

        cell = ws.cell(row=current_row, column=2, value='Drug Name')
        cell.fill = PatternFill(start_color=color_set['row_label'], end_color=color_set['row_label'], fill_type='solid')
        cell.font = Font(size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = white_border
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Data rows
        for label in row_labels:
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.fill = PatternFill(start_color=color_set['row_label'], end_color=color_set['row_label'], fill_type='solid')
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = white_border

            cell = ws.cell(row=current_row, column=2, value='')
            cell.fill = PatternFill(start_color=color_set['main'], end_color=color_set['main'], fill_type='solid')
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = white_border

            ws.row_dimensions[current_row].height = 60
            current_row += 1

    # Save the file
    output_path = '/Users/kimnguyen/Documents/Study Guide Claude Q3/Claude Templates/Drug_Class_Chart_Templates.xlsx'
    wb.save(output_path)
    print(f'✅ Drug class chart templates created: {output_path}')
    print(f'\nContains {len(COLOR_SETS)} color schemes with 3 chart sizes each (3-drug, 2-drug, single drug)')
    print('Open the file and copy/paste the pre-formatted charts into your study guides!')

if __name__ == '__main__':
    create_drug_class_templates()
