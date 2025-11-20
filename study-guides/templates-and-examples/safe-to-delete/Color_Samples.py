#!/usr/bin/env python3
"""
Create Excel file with light pastel color samples for visual comparison
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_color_samples():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Samples"

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40

    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = 'LIGHT PASTEL COLOR SAMPLES'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    ws['A2'] = 'Click on colors you like. These are MUCH lighter than current tan/green.'
    ws['A2'].font = Font(size=11, italic=True)
    ws.merge_cells('A2:G2')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Current colors (for comparison)
    current_colors = [
        ('Current Light Blue', 'A8BFD8'),
        ('Current Light Green', 'A8CCBA'),
        ('Current Light Purple', 'C8B8DC'),
        ('Current Light Tan', 'DBC8A0'),
        ('Current Light Teal', 'A8CCCC'),
    ]

    # NEW Light Pastel alternatives
    color_families = {
        'LIGHT BLUES': [
            ('Baby Blue', 'E3F2FD'),
            ('Powder Blue', 'BBDEFB'),
            ('Ice Blue', 'E1F5FE'),
            ('Sky Blue', 'B3E5FC'),
            ('Soft Blue', 'E8F4F8'),
            ('Pale Azure', 'F0F8FF'),
        ],
        'LIGHT GREENS': [
            ('Mint Green', 'E8F5E9'),
            ('Seafoam', 'C8E6C9'),
            ('Pale Sage', 'E8F5E9'),
            ('Light Mint', 'F1F8E9'),
            ('Soft Green', 'DCEDC8'),
            ('Honeydew', 'F0FFF0'),
        ],
        'LIGHT PURPLES': [
            ('Lavender', 'F3E5F5'),
            ('Pale Violet', 'E1BEE7'),
            ('Soft Lilac', 'EDE7F6'),
            ('Light Orchid', 'D1C4E9'),
            ('Pale Purple', 'F8F0FF'),
            ('Wisteria', 'E8E0F0'),
        ],
        'LIGHT PINKS/CORALS': [
            ('Blush Pink', 'FCE4EC'),
            ('Pale Rose', 'F8BBD9'),
            ('Soft Coral', 'FFEBEE'),
            ('Light Salmon', 'FBE9E7'),
            ('Pale Peach', 'FFF3E0'),
            ('Shell Pink', 'FFF0F5'),
        ],
        'LIGHT YELLOWS/CREAMS': [
            ('Pale Yellow', 'FFFDE7'),
            ('Cream', 'FFF8E1'),
            ('Soft Lemon', 'FFFACD'),
            ('Butter', 'FFFFF0'),
            ('Champagne', 'F7E7CE'),
            ('Light Gold', 'FFF9C4'),
        ],
        'LIGHT ORANGES/PEACHES': [
            ('Pale Apricot', 'FFF3E0'),
            ('Light Peach', 'FFCCBC'),
            ('Soft Orange', 'FFE0B2'),
            ('Cream Orange', 'FFF8E1'),
            ('Pale Melon', 'FFECD2'),
            ('Soft Tangerine', 'FFE8D6'),
        ],
    }

    row = 4

    # Show current colors first
    ws.cell(row=row, column=1, value='CURRENT COLORS (for comparison)')
    ws.cell(row=row, column=1).font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    for name, hex_code in current_colors:
        # Color sample cell
        cell = ws.cell(row=row, column=1, value=name)
        cell.fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=11)

        # Hex code
        hex_cell = ws.cell(row=row, column=2, value=f'#{hex_code}')
        hex_cell.border = thin_border
        hex_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Large preview
        preview = ws.cell(row=row, column=3, value='Sample text in this color background')
        preview.fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')
        preview.border = thin_border
        preview.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[row].height = 25
        row += 1

    row += 2

    # Show new lighter pastel options
    ws.cell(row=row, column=1, value='NEW LIGHTER PASTEL OPTIONS')
    ws.cell(row=row, column=1).font = Font(size=14, bold=True, color='FF0000')
    ws.merge_cells(f'A{row}:G{row}')
    row += 2

    # Display color families in two columns
    families_list = list(color_families.items())
    col_offset = 0

    for idx, (family_name, colors) in enumerate(families_list):
        if idx % 2 == 0:
            col_offset = 0
            if idx > 0:
                row += 1
            start_row = row
        else:
            col_offset = 4
            row = start_row

        base_col = 1 + col_offset

        # Family header
        header_cell = ws.cell(row=row, column=base_col, value=family_name)
        header_cell.font = Font(size=12, bold=True)
        ws.merge_cells(f'{get_column_letter(base_col)}{row}:{get_column_letter(base_col+2)}{row}')
        row += 1

        for name, hex_code in colors:
            # Color sample cell
            cell = ws.cell(row=row, column=base_col, value=name)
            cell.fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=10)

            # Hex code
            hex_cell = ws.cell(row=row, column=base_col+1, value=f'#{hex_code}')
            hex_cell.border = thin_border
            hex_cell.alignment = Alignment(horizontal='center', vertical='center')
            hex_cell.font = Font(size=9)

            # Large preview
            preview = ws.cell(row=row, column=base_col+2, value='Sample text here')
            preview.fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')
            preview.border = thin_border
            preview.alignment = Alignment(horizontal='center', vertical='center')
            preview.font = Font(size=10)

            ws.row_dimensions[row].height = 22
            row += 1

        if idx % 2 == 1:
            row = max(row, start_row + len(colors) + 1)

    # Save
    output_path = '/Users/kimnguyen/Documents/Study Guide Claude Q3/Claude Templates/Light_Pastel_Color_Samples.xlsx'
    wb.save(output_path)
    print(f'✅ Color samples saved: {output_path}')

if __name__ == '__main__':
    create_color_samples()
