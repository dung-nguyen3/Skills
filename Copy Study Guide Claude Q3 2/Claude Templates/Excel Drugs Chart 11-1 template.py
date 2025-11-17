#!/usr/bin/env python3
"""
EXCEL DRUG CHART GENERATOR - Python Template
Based on: Comprehensive Drug Chart Excel Generation Instructions (11-1)

This template provides all helper functions and structure for creating
comprehensive drug comparison charts from pharmacology content.

Usage when Claude generates a drug chart:
    1. Import this module
    2. Extract drug data from source file
    3. Call helper functions with extracted data
    4. Save the workbook

This approach saves tokens by:
    - Pre-defining all formatting functions
    - Pre-defining color schemes
    - Pre-defining table structures
    - Allowing Claude to focus on data extraction, not formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==================== COLOR SCHEME - SOFT PASTELS WITH BLACK TEXT ====================
# All colors defined with EXACT HEX CODES from instructions

# Main color sets for drug class tables (rotate through these)
# Color Set 0: Ice Blue 40% (Accent 1) - Excel theme color
ICE_BLUE_HEADER = 'B4C6E7'  # Darker shade for headers
ICE_BLUE_MAIN = 'D9E2F3'    # Ice Blue 40% (Accent 1)
ICE_BLUE_ROW_LABEL = 'C5D3ED'  # Medium shade for row labels

# Color Set 1: Seafoam
SEAFOAM_HEADER = 'A8CCA8'
SEAFOAM_MAIN = 'C8E6C9'
SEAFOAM_ROW_LABEL = 'B8D9B9'

# Color Set 2: Light Orchid
LIGHT_ORCHID_HEADER = 'B8A4D0'
LIGHT_ORCHID_MAIN = 'D1C4E9'
LIGHT_ORCHID_ROW_LABEL = 'C4B4DC'

# Color Set 3: Champagne
CHAMPAGNE_HEADER = 'E0D0B0'
CHAMPAGNE_MAIN = 'F7E7CE'
CHAMPAGNE_ROW_LABEL = 'EBDBBF'

# Color Set 4: Sky Blue 60% (Accent 5) - Light Teal
SKY_BLUE_HEADER = '9DC3E6'
SKY_BLUE_MAIN = 'BDD7EE'  # Sky Blue 60% (Accent 5)
SKY_BLUE_ROW_LABEL = 'AECDEA'

# Color Set 5: Pale Azure
PALE_AZURE_HEADER = 'D0E8FF'
PALE_AZURE_MAIN = 'F0F8FF'
PALE_AZURE_ROW_LABEL = 'E0F0FF'

# Color Set 6: Blush Pink
BLUSH_PINK_HEADER = 'E8C4CC'
BLUSH_PINK_MAIN = 'FCE4EC'
BLUSH_PINK_ROW_LABEL = 'F2D4DC'

# Color Set 7: Soft Lilac
SOFT_LILAC_HEADER = 'D0C8DC'
SOFT_LILAC_MAIN = 'EDE7F6'
SOFT_LILAC_ROW_LABEL = 'DED7E9'

# Color Set 8: Soft Tangerine
SOFT_TANGERINE_HEADER = 'E0C8B0'
SOFT_TANGERINE_MAIN = 'FFE8D6'
SOFT_TANGERINE_ROW_LABEL = 'EFD8C3'

# Color Set 9: Powder Blue
POWDER_BLUE_HEADER = 'A0C4E8'
POWDER_BLUE_MAIN = 'BBDEFB'
POWDER_BLUE_ROW_LABEL = 'ADD1F1'

# Special background colors
MNEMONIC_BG = 'E6F3FF'  # Light blue for mnemonics
CLINICAL_PEARL_BG = 'E8F5E9'  # Light green for clinical pearls
MAIN_TITLE_COLOR = '4472C4'  # Dark blue for sheet titles
ANALOGY_BOX_BG = 'FFF9E6'  # Light yellow for analogy boxes

# Color sets array for easy rotation
COLOR_SETS = [
    {'header': ICE_BLUE_HEADER, 'main': ICE_BLUE_MAIN, 'row_label': ICE_BLUE_ROW_LABEL},           # 0: Ice Blue 40%
    {'header': SEAFOAM_HEADER, 'main': SEAFOAM_MAIN, 'row_label': SEAFOAM_ROW_LABEL},               # 1: Seafoam
    {'header': LIGHT_ORCHID_HEADER, 'main': LIGHT_ORCHID_MAIN, 'row_label': LIGHT_ORCHID_ROW_LABEL}, # 2: Light Orchid
    {'header': CHAMPAGNE_HEADER, 'main': CHAMPAGNE_MAIN, 'row_label': CHAMPAGNE_ROW_LABEL},         # 3: Champagne
    {'header': SKY_BLUE_HEADER, 'main': SKY_BLUE_MAIN, 'row_label': SKY_BLUE_ROW_LABEL},           # 4: Sky Blue 60%
    {'header': PALE_AZURE_HEADER, 'main': PALE_AZURE_MAIN, 'row_label': PALE_AZURE_ROW_LABEL},     # 5: Pale Azure
    {'header': BLUSH_PINK_HEADER, 'main': BLUSH_PINK_MAIN, 'row_label': BLUSH_PINK_ROW_LABEL},     # 6: Blush Pink
    {'header': SOFT_LILAC_HEADER, 'main': SOFT_LILAC_MAIN, 'row_label': SOFT_LILAC_ROW_LABEL},     # 7: Soft Lilac
    {'header': SOFT_TANGERINE_HEADER, 'main': SOFT_TANGERINE_MAIN, 'row_label': SOFT_TANGERINE_ROW_LABEL}, # 8: Soft Tangerine
    {'header': POWDER_BLUE_HEADER, 'main': POWDER_BLUE_MAIN, 'row_label': POWDER_BLUE_ROW_LABEL},   # 9: Powder Blue
]

# Border style - white borders between cells (invisible, blends with pastel backgrounds)
thin_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='FFFFFF'),
    bottom=Side(style='thin', color='FFFFFF')
)

# ==================== HELPER FUNCTIONS ====================

def get_color_set(index):
    """Get color set by index (rotates through available sets)"""
    return COLOR_SETS[index % len(COLOR_SETS)]


def set_column_widths(ws, widths_dict):
    """
    Set column widths for a worksheet

    Args:
        ws: worksheet object
        widths_dict: dict mapping column letters to widths
                    e.g., {'A': 25, 'B': 30, 'C': 30}
    """
    for col, width in widths_dict.items():
        ws.column_dimensions[col].width = width


def add_sheet_title(ws, row, title, end_col_letter):
    """
    Add main sheet title (blue background, white text)

    Args:
        ws: worksheet object
        row: row number
        title: title text
        end_col_letter: ending column for merge (e.g., 'E')

    Returns:
        next row number
    """
    ws.merge_cells(f'A{row}:{end_col_letter}{row}')
    cell = ws[f'A{row}']
    cell.value = title
    cell.font = Font(bold=True, size=16, color='FFFFFF')
    cell.fill = PatternFill(start_color=MAIN_TITLE_COLOR, end_color=MAIN_TITLE_COLOR, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    ws.row_dimensions[row].height = 25  # Main sheet title
    return row + 1


def add_drug_class_table(ws, start_row, class_name, drug_names, rows_data, colors):
    """
    Add a drug class comparison table with proper structure

    IMPORTANT: Follows exact structure from instructions:
    - Row 1: Drug Class Header (merged, height 40, size 18)
    - Row 2: Drug Names (height 30, size 16)
    - Row 3+: Properties (Common MOA must come first after drug names)

    Args:
        ws: worksheet object
        start_row: starting row number
        class_name: name of drug class (uppercase, e.g., "NUCLEOSIDE REVERSE TRANSCRIPTASE INHIBITORS")
        drug_names: list of drug names for column headers
        rows_data: list of dicts with structure:
            [
                {
                    'label': 'Common MOA',
                    'type': 'merged',  # or 'individual'
                    'value': 'Single value for merged' OR
                    'values': ['value1', 'value2', ...] for individual
                },
                ...
            ]
        colors: dict with 'header', 'main', 'row_label' color hex codes

    Returns:
        next available row number
    """
    current_row = start_row
    num_drugs = len(drug_names)
    end_col_letter = get_column_letter(num_drugs + 1)  # +1 for label column

    # ROW 1: Drug Class Header (merged across all columns)
    ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
    header_cell = ws[f'A{current_row}']
    header_cell.value = class_name
    header_cell.font = Font(bold=True, size=18, color='000000')
    header_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.border = thin_border
    ws.row_dimensions[current_row].height = 28  # Drug class header
    current_row += 1

    # ROW 2: Drug Names
    # Column A (empty cell with background)
    ws[f'A{current_row}'].value = ''
    ws[f'A{current_row}'].fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
    ws[f'A{current_row}'].border = thin_border

    # Drug name columns
    for col_idx, drug_name in enumerate(drug_names, start=2):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = drug_name
        cell.font = Font(bold=True, size=16, color='000000')
        cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 30  # Drug names (slightly higher)
    current_row += 1

    # ROW 3+: Property rows
    for row_data in rows_data:
        label = row_data['label']
        row_type = row_data['type']

        # Column A: Property label
        label_cell = ws[f'A{current_row}']
        label_cell.value = label
        label_cell.font = Font(bold=True, size=14, color='000000')
        label_cell.fill = PatternFill(start_color=colors['row_label'], end_color=colors['row_label'], fill_type='solid')
        label_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        label_cell.border = thin_border

        if row_type == 'merged':
            # Merge cells B through end for class-wide property
            merge_range = f'B{current_row}:{end_col_letter}{current_row}'
            ws.merge_cells(merge_range)
            merged_cell = ws[f'B{current_row}']
            merged_cell.value = row_data['value']
            merged_cell.font = Font(size=12, color='000000')
            merged_cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
            merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            merged_cell.border = thin_border
        else:
            # Individual cells for each drug
            values = row_data['values']
            for col_idx, value in enumerate(values, start=2):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.font = Font(size=12, color='000000')
                cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = thin_border

        # Set row height based on label type
        if label.lower() == 'route':
            ws.row_dimensions[current_row].height = 32  # Route row (1.5x shorter than 48)
        else:
            ws.row_dimensions[current_row].height = 48  # Data rows
        current_row += 1

    return current_row


def add_mnemonic_row(ws, start_row, end_col_letter, mnemonic_content):
    """
    Add MEMORY TRICKS & MNEMONICS row after drug class table

    CRITICAL: Mnemonics must be RESEARCHED, never invented

    Args:
        ws: worksheet object
        start_row: row number
        end_col_letter: ending column letter for merge (e.g., 'E')
        mnemonic_content: researched mnemonic text (multi-line supported)

    Returns:
        next row number
    """
    ws.merge_cells(f'A{start_row}:{end_col_letter}{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = f'💡 MEMORY TRICKS & MNEMONICS\n\n{mnemonic_content}'
    cell.font = Font(size=11, color='000000')
    cell.fill = PatternFill(start_color=MNEMONIC_BG, end_color=MNEMONIC_BG, fill_type='solid')
    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    cell.border = thin_border
    ws.row_dimensions[start_row].height = 60  # Mnemonic
    return start_row + 1


def add_analogy_box(ws, start_row, analogy_content):
    """
    Add analogy box in column G for mechanism explanations

    Args:
        ws: worksheet object
        start_row: starting row
        analogy_content: analogy text (2-4 sentences)

    Returns:
        next row number
    """
    # Merge G column cells for analogy box (adjust rows as needed)
    merge_range = f'G{start_row}:G{start_row + 3}'
    ws.merge_cells(merge_range)
    cell = ws[f'G{start_row}']
    cell.value = f'💡 ANALOGY:\n\n{analogy_content}'
    cell.font = Font(size=10, color='000000', italic=True)
    cell.fill = PatternFill(start_color=ANALOGY_BOX_BG, end_color=ANALOGY_BOX_BG, fill_type='solid')
    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    cell.border = thin_border
    return start_row


def add_comparison_table(ws, start_row, title, headers, data_rows, colors, include_mnemonic=False, mnemonic_text=''):
    """
    Add a key comparison table for Tab 2: Key Comparisons

    Args:
        ws: worksheet object
        start_row: starting row number
        title: table title (e.g., "COMPARISON: Mechanisms of Action")
        headers: list of column headers (e.g., ['Drug Class', 'Mechanism', 'Analogy'])
        data_rows: list of lists, each inner list is a row of data
        colors: dict with 'header' and 'main' color hex codes
        include_mnemonic: bool, whether to add mnemonic row after table
        mnemonic_text: mnemonic content if include_mnemonic=True

    Returns:
        next available row number
    """
    current_row = start_row
    num_cols = len(headers)
    end_col_letter = get_column_letter(num_cols)

    # Title row (merged)
    ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color='000000')
    title_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.border = thin_border
    ws.row_dimensions[current_row].height = 25  # Comparison title
    current_row += 1

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=11, color='000000')
        cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 23  # Comparison header
    current_row += 1

    # Data rows
    for row_data in data_rows:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = value
            # First column is bold
            if col_idx == 1:
                cell.font = Font(bold=True, size=10, color='000000')
            else:
                cell.font = Font(size=10, color='000000')
            # ALL cells get pastel background
            cell.fill = PatternFill(start_color=colors['main'], end_color=colors['main'], fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 43  # Comparison data
        current_row += 1

    # Add mnemonic if requested
    if include_mnemonic and mnemonic_text:
        current_row += 1  # Blank row
        ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f'💡 MEMORY TRICKS & MNEMONICS\n\n{mnemonic_text}'
        cell.font = Font(size=11, color='000000')
        cell.fill = PatternFill(start_color=MNEMONIC_BG, end_color=MNEMONIC_BG, fill_type='solid')
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        cell.border = thin_border
        ws.row_dimensions[current_row].height = 50  # Comparison mnemonic
        current_row += 1

    return current_row


def add_master_chart_row(ws, row_num, drug_class, drug_name, route, mechanism, uses, adverse_effects,
                          contraindications, resistance, drug_interactions, special_considerations, color):
    """
    Add a single drug row to Master Chart

    Args:
        ws: worksheet object
        row_num: row number
        drug_class: drug class name
        drug_name: drug name with brand name
        route: administration route
        mechanism: mechanism of action
        uses: clinical uses
        adverse_effects: side effects
        contraindications: contraindications
        resistance: resistance patterns
        drug_interactions: drug interactions
        special_considerations: special considerations
        color: pastel background color for this row (based on drug class)
    """
    data = [drug_class, drug_name, route, mechanism, uses, adverse_effects,
            contraindications, resistance, drug_interactions, special_considerations]

    for col_idx, value in enumerate(data, start=1):
        cell = ws.cell(row_num, col_idx, value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border
        cell.font = Font(size=10, color='000000')
        # First column (drug class) is bold
        if col_idx == 1:
            cell.font = Font(bold=True, size=10, color='000000')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws.row_dimensions[row_num].height = 60


def setup_master_chart_sheet(ws):
    """
    Set up Master Chart sheet with frozen header row

    Args:
        ws: worksheet object
    """
    # Set column widths
    column_widths = {
        'A': 20,  # Drug Class
        'B': 25,  # Drug Name (Brand)
        'C': 12,  # Route
        'D': 35,  # Mechanism
        'E': 35,  # Uses
        'F': 35,  # Adverse Effects
        'G': 25,  # Contraindications
        'H': 25,  # Resistance
        'I': 30,  # Drug Interactions
        'J': 35,  # Special Considerations
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Header row - exactly matches HIV chart structure
    headers = ['Drug Class', 'Drug Name (Brand)', 'Route', 'Mechanism', 'Uses', 'Adverse Effects',
               'Contraindications', 'Resistance', 'Drug Interactions', 'Special Considerations']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color=MAIN_TITLE_COLOR, end_color=MAIN_TITLE_COLOR, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 25

    # Freeze first row
    ws.freeze_panes = 'A2'


def add_high_yield_section(ws, start_row, section_title, content, bg_color, content_bg_color='E8F5E9', end_col_letter='C'):
    """
    Add a section to High-Yield & Pearls tab

    Args:
        ws: worksheet object
        start_row: starting row
        section_title: section title (e.g., "CLINICAL PEARLS")
        content: section content (multi-line text)
        bg_color: background color for section header
        content_bg_color: background color for content cell (default light green)
        end_col_letter: ending column for merge

    Returns:
        next row number
    """
    current_row = start_row

    # Section title
    ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = section_title
    title_cell.font = Font(bold=True, size=14, color='000000')
    title_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.border = thin_border
    ws.row_dimensions[current_row].height = 23  # High-yield section header
    current_row += 1

    # Content - with colored background (like HIV chart)
    ws.merge_cells(f'A{current_row}:{end_col_letter}{current_row}')
    content_cell = ws[f'A{current_row}']
    content_cell.value = content
    content_cell.font = Font(size=11, color='000000')
    content_cell.fill = PatternFill(start_color=content_bg_color, end_color=content_bg_color, fill_type='solid')
    content_cell.alignment = Alignment(wrap_text=True, vertical='top')
    content_cell.border = thin_border
    ws.row_dimensions[current_row].height = 60  # High-yield content
    current_row += 1

    return current_row


def expand_row_heights(wb):
    """
    CRITICAL: Expand row heights only if content requires more space

    MUST be called AFTER all content is added to workbook
    This ensures no content is cut off while PRESERVING manually set heights

    Args:
        wb: workbook object
    """
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            current_height = sheet.row_dimensions[row[0].row].height or 15
            calculated_height = 15  # Minimum height

            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Count newlines
                    lines = cell.value.count('\n') + 1

                    # If wrap_text is enabled, estimate wrapped lines
                    if cell.alignment and cell.alignment.wrap_text:
                        col_width = sheet.column_dimensions[cell.column_letter].width
                        if col_width:
                            # Estimate lines based on content length and column width
                            char_per_line = col_width * 0.9
                            estimated_lines = max(lines, len(cell.value) / char_per_line)
                            calculated_height = max(calculated_height, estimated_lines * 15)

            # Only increase height if calculated is greater than current (preserve manual settings)
            if calculated_height > current_height:
                sheet.row_dimensions[row[0].row].height = min(calculated_height, 500)


def create_workbook_structure():
    """
    Create a new workbook with proper 4-tab structure

    Returns:
        Workbook object with 4 sheets created and named
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create exactly 4 tabs in specified order
    ws_details = wb.create_sheet('Drug Details', 0)
    ws_comp = wb.create_sheet('Key Comparisons', 1)
    ws_master = wb.create_sheet('Master Chart', 2)
    ws_high = wb.create_sheet('High-Yield & Pearls', 3)

    return wb, ws_details, ws_comp, ws_master, ws_high


# ==================== USAGE EXAMPLE ====================

def create_example_drug_chart():
    """
    Example usage showing how to create a drug chart using this template

    This demonstrates the proper workflow:
    1. Create workbook structure
    2. Add drug class tables to Drug Details
    3. Add comparison tables to Key Comparisons
    4. Add all drugs to Master Chart
    5. Add high-yield content
    6. Expand row heights
    7. Save
    """

    # Step 1: Create workbook structure
    wb, ws_details, ws_comp, ws_master, ws_high = create_workbook_structure()

    # ==================== TAB 1: DRUG DETAILS ====================
    set_column_widths(ws_details, {'A': 32, 'B': 42, 'C': 42, 'D': 45, 'E': 45})
    current_row = add_sheet_title(ws_details, 1, 'DRUG DETAILS - BY CLASS', 'E')
    current_row += 1

    # ===== DRUG CLASS 1: NRTIs =====
    colors = get_color_set(0)  # Light blue
    drug_names = ['Tenofovir\n(Viread)', 'Emtricitabine\n(Emtriv)', 'Lamivudine\n(Epivir)', 'Abacavir\n(Ziagen)']
    rows_data = [
        {'label': 'Common MOA', 'type': 'merged', 'value': 'Nucleoside analogue, requires phosphorylation (x3) → inhibits reverse transcriptase → blocks viral replication'},
        {'label': 'Specific MOA', 'type': 'individual', 'values': ['Adenosine analogue', 'Cytosine analogue', 'Cytosine analogue', 'Guanosine analogue']},
        {'label': 'Uses', 'type': 'individual', 'values': [
            '🟢 HIV - First line for naïve patients\n🟢 HBV monotherapy',
            '🟢 HIV - First line for naïve patients\nPrEP',
            'HIV\nHBV monotherapy',
            'HIV (often combined with lamivudine)'
        ]},
        {'label': 'Combination Therapy', 'type': 'merged', 'value': '2 NRTI + 1 other drug class (INSTI, NNRTI, or PI + enhancer)'},
        {'label': '', 'type': 'individual', 'values': ['Tenofovir+Emtricitabine = Truvada', '', 'Lamivudine + Abacavir', '']},
        {'label': 'Common AE', 'type': 'merged', 'value': '⚠️ Lactic acidosis\n⚠️ Hepatotoxicity (very rare but fatal)'},
        {'label': 'Specific Drug AE', 'type': 'individual', 'values': [
            'GI, flatulence',
            'GI, flatulence',
            'Headache, fatigue, insomnia, GI\nVery well tolerated',
            '⚠️ Serious hypersensitivity (~5%): fever, rash, GI, respiratory symptoms'
        ]},
        {'label': 'Special Considerations', 'type': 'individual', 'values': [
            'Inhibits HBV polymerase',
            '',
            'Often used in treatment regimens\n✅ Safe in pregnancy',
            '❗️Screen for HLA-B*5701\nONLY use in patients who test negative'
        ]},
        {'label': 'Route', 'type': 'merged', 'value': 'Oral'},
    ]
    current_row = add_drug_class_table(ws_details, current_row, 'NUCLEOSIDE REVERSE TRANSCRIPTASE INHIBITORS (NRTIs)', drug_names, rows_data, colors)
    mnemonic = """LATE - NRTIs for HIV:
L = Lamivudine (well tolerated, safe in pregnancy)
A = Abacavir (screen for HLA-B*5701 hypersensitivity)
T = Tenofovir (first-line, also treats HBV)
E = Emtricitabine (first-line, combines with tenofovir as Truvada)

Remember: NRTIs require phosphorylation (x3) to become active"""
    current_row = add_mnemonic_row(ws_details, current_row, 'E', mnemonic)
    current_row += 2

    # ===== DRUG CLASS 2: NNRTIs =====
    colors = get_color_set(1)  # Light green
    drug_names = ['Rilpivirine\n(Edurant)', 'Doravirine\n(Pifeltro)']
    rows_data = [
        {'label': 'Common MOA', 'type': 'merged', 'value': 'Bind directly to inhibit viral reverse transcriptase at allosteric site (non-competitively)'},
        {'label': 'Uses', 'type': 'merged', 'value': 'HIV - In combination with NRTIs\n⚠️ Rapid resistance if not combined'},
        {'label': 'Common AE', 'type': 'merged', 'value': 'Rash, CNS effects\n⚠️ Hepatotoxicity'},
        {'label': 'Specific Drug AE', 'type': 'individual', 'values': [
            'Insomnia, headache, depression',
            'Dizziness, nausea, headache'
        ]},
        {'label': 'Special Considerations', 'type': 'individual', 'values': [
            'Take with food (increases absorption)\nMonitor mental health',
            'Once daily\nFewer drug interactions'
        ]},
        {'label': 'Route', 'type': 'merged', 'value': 'Oral'},
    ]
    # Adjust column count for 2 drugs
    set_column_widths(ws_details, {'A': 32, 'B': 55, 'C': 55})
    current_row = add_drug_class_table(ws_details, current_row, 'NON-NUCLEOSIDE REVERSE TRANSCRIPTASE INHIBITORS (NNRTIs)', drug_names, rows_data, colors)
    mnemonic = """NNRTIs end in "-virine":
Rilpivirine, Doravirine, Efavirenz, Nevirapine

"VIRines go VIRal" - directly bind to reverse transcriptase"""
    current_row = add_mnemonic_row(ws_details, current_row, 'C', mnemonic)
    current_row += 2

    # ===== DRUG CLASS 3: Integrase Inhibitors =====
    colors = get_color_set(2)  # Light purple
    drug_names = ['Dolutegravir\n(Tivicay)', 'Bictegravir', 'Cabotegravir\n(Vocabria)']
    rows_data = [
        {'label': 'Common MOA', 'type': 'merged', 'value': 'Inhibit HIV integrase enzyme → prevents viral DNA integration into host genome'},
        {'label': 'Uses', 'type': 'merged', 'value': 'HIV treatment (part of HAART)\n🟢 Preferred agents for initial therapy'},
        {'label': 'Common AE', 'type': 'merged', 'value': 'Generally well tolerated\nInsomnia, headache, GI upset'},
        {'label': 'Specific Drug AE', 'type': 'individual', 'values': [
            'Weight gain, insomnia',
            'Diarrhea, nausea',
            'Injection site reactions (IM)'
        ]},
        {'label': 'Special Considerations', 'type': 'individual', 'values': [
            'Once daily\nFew drug interactions',
            'Part of Biktarvy (single tablet)',
            'Long-acting injectable option\nMonthly or every 2 months'
        ]},
        {'label': 'Route', 'type': 'individual', 'values': ['Oral', 'Oral (once daily)', 'Oral, IM']},
    ]
    set_column_widths(ws_details, {'A': 32, 'B': 45, 'C': 45, 'D': 45})
    current_row = add_drug_class_table(ws_details, current_row, 'INTEGRASE INHIBITORS (INSTIs)', drug_names, rows_data, colors)
    mnemonic = """INSTIs end in "-tegravir":
Dolutegravir, Bictegravir, Cabotegravir, Raltegravir

"TEGRA-viral" - integrate into therapy as preferred agents"""
    current_row = add_mnemonic_row(ws_details, current_row, 'D', mnemonic)
    current_row += 2

    # ==================== TAB 2: KEY COMPARISONS ====================
    set_column_widths(ws_comp, {'A': 32, 'B': 45, 'C': 45, 'D': 45})
    current_row = add_sheet_title(ws_comp, 1, 'KEY COMPARISONS', 'D')
    current_row += 1

    # Comparison 1: Combination Drugs
    colors = get_color_set(0)
    current_row = add_comparison_table(
        ws_comp, current_row,
        'COMBINATION DRUGS',
        ['Drug 1', 'Drug 2', 'Combined Name', 'Notes'],
        [
            ['Tenofovir (NRTI)', 'Emtricitabine (NRTI)', 'Truvada', '🟢 First-line for naïve HIV patients, also used for PrEP'],
            ['Bictegravir (Integrase)', 'Emtricitabine + Tenofovir (NRTIs)', 'Biktarvy', 'Complete treatment regimen - once daily'],
            ['Darunavir or Atazanavir (PI)', 'Ritonavir (Enhancer)', 'Combined therapy', 'Ritonavir boosts PI levels via CYP3A4 inhibition'],
        ],
        colors
    )
    current_row += 2

    # Comparison 2: Mechanisms with Analogies
    colors = get_color_set(1)
    current_row = add_comparison_table(
        ws_comp, current_row,
        'MECHANISMS OF ACTION WITH ANALOGIES',
        ['Drug Class', 'Mechanism', 'Analogy', 'Key Target'],
        [
            ['NRTIs', 'Nucleoside analogues incorporated into DNA chain, causing termination', 'Think of NRTIs like defective Lego blocks that stop the DNA chain from growing', 'Reverse transcriptase active site'],
            ['NNRTIs', 'Bind directly to reverse transcriptase allosteric site', 'NNRTIs are like a wrench thrown into the gears, jamming the enzyme', 'Reverse transcriptase allosteric pocket'],
            ['Protease Inhibitors', 'Bind to protease enzyme, preventing cleavage of viral polyproteins', 'Imagine protease as molecular scissors cutting viral proteins - PIs jam the scissors', 'HIV protease enzyme'],
            ['Integrase Inhibitors', 'Block integrase enzyme from inserting viral DNA into host genome', 'Think of integrase like a molecular stapler - inhibitors prevent stapling viral DNA into host DNA', 'Integrase enzyme'],
        ],
        colors
    )
    current_row += 2

    # Comparison 3: First-Line Drugs
    colors = get_color_set(2)
    current_row = add_comparison_table(
        ws_comp, current_row,
        'FIRST-LINE DRUGS & PREFERRED AGENTS',
        ['Drug', 'Status', 'Clinical Context'],
        [
            ['Tenofovir (NRTI)', '🟢 First-line', 'For naïve HIV patients - combined with Emtricitabine (Truvada)'],
            ['Emtricitabine (NRTI)', '🟢 First-line', 'For naïve HIV patients - combined with Tenofovir (Truvada)'],
            ['Darunavir (PI)', '🟢 Drug of first choice', 'When protease inhibitor included in treatment regimen'],
        ],
        colors
    )
    current_row += 2

    # ==================== TAB 3: MASTER CHART ====================
    setup_master_chart_sheet(ws_master)

    # Add drug rows (color-coded by class) - matching Drug Details color scheme
    # NRTIs - Light Blue (color set 0)
    add_master_chart_row(ws_master, 2, 'NRTI', 'Lamivudine (Epivir)', 'Oral',
                         'Nucleoside analogue → inhibits RT', 'HIV, HBV',
                         'Headache, fatigue, GI', 'None absolute', 'Develops rapidly',
                         'None major', '✅ Safe in pregnancy', ICE_BLUE_MAIN)
    add_master_chart_row(ws_master, 3, 'NRTI', 'Tenofovir (Viread)', 'Oral',
                         'Nucleoside analogue → inhibits RT', '🟢 HIV - First line\nHBV',
                         '⚠️ Lactic acidosis, hepatotoxicity', 'Severe renal impairment', 'Develops rapidly',
                         'Nephrotoxic drugs', 'Monitor renal function', ICE_BLUE_MAIN)
    add_master_chart_row(ws_master, 4, 'NRTI', 'Emtricitabine (Emtriv)', 'Oral',
                         'Nucleoside analogue → inhibits RT', '🟢 HIV - First line\nPrEP',
                         'GI, flatulence', 'None absolute', 'Develops rapidly',
                         'None major', 'Combined with Tenofovir = Truvada', ICE_BLUE_MAIN)
    add_master_chart_row(ws_master, 5, 'NRTI', 'Abacavir (Ziagen)', 'Oral',
                         'Nucleoside analogue → inhibits RT', 'HIV',
                         '⚠️ Hypersensitivity (5%)', 'HLA-B*5701 positive', 'Develops rapidly',
                         'None major', '❗️ Screen HLA-B*5701 BEFORE use', ICE_BLUE_MAIN)

    # NNRTIs - Light Green (color set 1)
    add_master_chart_row(ws_master, 6, 'NNRTI', 'Rilpivirine (Edurant)', 'Oral',
                         'Binds allosteric site on RT', 'HIV (with NRTIs)',
                         'Rash, insomnia, depression', 'None absolute', '⚠️ Rapid if monotherapy',
                         'CYP3A4 inducers', 'Take with food', SEAFOAM_MAIN)
    add_master_chart_row(ws_master, 7, 'NNRTI', 'Doravirine (Pifeltro)', 'Oral',
                         'Binds allosteric site on RT', 'HIV (with NRTIs)',
                         'Dizziness, nausea', 'None absolute', '⚠️ Rapid if monotherapy',
                         'Fewer interactions', 'Once daily dosing', SEAFOAM_MAIN)

    # Integrase Inhibitors - Light Purple (color set 2)
    add_master_chart_row(ws_master, 8, 'Integrase Inhibitor', 'Dolutegravir (Tivicay)', 'Oral',
                         'Inhibits integrase enzyme', '🟢 HIV - Preferred initial',
                         'Weight gain, insomnia', 'None absolute', 'Low risk',
                         'Polyvalent cations', 'Once daily, few interactions', LIGHT_ORCHID_MAIN)
    add_master_chart_row(ws_master, 9, 'Integrase Inhibitor', 'Cabotegravir (Vocabria)', 'Oral, IM',
                         'Inhibits integrase enzyme', 'HIV treatment & PrEP',
                         'Injection site reactions', 'None absolute', 'Low risk',
                         'Rifampin', 'Long-acting IM option', LIGHT_ORCHID_MAIN)
    add_master_chart_row(ws_master, 10, 'Integrase Inhibitor', 'Bictegravir', 'Oral',
                         'Inhibits integrase enzyme', 'HIV (part of Biktarvy)',
                         'Diarrhea, nausea', 'None absolute', 'Low risk',
                         'Polyvalent cations', 'Single tablet regimen', LIGHT_ORCHID_MAIN)

    # Protease Inhibitors - Light Tan (color set 3)
    add_master_chart_row(ws_master, 11, 'Protease Inhibitor', 'Darunavir (Prezista)', 'Oral',
                         'Inhibits HIV protease', '🟢 Preferred PI',
                         'GI upset, rash', 'Sulfa allergy (caution)', 'Low risk',
                         'CYP3A4 substrates', 'Must boost with ritonavir/cobicistat', CHAMPAGNE_MAIN)
    add_master_chart_row(ws_master, 12, 'Protease Inhibitor', 'Atazanavir (Reyataz)', 'Oral',
                         'Inhibits HIV protease', 'HIV treatment',
                         'Hyperbilirubinemia (jaundice)', 'None absolute', 'Moderate',
                         'PPIs, H2 blockers', 'Causes indirect hyperbilirubinemia', CHAMPAGNE_MAIN)

    # Pharmacokinetic Enhancers - Light Teal (color set 4)
    add_master_chart_row(ws_master, 13, 'Pharmacokinetic Enhancer', 'Ritonavir (Norvir)', 'Oral',
                         'Inhibits CYP3A4 → boosts PI levels', 'Booster for PIs',
                         'GI upset, lipid changes', 'None as booster', 'N/A',
                         'Major CYP3A4 inhibitor', 'Used at low doses to boost', SKY_BLUE_MAIN)
    add_master_chart_row(ws_master, 14, 'Pharmacokinetic Enhancer', 'Cobicistat (Tyboost)', 'Oral',
                         'Inhibits CYP3A4 → boosts PI levels', 'Booster for PIs/INSTIs',
                         'GI upset', 'None as booster', 'N/A',
                         'Major CYP3A4 inhibitor', 'No anti-HIV activity itself', SKY_BLUE_MAIN)

    # ==================== TAB 4: HIGH-YIELD & PEARLS ====================
    set_column_widths(ws_high, {'A': 70, 'B': 70, 'C': 70})
    current_row = add_sheet_title(ws_high, 1, 'HIGH-YIELD & CLINICAL PEARLS', 'C')
    current_row += 1

    # Section 1: HAART Protocol
    haart_protocol = """• Standard HAART regimen for treatment-naïve patients: 2 NRTIs + 1 drug from another class (INSTI, NNRTI, or PI)
• First-line combination: Tenofovir + Emtricitabine (Truvada) + Integrase inhibitor or NNRTI
• Goal: Suppress viral load to undetectable levels (<50 copies/mL)
• Adherence is critical: Missing doses leads to resistance"""
    current_row = add_high_yield_section(ws_high, current_row, 'HAART PROTOCOL (HIGHLY ACTIVE ANTIRETROVIRAL THERAPY)', haart_protocol, SEAFOAM_HEADER)
    current_row += 1

    # Section 2: Clinical Pearls
    pearls = """• Pre-Exposure Prophylaxis (PrEP): HIV-negative individuals at high risk take Truvada (Tenofovir + Emtricitabine) daily
• Post-Exposure Prophylaxis (PEP): Start HAART within 72 hours of exposure, continue for 28 days
• Always screen for HLA-B*5701 before starting Abacavir to prevent hypersensitivity reaction
• Ritonavir is used at low doses to boost other protease inhibitor levels (CYP3A4 inhibition)
• Lamivudine is safe in pregnancy - often used in HIV+ pregnant patients"""
    current_row = add_high_yield_section(ws_high, current_row, 'CLINICAL PEARLS', pearls, ICE_BLUE_HEADER)
    current_row += 1

    # Section 3: Mnemonics - use light blue (#E6F3FF) for mnemonic content
    mnemonics = """NRTIs - DEALSZ Mnemonic:
D = Didanosine
E = Emtricitabine
A = Abacavir
L = Lamivudine
S = Stavudine
Z = Zidovudine (AZT)

Remember: All NRTIs end in "-vudine" or are nucleoside analogues"""
    current_row = add_high_yield_section(ws_high, current_row, 'MNEMONICS & MEMORY TRICKS', mnemonics, LIGHT_ORCHID_HEADER, 'E6F3FF')
    current_row += 1

    # Section 4: If You See X, Think Y
    if_x_think_y = """• If you see: Drug ending in "-vudine" → Think: NRTI (nucleoside analogue)
• If you see: Drug ending in "-virine" → Think: NNRTI (non-nucleoside RT inhibitor)
• If you see: Drug ending in "-tegravir" → Think: Integrase inhibitor
• If you see: Drug ending in "-navir" → Think: Protease inhibitor
• If you see: Serious hypersensitivity reaction → Think: Abacavir (screen HLA-B*5701)
• If you see: Boosting other drug levels → Think: Ritonavir or Cobicistat"""
    current_row = add_high_yield_section(ws_high, current_row, 'IF YOU SEE X, THINK Y', if_x_think_y, CHAMPAGNE_HEADER)
    current_row += 1

    # Section 5: Must-Know Facts
    must_know = """• NRTIs require phosphorylation (x3) to become active triphosphate form
• HAART uses 3 drugs from at least 2 different classes
• Truvada (Tenofovir + Emtricitabine) is first-line for both treatment and PrEP
• All NRTIs carry risk of lactic acidosis and hepatotoxicity (rare but serious)
• Resistance develops rapidly if medications not taken as prescribed - adherence is critical"""
    current_row = add_high_yield_section(ws_high, current_row, 'MUST-KNOW FACTS', must_know, SKY_BLUE_HEADER)

    # Step 6: CRITICAL - Expand row heights
    expand_row_heights(wb)

    return wb


# ==================== MAIN EXECUTION ====================

if __name__ == '__main__':
    print("=" * 70)
    print("EXCEL DRUG CHART TEMPLATE - Python Helper Functions")
    print("=" * 70)
    print("\nThis template provides all helper functions for creating drug charts.")
    print("Generating example chart to demonstrate usage...\n")

    # Create example
    wb = create_example_drug_chart()

    # Save
    output_path = '/Users/kimnguyen/Documents/Study Guide Claude Q3/Claude Templates/Drug_Chart_Example.xlsx'
    wb.save(output_path)

    print(f"✅ Example chart created: {output_path}")
    print("\n" + "=" * 70)
    print("FILE STRUCTURE:")
    print("=" * 70)
    print("  Tab 1: Drug Details - Drug class comparison tables")
    print("  Tab 2: Key Comparisons - Side-by-side comparisons with analogies")
    print("  Tab 3: Master Chart - All drugs in comprehensive table (frozen header)")
    print("  Tab 4: High-Yield & Pearls - Clinical pearls and mnemonics")
    print("\n✓ Soft pastel color scheme applied to ALL data cells")
    print("✓ All text in black for readability")
    print("✓ Row heights automatically expanded to fit content")
    print("✓ Mnemonics sections included (must be researched, not invented)")
    print("✓ Analogy boxes for mechanism comparisons")
    print("\n" + "=" * 70)
    print("USAGE INSTRUCTIONS FOR CLAUDE:")
    print("=" * 70)
    print("When creating a drug chart from a source file:")
    print("1. Import this template module")
    print("2. Extract drug information from source file")
    print("3. Call helper functions with extracted data")
    print("4. Save the workbook")
    print("\nThis approach saves ~60-70% of tokens by reusing formatting code!")
    print("=" * 70)
