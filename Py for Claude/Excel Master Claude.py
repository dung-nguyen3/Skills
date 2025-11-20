#!/usr/bin/env python3
"""
EXCEL MASTER CHART - CLAUDE GENERATOR TEMPLATE
===============================================

This template is for CLAUDE to create complete ready-to-run scripts.

INSTRUCTIONS FOR CLAUDE:
1. Copy this entire template
2. Fill in the SOURCE_INFO section
3. Fill in the HEADERS list (column names)
4. Fill in the DATA array with all extracted data from source
5. Update OUTPUT_FILENAME
6. Save as: [LectureNumber]_[Topic]_Master_Chart_Generator.py
7. User runs: python3 script.py (Excel opens automatically)

VERIFICATION CHECKLIST (Claude must complete BEFORE creating):
☑ Source file: [exact path]
☑ Instruction template: Excel Master Chart Only.txt
☑ Source-only policy: ONLY use information from source file
☑ MANDATORY: WebSearch used for mnemonics (not data extraction)
☑ Save location: [Class]/[Exam]/Claude Study Tools/

Features:
- ✅ Auto-detects drug class/category changes
- ✅ Auto-assigns colors (10-color rotation)
- ✅ All data hard-coded (ready to run)
- ✅ Opens Excel automatically
"""

import argparse
import sys
import subprocess
import platform
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# SOURCE INFORMATION (Claude fills this in)
# =============================================================================

SOURCE_FILE = "path/to/source/file.txt"  # Claude: Update with actual source file path
LECTURE_TOPIC = "Topic Name"  # Claude: Update with topic (e.g., "HIV Drugs")
CREATION_DATE = "2025-01-19"  # Claude: Update with creation date

# =============================================================================
# COLOR SCHEME (from Excel Master Chart Only.txt)
# =============================================================================

MAIN_TITLE_COLOR = '4472C4'  # Dark blue for header

# 10-color rotation for drug classes/groups
COLOR_SETS = [
    'D9E2F3',  # 0: Ice Blue
    'C8E6C9',  # 1: Seafoam
    'D1C4E9',  # 2: Light Orchid
    'F7E7CE',  # 3: Champagne
    'BDD7EE',  # 4: Sky Blue
    'F0F8FF',  # 5: Pale Azure
    'FCE4EC',  # 6: Blush Pink
    'EDE7F6',  # 7: Soft Lilac
    'FFE8D6',  # 8: Soft Tangerine
    'BBDEFB',  # 9: Powder Blue
]

# White borders (invisible against pastel backgrounds)
thin_border = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='thin', color='FFFFFF'),
    bottom=Side(style='thin', color='FFFFFF')
)

# =============================================================================
# HEADERS - Claude defines based on chart type
# =============================================================================

# For Drug Charts (11 columns):
HEADERS = [
    'Drug Class',
    'Drug Name (Brand)',
    'Route',
    'Mechanism',
    'Uses',
    'Adverse Effects',
    'Contraindications',
    'Resistance',
    'Drug Interactions',
    'Drug Combinations',
    'Special Considerations'
]

# For Condition Charts (7 columns), Claude would use:
# HEADERS = [
#     'Condition',
#     'Epidemiology',
#     'Risk Factors',
#     'Clinical Presentation',
#     'Diagnostics',
#     'Labs',
#     'Treatment'
# ]

# For Lab Value Charts (5 columns), Claude would use:
# HEADERS = [
#     'Test Name',
#     'Normal Range',
#     'Increased In',
#     'Decreased In',
#     'Clinical Significance'
# ]

# =============================================================================
# DATA - Claude fills with extracted data from source
# =============================================================================

# IMPORTANT: Group by drug class! All drugs in same class must be together.
# Each row must have exactly the same number of values as HEADERS.

DATA = [
    # Example format - Claude replaces with real data:
    # ['NRTI', 'Tenofovir (Viread)', 'Oral', 'Adenosine analogue → inhibits RT', '🟢 HIV - First line\nPrEP', 'GI, flatulence\n⚠️ Lactic acidosis', 'Severe renal impairment', 'Rapid if alone', 'None major', 'Truvada (with emtricitabine)', 'Monitor renal function'],
    # ['NRTI', 'Lamivudine (Epivir)', 'Oral', 'Cytosine analogue → inhibits RT', 'HIV treatment\nHBV', 'Headache, fatigue', 'None absolute', 'Rapid if alone', 'None major', 'With abacavir', '✅ Safe in pregnancy'],
    # ['NNRTI', 'Rilpivirine (Edurant)', 'Oral', 'Binds RT allosteric site', 'HIV (with NRTIs)', 'Depression, headache', 'Hepatitis co-infection', 'Rapid if alone', 'None major', 'Cabenuva', 'Must use with NRTIs'],
]

# Output filename - Claude updates based on topic
OUTPUT_FILENAME = 'Master_Chart.xlsx'  # e.g., 'HIV_Drugs_Master_Chart.xlsx'

# =============================================================================
# WORKBOOK CREATION
# =============================================================================

def get_column_width(column_name):
    """Auto-determine column width based on content type."""
    if column_name.lower() in ['route', 'status', 'type']:
        return 12
    if any(term in column_name.lower() for term in ['class', 'category', 'condition']):
        return 22
    if 'name' in column_name.lower():
        return 28
    return 35

def create_master_chart(headers):
    """Create Master Chart workbook with specified column headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Chart"

    # Set column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = get_column_width(header)
        ws.column_dimensions[col_letter].width = width

    # Create header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color=MAIN_TITLE_COLOR,
                               end_color=MAIN_TITLE_COLOR,
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center',
                                  vertical='center',
                                  wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = 'A2'

    return wb, ws

def add_data_row(ws, row_num, data_values, color):
    """Add a single data row to the Master Chart."""
    for col_idx, value in enumerate(data_values, start=1):
        cell = ws.cell(row_num, col_idx, value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border
        cell.font = Font(size=10, color='000000')

        # First column is bold (grouping/class column)
        if col_idx == 1:
            cell.font = Font(bold=True, size=10, color='000000')

        # Apply background color
        cell.fill = PatternFill(start_color=color,
                               end_color=color,
                               fill_type='solid')

# =============================================================================
# ⭐ AUTO-COLOR ASSIGNMENT
# =============================================================================

def add_data_with_auto_colors(ws, data):
    """
    Add all data rows with AUTOMATIC color assignment.

    How it works:
    1. Tracks the drug class/category (first column) for each row
    2. When class changes, assigns next color in rotation
    3. All items in same class get same color

    Example:
        Row 1: NRTI → Assigns Ice Blue (color 0)
        Row 2: NRTI → Keeps Ice Blue (same class)
        Row 3: NRTI → Keeps Ice Blue (same class)
        Row 4: NNRTI → Assigns Seafoam (color 1, class changed!)
        Row 5: NNRTI → Keeps Seafoam (same class)
    """
    if not data:
        print("⚠️  No data provided!")
        return 0

    previous_class = None
    current_color = None
    color_index = 0
    row_num = 2  # Start after header

    for row_data in data:
        if not row_data:  # Skip empty rows
            continue

        drug_class = row_data[0]  # First column = Drug Class/Category

        # Check if class changed
        if drug_class != previous_class:
            # New class detected - assign next color
            current_color = COLOR_SETS[color_index % len(COLOR_SETS)]
            print(f"  ✓ New group: '{drug_class}' → Color {color_index} (#{current_color})")
            color_index += 1
            previous_class = drug_class

        # Add row with current color
        add_data_row(ws, row_num, row_data, current_color)
        row_num += 1

    return row_num - 2  # Return total rows added

# =============================================================================
# VALIDATION
# =============================================================================

def validate_data(headers, data):
    """Validate that data matches header count."""
    if not data:
        print("❌ ERROR: No data provided!")
        return False

    expected_cols = len(headers)
    issues = []

    for i, row in enumerate(data, start=1):
        if not row:  # Skip empty rows
            continue
        actual_cols = len(row)
        if actual_cols != expected_cols:
            issues.append(f"  Row {i}: Has {actual_cols} columns, expected {expected_cols}")

    if issues:
        print("\n❌ DATA VALIDATION ERRORS:")
        print(f"Expected {expected_cols} columns based on headers.")
        for issue in issues:
            print(issue)
        print("\nFix: Ensure each data row has exactly", expected_cols, "values.\n")
        return False

    return True

# =============================================================================
# MAIN FUNCTION
# =============================================================================

def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Generate Excel Master Chart with AUTO-COLORING',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Generated from source: {SOURCE_FILE}
Topic: {LECTURE_TOPIC}
Created: {CREATION_DATE}

Simply run: python3 {sys.argv[0]}
Excel will open automatically with perfectly formatted chart.
        """
    )

    parser.add_argument('--output', help='Output Excel filename', default=OUTPUT_FILENAME)
    args = parser.parse_args()

    print("=" * 70)
    print(f"EXCEL MASTER CHART GENERATOR - {LECTURE_TOPIC}")
    print("=" * 70)
    print()
    print(f"Source: {SOURCE_FILE}")
    print()

    # Validate data
    if not validate_data(HEADERS, DATA):
        return 1

    # Show configuration
    print(f"📊 Configuration:")
    print(f"   Columns: {len(HEADERS)}")
    print(f"   Data rows: {len([d for d in DATA if d])}")
    print(f"   Output: {args.output}")
    print()

    # Create workbook
    print("🔨 Creating Master Chart...")
    wb, ws = create_master_chart(HEADERS)
    print(f"   ✓ Workbook created with {len(HEADERS)} columns")
    print(f"   ✓ Header row formatted (dark blue, white text)")
    print(f"   ✓ Frozen panes at A2")
    print()

    # Add data with auto-coloring
    print("🎨 Adding data with AUTO-COLOR assignment...")
    rows_added = add_data_with_auto_colors(ws, DATA)
    print(f"   ✓ {rows_added} data rows added")
    print()

    # Save
    print(f"💾 Saving to: {args.output}")
    wb.save(args.output)
    print()
    print("=" * 70)
    print("✅ SUCCESS!")
    print("=" * 70)

    # Summary
    print()
    print("📋 Summary:")
    print(f"   • Total rows: {rows_added}")
    print(f"   • Drug classes/categories: {len(set(row[0] for row in DATA if row))}")
    print(f"   • Auto-colored: YES ✅")
    print(f"   • File: {args.output}")
    print()

    # Auto-open the Excel file
    print(f"📂 Opening {args.output}...")
    try:
        if platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', args.output], check=True)
        elif platform.system() == 'Windows':
            subprocess.run(['start', args.output], shell=True, check=True)
        else:  # Linux
            subprocess.run(['xdg-open', args.output], check=True)
        print(f"   ✓ Opened in Excel")
    except Exception as e:
        print(f"   ⚠️  Could not auto-open file: {e}")
        print(f"   → Manually open: {args.output}")

    print()
    return 0

if __name__ == '__main__':
    sys.exit(main())
