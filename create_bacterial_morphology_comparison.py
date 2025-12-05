#!/usr/bin/env python3
"""
BACTERIAL MORPHOLOGY COMPARISON CHART
Created from: Bacterial Morphology microbiology_text.txt

CRITICAL COLOR IMPLEMENTATION:
- Tab 1: Each TABLE gets ONE solid color for ALL rows
- Tab 2: Items in same CATEGORY get same color
- NO alternating row colors within a single table/category
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# COLOR SCHEME
# =============================================================================

HEADER_BG = '4472C4'  # Dark blue
MNEMONIC_BG = 'E6F3FF'  # Light blue for mnemonics

# CRITICAL: Use ONE color per table (Tab 1) or category (Tab 2)
# Rotate to next color only when starting NEW table/category
ROW_COLORS = [
    'D9E2F3',  # Ice Blue - Table/Category 1
    'C8E6C9',  # Seafoam - Table/Category 2
    'D1C4E9',  # Light Orchid - Table/Category 3
    'F7E7CE',  # Champagne - Table/Category 4
    'BDD7EE',  # Sky Blue - Table/Category 5
    'F0F8FF',  # Pale Azure - Table/Category 6
    'FCE4EC',  # Blush Pink - Table/Category 7
    'EDE7F6',  # Soft Lilac - Table/Category 8
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def apply_cell_style(cell, text='', bold=False, font_size=11, bg_color=None,
                     border=True, alignment='left', wrap=True, font_color='000000'):
    """Apply comprehensive cell styling"""
    cell.value = text
    cell.font = Font(name='Calibri', size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(
        horizontal=alignment,
        vertical='top',
        wrap_text=wrap
    )

    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    if border:
        cell.border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

def create_comparison_header(ws, title, row, span_cols=5):
    """Create merged title row for a comparison table"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row, 1)
    apply_cell_style(cell, text=title, bold=True, font_size=14,
                    bg_color=HEADER_BG, alignment='center', font_color='FFFFFF')
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws.row_dimensions[row].height = 30

def create_column_headers(ws, headers, row, start_col=1):
    """Create column headers for comparison table"""
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row, col_idx)
        apply_cell_style(cell, text=header, bold=True, font_size=11,
                        bg_color='8CA4BE', alignment='center', font_color='FFFFFF')
    ws.row_dimensions[row].height = 25

def create_master_header_row(ws, headers, row=1):
    """Create formatted header row with freeze for Master Chart"""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx)
        apply_cell_style(
            cell, text=header, bold=True, font_size=12,
            bg_color=HEADER_BG, alignment='center', font_color='FFFFFF'
        )
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')

    ws.row_dimensions[row].height = 25
    ws.freeze_panes = f'A{row+1}'

def set_column_widths(ws, widths):
    """Set column widths from dictionary"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def add_mnemonic_row(ws, row, mnemonic_text, span_cols=5):
    """Add mnemonic row directly below a comparison table"""
    cell_a = ws.cell(row, 1)
    apply_cell_style(cell_a, text='MEMORY AID', bold=True, font_size=11,
                    bg_color=MNEMONIC_BG, alignment='left', font_color='0000FF')

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=span_cols)
    cell_content = ws.cell(row, 2)
    apply_cell_style(cell_content, text=mnemonic_text, bg_color=MNEMONIC_BG)
    ws.row_dimensions[row].height = 60

def add_section_header(ws, row, title):
    """Add section header in Summary tab"""
    cell = ws.cell(row, 1)
    apply_cell_style(cell, text=title, bold=True, font_size=14,
                    bg_color=HEADER_BG, alignment='center', font_color='FFFFFF')
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    ws.row_dimensions[row].height = 30

# =============================================================================
# TAB 1: KEY COMPARISONS
# =============================================================================

def create_key_comparisons_tab(wb):
    """
    Tab 1: Key Comparisons
    CRITICAL: Each table gets ONE solid color for ALL rows
    """
    ws = wb.active
    ws.title = "Key Comparisons"

    set_column_widths(ws, {
        'A': 25,
        'B': 35,
        'C': 35,
        'D': 35,
        'E': 35,
    })

    current_row = 1

    # =========================================================================
    # TABLE 1: Bacterial Shapes
    # =========================================================================

    create_comparison_header(ws, "BACTERIAL SHAPES", current_row, span_cols=5)
    current_row += 1

    headers = ['Shape Type', 'Name', 'Description', 'Visual', 'Examples']
    create_column_headers(ws, headers, current_row)
    current_row += 1

    shapes_data = [
        ('Coccus', 'Cocci (plural)', 'Spheres or round cells', 'O', 'Staphylococcus, Streptococcus'),
        ('Bacillus', 'Bacilli (plural)', 'Rod-shaped cells', 'Rectangle', 'E. coli, Bacillus'),
        ('Vibrio', 'Vibrios (plural)', 'Comma-shaped cells', 'Curved', 'Vibrio cholerae'),
        ('Spirillum', 'Spirilla (plural)', 'Rigid spiral rod', 'Spiral', 'Spirillum volutans'),
        ('Spirochete', 'Spirochetes (plural)', 'Flexible spiral rod', 'Corkscrew', 'Treponema, Borrelia'),
        ('Coccobacillus', 'Coccobacilli (plural)', 'Short rod or oval', 'Short rod', 'Haemophilus'),
        ('Pleomorphic', 'Variable', 'No defined shape', 'Variable', 'Mycoplasma'),
    ]

    # TABLE 1: ONE color (Ice Blue) for ALL rows
    table_1_color = ROW_COLORS[0]
    for row_data in shapes_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_1_color)
        current_row += 1

    current_row += 3

    # =========================================================================
    # TABLE 2: Bacterial Arrangements
    # =========================================================================

    create_comparison_header(ws, "BACTERIAL ARRANGEMENTS", current_row, span_cols=5)
    current_row += 1

    headers = ['Arrangement', 'Prefix', 'Description', 'Division Pattern', 'Examples']
    create_column_headers(ws, headers, current_row)
    current_row += 1

    arrangements_data = [
        ('Pairs', 'Diplo-', 'Two cells together', 'One plane, remain attached', 'Diplococci'),
        ('Chains', 'Strepto-', 'Cells in chains', 'One plane, cells remain in line', 'Streptococcus'),
        ('Clusters', 'Staphylo-', 'Grape-like clusters', 'Multiple planes, irregular clusters', 'Staphylococcus'),
        ('Packets of 4', 'Tetrads', 'Four cells in square', 'Two planes', 'Certain cocci'),
        ('Packets of 8', 'Sarcina', 'Eight cells in cube', 'Three planes', 'Sarcina'),
    ]

    # TABLE 2: ONE color (Seafoam) for ALL rows
    table_2_color = ROW_COLORS[1]
    for row_data in arrangements_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_2_color)
        current_row += 1

    # Mnemonic for arrangements
    current_row += 1
    add_mnemonic_row(ws, current_row,
        'MNEMONIC: Greek origins help!\n\nDiplo- = Double (two)\nStrepto- = String/chain (like beads)\nStaphylo- = Grapes (cluster)\n\n[Researched mnemonic from medical education resources]', span_cols=5)
    current_row += 3

    # =========================================================================
    # TABLE 3: Cell Wall Types
    # =========================================================================

    create_comparison_header(ws, "BACTERIAL CELL WALL TYPES", current_row, span_cols=5)
    current_row += 1

    headers = ['Type', 'Peptidoglycan', 'Key Components', 'Gram Stain', 'Examples']
    create_column_headers(ws, headers, current_row)
    current_row += 1

    cell_wall_data = [
        ('Gram-positive', 'Thick', 'Teichoic acid', 'Purple (retains crystal violet)', 'Staphylococcus, Streptococcus'),
        ('Gram-negative', 'Thin', 'Outer membrane with LPS (endotoxin)', 'Pink/red (safranin counterstain)', 'E. coli, Pseudomonas'),
        ('Acid-fast', 'Thin', 'Mycolic acid layer, tetrameric porins', 'Requires acid-fast stain (Ziehl-Neelsen)', 'Mycobacterium tuberculosis'),
    ]

    # TABLE 3: ONE color (Light Orchid) for ALL rows
    table_3_color = ROW_COLORS[2]
    for row_data in cell_wall_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_3_color)
        current_row += 1

    # Mnemonic for Gram stain
    current_row += 1
    add_mnemonic_row(ws, current_row,
        'MNEMONIC: "Keep your P\'s together"\n\nGram-Positive = Purple (both start with P)\nGram-negative = Pink/red\n\n[Researched mnemonic from Pearson Study Prep]', span_cols=5)
    current_row += 3

    # =========================================================================
    # TABLE 4: External Structures
    # =========================================================================

    create_comparison_header(ws, "BACTERIAL EXTERNAL STRUCTURES", current_row, span_cols=5)
    current_row += 1

    headers = ['Structure', 'Composition', 'Function', 'Clinical Significance', 'Examples']
    create_column_headers(ws, headers, current_row)
    current_row += 1

    external_data = [
        ('Capsule (Glycocalyx)', 'Polysaccharides, glycoproteins', 'Virulence factor - inhibits phagocytosis, protects from desiccation', 'Encapsulated bacteria more virulent', 'Streptococcus pneumoniae'),
        ('Slime layer (Glycocalyx)', 'Polysaccharides, glycoproteins', 'Biofilm formation, adherence', 'Dental plaque, catheter infections', 'Various bacteria'),
        ('Flagella', 'Protein filaments', 'Motility', 'Helps bacteria spread, reach host cells', 'E. coli, Vibrio'),
        ('Pili/Fimbriae (common)', 'Protein subunits (pilins)', 'Adherence to host cells (adhesins)', 'Tissue tropism, virulence factor', 'E. coli (attaches to intestinal epithelium)'),
        ('Sex pilus (F pilus)', 'Protein subunits (pilins)', 'Conjugation (genetic exchange)', 'Antibiotic resistance transfer', 'Gram-negative bacteria only'),
    ]

    # TABLE 4: ONE color (Champagne) for ALL rows
    table_4_color = ROW_COLORS[3]
    for row_data in external_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_4_color)
        current_row += 1

    # Mnemonic for encapsulated bacteria
    current_row += 1
    add_mnemonic_row(ws, current_row,
        'MNEMONIC for encapsulated bacteria: "SHiNE SKiS"\n\nS = Streptococcus pneumoniae\nH = Haemophilus influenzae\nN = Neisseria meningitidis\nE = E. coli\nS = Salmonella\nK = Klebsiella pneumoniae\nS = Group B Streptococcus\n\n[Researched mnemonic from USMLE Step 1 resources]', span_cols=5)
    current_row += 3

    # =========================================================================
    # TABLE 5: Virulence Factors
    # =========================================================================

    create_comparison_header(ws, "BACTERIAL VIRULENCE FACTORS", current_row, span_cols=5)
    current_row += 1

    headers = ['Virulence Factor', 'Type', 'Mechanism', 'Function', 'Examples']
    create_column_headers(ws, headers, current_row)
    current_row += 1

    virulence_data = [
        ('Adhesins', 'Surface proteins', 'Bind to receptors on host cells', 'Enable attachment, colonization, tissue tropism', 'Pili, fimbriae, glycocalyx'),
        ('Invasins', 'Enzymes', 'Degrade tissue matrices, intracellular spaces', 'Facilitate bacterial spread, tissue damage', 'Collagenase, neuraminidase, streptokinase'),
        ('Exotoxins', 'Secreted proteins', 'Highly specific enzyme-like activity', 'Damage host cells remotely', 'Diphtheria toxin, botulinum toxin, cholera toxin'),
        ('Endotoxins (LPS)', 'Lipopolysaccharide (Lipid A)', 'Immune system activation', 'Can cause toxic shock when released', 'Gram-negative outer membrane LPS'),
        ('Antiphagocytic factors', 'Capsule, proteins', 'Prevent phagocytosis', 'Evade immune system', 'Capsule, Protein A (S. aureus), Protein M (Streptococcus)'),
    ]

    # TABLE 5: ONE color (Sky Blue) for ALL rows
    table_5_color = ROW_COLORS[4]
    for row_data in virulence_data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=table_5_color)
        current_row += 1

    return ws

# =============================================================================
# TAB 2: MASTER CHART
# =============================================================================

def create_master_chart_tab(wb):
    """
    Tab 2: Master Chart
    CRITICAL: Items in same CATEGORY get same color
    """
    ws = wb.create_sheet("Master Chart")

    set_column_widths(ws, {
        'A': 25,
        'B': 20,
        'C': 35,
        'D': 35,
        'E': 30,
    })

    headers = ['Structure/Component', 'Category', 'Location', 'Function', 'Clinical Relevance']
    create_master_header_row(ws, headers, 1)

    # Master chart data organized by category
    master_data = [
        # CATEGORY 1: Cell Wall Components (Ice Blue)
        ('Peptidoglycan', 'Cell Wall', 'All bacteria (thick in G+, thin in G-)', 'Rigidity, shape, osmotic protection', 'Target for beta-lactams, lysozyme'),
        ('Teichoic acid', 'Cell Wall', 'Gram-positive only', 'Structural support', 'Antigenic'),
        ('Outer membrane', 'Cell Wall', 'Gram-negative only', 'Protective barrier', 'Antibiotic resistance'),
        ('LPS (endotoxin)', 'Cell Wall', 'Gram-negative outer membrane', 'Immune signaling', 'Toxic shock (Lipid A)'),
        ('Mycolic acid', 'Cell Wall', 'Acid-fast bacteria', 'Resistance to desiccation, antibiotics', 'Inhibits phagocytosis'),

        # CATEGORY 2: External Structures (Seafoam)
        ('Capsule', 'External', 'Outermost layer', 'Virulence factor - inhibits phagocytosis', 'Encapsulated bacteria more virulent'),
        ('Slime layer', 'External', 'Outermost layer', 'Biofilm formation, adherence', 'Dental plaque, catheter infections'),
        ('Flagella', 'External', 'Variable locations (monotrichous, peritrichous, etc.)', 'Motility', 'Helps spread, reach host cells'),
        ('Pili/Fimbriae', 'External', 'Surface appendages', 'Adherence (adhesins), conjugation (sex pilus)', 'Virulence, antibiotic resistance transfer'),

        # CATEGORY 3: Internal Structures (Light Orchid)
        ('Chromosome', 'Internal', 'Nucleoid region', 'Genetic information (single circular dsDNA)', 'Contains essential genes'),
        ('Plasmids', 'Internal', 'Cytoplasm', 'Extra genes (circular dsDNA)', 'Antibiotic resistance spread'),
        ('Ribosomes (70S)', 'Internal', 'Cytoplasm', 'Protein synthesis', 'Antibiotic target (different from 80S)'),
        ('Endospores', 'Internal', 'Produced by Bacillus, Clostridium', 'Dormant survival form', 'Extremely resistant (heat, chemicals)'),

        # CATEGORY 4: Cell Membrane (Champagne)
        ('Cell membrane', 'Membrane', 'Below cell wall', 'Selective permeability, proteins (gates, sensors)', 'Fluid mosaic model'),
    ]

    current_row = 2

    # Color assignment by category
    category_colors = {
        'Cell Wall': ROW_COLORS[0],  # Ice Blue
        'External': ROW_COLORS[1],   # Seafoam
        'Internal': ROW_COLORS[2],   # Light Orchid
        'Membrane': ROW_COLORS[3],   # Champagne
    }

    for row_data in master_data:
        category = row_data[1]
        color = category_colors.get(category, ROW_COLORS[0])

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(current_row, col_idx)
            bold = (col_idx == 1)
            apply_cell_style(cell, text=value, bold=bold, bg_color=color)

        ws.row_dimensions[current_row].height = 50
        current_row += 1

    return ws

# =============================================================================
# TAB 3: SUMMARY
# =============================================================================

def create_summary_tab(wb):
    """Tab 3: Summary with mnemonics and high-yield pearls"""
    ws = wb.create_sheet("Summary")

    set_column_widths(ws, {'A': 100})

    current_row = 1

    # =========================================================================
    # MNEMONICS
    # =========================================================================

    add_section_header(ws, current_row, "MNEMONICS")
    current_row += 1

    mnemonics = [
        ('Gram Stain Colors', '"Keep your P\'s together"\n\nGram-Positive = Purple (both start with P)\nGram-negative = Pink/red\n\n[Researched from Pearson Study Prep]'),
        ('Bacterial Arrangements', 'Greek origins:\n\nDiplo- = Double (two cells)\nStrepto- = String/chain (like beads)\nStaphylo- = Grapes (cluster)\n\n[Researched from medical education resources]'),
        ('Encapsulated Bacteria', '"SHiNE SKiS" (detected by Quellung reaction)\n\nS = Streptococcus pneumoniae\nH = Haemophilus influenzae\nN = Neisseria meningitidis\nE = E. coli\nS = Salmonella\nK = Klebsiella pneumoniae\nS = Group B Streptococcus\n\n[Researched from USMLE Step 1 resources]'),
        ('Virulence Factors', '"AIT" = Adhesins, Invasins, Toxins\n\nOr: "All Invaders are Toxic"\n\n[Memory aid for three main categories]'),
    ]

    for mnemonic_name, mnemonic_content in mnemonics:
        cell = ws.cell(current_row, 1)
        full_text = f'MNEMONIC: "{mnemonic_name}"\n\n{mnemonic_content}'
        apply_cell_style(cell, text=full_text, bg_color=MNEMONIC_BG)
        ws.row_dimensions[current_row].height = 100
        current_row += 1

    current_row += 2

    # =========================================================================
    # IF X THINK Y
    # =========================================================================

    add_section_header(ws, current_row, '"IF X THINK Y" ASSOCIATIONS')
    current_row += 1

    associations = [
        "If purple on Gram stain → Think Gram-positive (thick peptidoglycan)",
        "If pink/red on Gram stain → Think Gram-negative (thin peptidoglycan, outer membrane with LPS)",
        "If doesn't Gram stain well → Think acid-fast bacteria (Mycobacterium)",
        "If grape-like clusters → Think Staphylococcus (staphylo- = grapes)",
        "If chains → Think Streptococcus (strepto- = chains)",
        "If encapsulated bacteria → Think increased virulence (inhibits phagocytosis)",
        "If endotoxin/LPS → Think Gram-negative bacteria (outer membrane)",
        "If endospores → Think Bacillus or Clostridium (extremely resistant)",
        "If biofilm formation → Think slime layer glycocalyx",
        "If antibiotic resistance spread → Think plasmids and sex pili (conjugation)",
    ]

    for assoc in associations:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=assoc, bg_color='F3E5F5')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2

    # =========================================================================
    # KEY DEFINITIONS
    # =========================================================================

    add_section_header(ws, current_row, "KEY DEFINITIONS")
    current_row += 1

    definitions = [
        "Prokaryote: Without a nucleus, no membrane-bound organelles",
        "Peptidoglycan: Complex polymer (NAG-NAM backbone, tetrapeptide sidechains, pentapeptide crosslinks)",
        "LPS (Lipopolysaccharide): Endotoxin in Gram-negative outer membrane. Lipid A = toxic component",
        "Glycocalyx: Outermost bacterial layer (capsule or slime layer)",
        "Virulence factor: Property that enables microbe to establish on/within host and cause disease",
        "Adhesins: Proteins that bind bacteria to host cell receptors",
        "Invasins: Enzymes that degrade tissue matrices (collagenase, neuraminidase, streptokinase)",
        "Endospore: Dormant bacterial cell form (Bacillus, Clostridium) - extremely resistant",
        "Conjugation: Genetic exchange via sex pilus (F pilus)",
        "Acid-fast: Bacteria with mycolic acid layer (don't Gram stain well)",
    ]

    for defn in definitions:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=defn, bg_color='E8F5E9')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 2

    # =========================================================================
    # HIGH-YIELD PEARLS
    # =========================================================================

    add_section_header(ws, current_row, "HIGH-YIELD PEARLS")
    current_row += 1

    pearls = [
        "• Bacteria are 0.2-2.0 µm in diameter (some of smallest living organisms)",
        "• Gram stain is critical for initial ID and guides antibiotic therapy before final ID",
        "• Peptidoglycan is unique to bacteria - target for beta-lactams and lysozyme",
        "• LPS (endotoxin) signals immune invasion by Gram-negatives. Large amounts = toxic shock",
        "• Mycolic acid layer in acid-fast bacteria provides resistance to desiccation and antibiotics",
        "• Capsule is major virulence factor - makes bacteria 'slippery' to phagocytes",
        "• Endospores (Bacillus, Clostridium) resist heat, desiccation, chemicals, UV, boiling",
        "• Prokaryotic ribosomes are 70S (vs eukaryotic 80S) - antibiotics target 70S specifically",
        "• Sex pilus (F pilus) transfers plasmids → antibiotic resistance spread",
        "• Exotoxins are secreted proteins with high specificity (diphtheria, botulinum, cholera)",
        "• Endotoxin (LPS) is structural component - not toxic until released from cell",
        "• Flagella arrangements: monotrichous (1 polar), lophotrichous (multiple polar), peritrichous (all over)",
    ]

    for pearl in pearls:
        cell = ws.cell(current_row, 1)
        apply_cell_style(cell, text=pearl, bg_color='E0F2F1')
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    return ws

# =============================================================================
# MAIN FUNCTION
# =============================================================================

def create_comparison_chart(output_path):
    """Create complete 3-tab comparison chart"""
    wb = Workbook()

    create_key_comparisons_tab(wb)
    create_master_chart_tab(wb)
    create_summary_tab(wb)

    wb.save(output_path)
    print(f"Comparison chart created: {output_path}")

if __name__ == '__main__':
    output_path = '/Users/kimnguyen/Library/Mobile Documents/com~apple~CloudDocs/Documents/Midwestern/Quarter 3/Microbiology/Claude Study Tools/Bacterial_Morphology_Comparison_Chart.xlsx'
    create_comparison_chart(output_path)
